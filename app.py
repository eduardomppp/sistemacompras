from flask import Flask, Blueprint, session, request, flash, redirect, url_for, render_template, jsonify, send_file, send_from_directory, abort
from flask_sqlalchemy import SQLAlchemy
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
#from flask_wtf.csrf import CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import safe_join
import os
from dotenv import load_dotenv
import logging
from datetime import datetime, timedelta, timezone
import chardet
import re
import sqlite3
import uuid
from werkzeug.utils import secure_filename
from os.path import basename
import pdfkit
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import joinedload
from collections import Counter
import random
import shutil
from sqlalchemy import or_
from decimal import Decimal, InvalidOperation
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import make_response
from sqlalchemy import func as _sa_func
# Adicione no início do arquivo, após as importações
from functools import lru_cache
from time import time
import threading

# Cache simples com tempo de expiração
_cache = {}
_cache_timestamp = {}

_cache_lock = threading.Lock()

def get_cached_or_execute(key, func, ttl_seconds=300):
    """
    Cache thread-safe com tempo de expiração.
    Uso: dados = get_cached_or_execute('chave_unica', lambda: minha_funcao(), ttl_seconds=60)
    """
    current_time = time()
    with _cache_lock:
        if key in _cache and (current_time - _cache_timestamp.get(key, 0)) < ttl_seconds:
            return _cache[key]

    result = func()
    with _cache_lock:
        _cache[key] = result
        _cache_timestamp[key] = current_time
    return result

# Funções auxiliares para cache
def get_empresas_unicas_cached():
    """Retorna lista única de empresas"""
    return [e[0] for e in db.session.query(SolicitacoesCompra.empresa).distinct().all() if e[0]]

def get_usuarios_unicos_cached():
    """Retorna lista única de usuários"""
    return [u[0] for u in db.session.query(SolicitacoesCompra.usuario).distinct().all() if u[0]]

def get_aplicacoes_unicas_cached():
    """Retorna lista única de aplicações"""
    from sqlalchemy import or_
    
    aplicacoes = set()
    
    # Materiais
    for a in db.session.query(Materiais.Aplicacao).filter(
        Materiais.Aplicacao.isnot(None), Materiais.Aplicacao != ''
    ).distinct().all():
        if a[0]:
            aplicacoes.add(a[0].strip())
    
    # Solicitações
    for a in db.session.query(SolicitacoesCompra.aplicacao).filter(
        SolicitacoesCompra.aplicacao.isnot(None), SolicitacoesCompra.aplicacao != ''
    ).distinct().all():
        if a[0]:
            aplicacoes.add(a[0].strip())
    
    return sorted(aplicacoes)

# ── Filtros de listagem ──────────────────────────────────────────────────────
# Uma única função, chamada de qualquer rota, com cache de 5 min.
# Emite UMA query por coluna ao banco — sem N+1, sem recalculo a cada request.
_FILTRO_STATUS_APROVADO = 'Aprovado'

def get_filtros_listagem(
    apenas_aprovadas: bool = True,
    comprador: str | None = None,
    ttl: int = 300,
) -> dict:
    """
    Retorna {'empresas': [...], 'usuarios': [...], 'aplicacoes': [...]}
    com cache thread-safe.  Parâmetros formam a chave de cache.

    Proteção: o valor de `comprador` é normalizado antes de virar chave —
    caracteres de controle e pipe são removidos para evitar colisões de cache
    forçadas via parâmetro malicioso.
    """
    comprador_safe = re.sub(r'[|\x00-\x1f]', '', comprador or '').strip() or None
    cache_key = f"filtros_listagem\x00aprovadas={apenas_aprovadas}\x00comprador={comprador_safe}"
    return get_cached_or_execute(cache_key, lambda: _calcular_filtros(apenas_aprovadas, comprador_safe), ttl)

def _calcular_filtros(apenas_aprovadas: bool, comprador: str | None) -> dict:
    """Executa as três queries DISTINCT de uma vez e devolve o dict.
    Captura tanto erros de banco quanto falhas de memória/ambiente,
    garantindo que o sistema caia para um estado conhecido e auditável.
    """
    from sqlalchemy import distinct as _dist
    try:
        base_filter_emp  = [SolicitacoesCompra.empresa.isnot(None)]
        base_filter_usr  = [SolicitacoesCompra.usuario.isnot(None)]
        base_filter_aplic = [
            SolicitacoesCompra.aplicacao.isnot(None),
            SolicitacoesCompra.aplicacao != '',
        ]
        if apenas_aprovadas:
            base_filter_emp.append(SolicitacoesCompra.status_aprovacao == _FILTRO_STATUS_APROVADO)
            base_filter_usr.append(SolicitacoesCompra.status_aprovacao == _FILTRO_STATUS_APROVADO)
            base_filter_aplic.append(SolicitacoesCompra.status_aprovacao == _FILTRO_STATUS_APROVADO)
        if comprador:
            base_filter_emp.append(SolicitacoesCompra.comprador_atribuido == comprador)
            base_filter_usr.append(SolicitacoesCompra.comprador_atribuido == comprador)
            base_filter_aplic.append(SolicitacoesCompra.comprador_atribuido == comprador)

        empresas = sorted([
            r[0] for r in db.session.query(_dist(SolicitacoesCompra.empresa)).filter(*base_filter_emp).all()
            if r[0]
        ])
        usuarios = sorted([
            r[0] for r in db.session.query(_dist(SolicitacoesCompra.usuario)).filter(*base_filter_usr).all()
            if r[0]
        ])
        aplicacoes = sorted([
            r[0] for r in db.session.query(_dist(SolicitacoesCompra.aplicacao)).filter(*base_filter_aplic).all()
            if r[0]
        ])
        return {'empresas': empresas, 'usuarios': usuarios, 'aplicacoes': aplicacoes}

    except SQLAlchemyError as e:
        logging.error('_calcular_filtros – erro de banco: %s', type(e).__name__, exc_info=True)
        return {'empresas': [], 'usuarios': [], 'aplicacoes': []}
    except MemoryError:
        logging.critical('_calcular_filtros – MemoryError: resultado de filtros muito grande', exc_info=True)
        return {'empresas': [], 'usuarios': [], 'aplicacoes': []}
    except Exception as e:
        logging.error('_calcular_filtros – erro inesperado: %s', type(e).__name__, exc_info=True)
        return {'empresas': [], 'usuarios': [], 'aplicacoes': []}

def invalidar_cache_filtros():
    """Chame após qualquer escrita que altere empresas/usuarios/aplicações."""
    with _cache_lock:
        for k in list(_cache.keys()):
            if k.startswith('filtros_listagem'):
                _cache.pop(k, None)
                _cache_timestamp.pop(k, None)

# Carregar variáveis de ambiente
load_dotenv()

# Substitua a função get_local_time() existente (no início do arquivo)
def get_local_time():
    """Retorna o datetime atual no fuso horário de Brasília (UTC-3) SEM timezone info."""
    from datetime import datetime, timezone, timedelta
    BRASILIA = timezone(timedelta(hours=-3))
    return datetime.now(BRASILIA).replace(tzinfo=None)

# Inicializar aplicação Flask
app = Flask(__name__)
app.secret_key = os.getenv('FLASK_SECRET_KEY')


# Configurações de segurança da sessão
app.config.update(
    SESSION_COOKIE_SECURE=os.getenv('FLASK_ENV', 'production') != 'development',  # True em produção, False em desenvolvimento
    SESSION_COOKIE_HTTPONLY=True,   # Impede acesso via JavaScript
    SESSION_COOKIE_SAMESITE='Lax',  # Proteção contra CSRF
    PERMANENT_SESSION_LIFETIME=timedelta(minutes=30),  # Tempo de expiração
    SESSION_REFRESH_EACH_REQUEST=True,        # Renova o tempo de vida a cada requisição
    TEMPLATES_AUTO_RELOAD=True                # Recarrega templates em desenvolvimento

)

if not app.secret_key:
    raise ValueError("Defina FLASK_SECRET_KEY no arquivo .env")

# Configuração do banco de dados SQLite
DATABASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ComparasDB.db')

# ===== Correcao Coluna Marca teste_teste=====

def garantir_colunas_necessarias():
    """Garante que TODAS as colunas necessárias existem em todas as tabelas"""
    print("🔧 VERIFICANDO INTEGRIDADE DO BANCO DE DADOS...")
    print("=" * 50)
    
    # ============================================
    # Lista completa de verificações por tabela
    # ============================================
    verificacoes = [
        # Tabela SolicitacoesPreenchidas
        {
            'tabela': 'SolicitacoesPreenchidas',
            'coluna': 'marca',
            'sql': "ALTER TABLE SolicitacoesPreenchidas ADD COLUMN marca TEXT"
        },
        {
            'tabela': 'SolicitacoesPreenchidas',
            'coluna': 'observacoes',
            'sql': "ALTER TABLE SolicitacoesPreenchidas ADD COLUMN observacoes TEXT"
        },
        {
            'tabela': 'SolicitacoesPreenchidas',
            'coluna': 'pdf_path',
            'sql': "ALTER TABLE SolicitacoesPreenchidas ADD COLUMN pdf_path TEXT"
        },
        {
            'tabela': 'SolicitacoesPreenchidas',
            'coluna': 'status',
            'sql': "ALTER TABLE SolicitacoesPreenchidas ADD COLUMN status TEXT DEFAULT 'Rascunho'"
        },
        
        # Tabela SolicitacoesCompra
        {
            'tabela': 'SolicitacoesCompra',
            'coluna': 'status_aprovacao',
            'sql': "ALTER TABLE SolicitacoesCompra ADD COLUMN status_aprovacao TEXT"
        },
        {
            'tabela': 'SolicitacoesCompra',
            'coluna': 'observacoes_col',
            'sql': "ALTER TABLE SolicitacoesCompra ADD COLUMN observacoes_col TEXT"
        },
        {
            'tabela': 'SolicitacoesCompra',
            'coluna': 'aplicacao_geral',
            'sql': "ALTER TABLE SolicitacoesCompra ADD COLUMN aplicacao_geral TEXT"
        },
        {
            'tabela': 'SolicitacoesCompra',
            'coluna': 'comprador_atribuido',
            'sql': "ALTER TABLE SolicitacoesCompra ADD COLUMN comprador_atribuido TEXT"
        },
        {
            'tabela': 'SolicitacoesCompra',
            'coluna': 'data_envio_aprovacao',
            'sql': "ALTER TABLE SolicitacoesCompra ADD COLUMN data_envio_aprovacao DATETIME"
        },
        {
            'tabela': 'SolicitacoesCompra',
            'coluna': 'data_retorno_aprovacao',
            'sql': "ALTER TABLE SolicitacoesCompra ADD COLUMN data_retorno_aprovacao DATETIME"
        },
        {
            'tabela': 'SolicitacoesCompra',
            'coluna': 'prioridade',
            'sql': "ALTER TABLE SolicitacoesCompra ADD COLUMN prioridade TEXT DEFAULT 'Programado'"
        },
        
        # Tabela Materiais
        {
            'tabela': 'Materiais',
            'coluna': 'Ativo',
            'sql': "ALTER TABLE Materiais ADD COLUMN Ativo BOOLEAN DEFAULT 0"
        },
        {
            'tabela': 'Materiais',
            'coluna': 'FatorConsumo',
            'sql': "ALTER TABLE Materiais ADD COLUMN FatorConsumo REAL DEFAULT 0.0"
        },
        
        # Tabela PedidosCompra
        {
            'tabela': 'PedidosCompra',
            'coluna': 'observacoes',
            'sql': "ALTER TABLE PedidosCompra ADD COLUMN observacoes TEXT"
        },
        {
            'tabela': 'PedidosCompra',
            'coluna': 'comprovante_pagamento',
            'sql': "ALTER TABLE PedidosCompra ADD COLUMN comprovante_pagamento TEXT"
        },
        {
            'tabela': 'PedidosCompra',
            'coluna': 'valor_frete',
            'sql': "ALTER TABLE PedidosCompra ADD COLUMN valor_frete REAL DEFAULT 0.0"
        },
        {
            'tabela': 'PedidosCompra',
            'coluna': 'valor_liquido',
            'sql': "ALTER TABLE PedidosCompra ADD COLUMN valor_liquido REAL DEFAULT 0.0"
        },
        {
            'tabela': 'PedidosCompra',
            'coluna': 'forma_pagamento',
            'sql': "ALTER TABLE PedidosCompra ADD COLUMN forma_pagamento TEXT"
        }
    ]
    
    # ============================================
    # Executar verificações — UMA conexão para todas as iterações
    # ============================================
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        for item in verificacoes:
            try:
                cursor.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                    (item['tabela'],)
                )
                if not cursor.fetchone():
                    print(f"⚠️ Tabela '{item['tabela']}' não existe - será criada pelo SQLAlchemy")
                    continue

                cursor.execute(f"PRAGMA table_info({item['tabela']})")
                colunas = [col[1] for col in cursor.fetchall()]

                if item['coluna'] not in colunas:
                    print(f"➕ Adicionando coluna '{item['coluna']}' à tabela '{item['tabela']}'...")
                    try:
                        cursor.execute(item['sql'])
                        print(f"   ✅ Coluna '{item['coluna']}' adicionada com sucesso!")
                    except sqlite3.OperationalError as e:
                        if "duplicate column name" in str(e).lower():
                            print(f"   ✅ Coluna '{item['coluna']}' já existe (ignorando erro)")
                        else:
                            print(f"   ⚠️ Erro ao adicionar coluna: {e}")
                else:
                    print(f"✅ Coluna '{item['coluna']}' já existe em '{item['tabela']}'")

            except sqlite3.Error as e:
                print(f"⚠️ Erro ao verificar {item['tabela']}.{item['coluna']}: {e}")
            except Exception as e:
                print(f"⚠️ Erro inesperado: {e}")

        conn.commit()
        conn.close()

    except sqlite3.Error as e:
        print(f"⚠️ Erro ao abrir conexão para verificações: {e}")
    
    print("=" * 50)
    print("🔧 VERIFICAÇÃO CONCLUÍDA")

# ============================================
# (garantir_colunas_necessarias é chamada dentro do bloco __main__ após db.create_all)

app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DATABASE}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'connect_args': {'check_same_thread': False},
    'pool_pre_ping': True,
    'pool_size': 5,
    'pool_recycle': 300,
}
DB_PATH_FORNECEDORES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fornecedores.db')

# Configuração para upload de arquivos
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # Limite de 5MB para uploads
ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png', 'pdf'}

# Criar pasta de uploads se não existir
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename, allowed_extensions=None):
    """Verifica se o arquivo tem uma extensão permitida"""
    if not filename:
        return False
    if '.' not in filename:
        return False
    try:
        extension = filename.rsplit('.', 1)[1].lower()
        allowed = allowed_extensions or ALLOWED_EXTENSIONS
        return extension in allowed
    except IndexError:
        return False

# Configuração do logger
log_file = 'app_errors.log'
log_dir = os.path.dirname(log_file)
if log_dir and not os.path.exists(log_dir):
    os.makedirs(log_dir)

_log_level = logging.DEBUG if os.getenv('FLASK_ENV') == 'development' else logging.WARNING
logging.basicConfig(
    filename=log_file,
    level=_log_level,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    encoding='utf-8'
)

# Definir o caminho do arquivo de senhas
SENHAS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "senhas.txt")

def log_error(message):
    logging.error(message)

def _flash_erro_seguro(mensagem_usuario: str, excecao: Exception, logger=None) -> None:
    """
    Exibe mensagem genérica via flash (sem expor detalhes da exceção ao usuário)
    e registra o erro completo com stack trace no log interno.
    """
    flash(mensagem_usuario, 'error')
    if logger:
        logger.error(mensagem_usuario, exc_info=True)
    else:
        logging.error(mensagem_usuario, exc_info=True)

# Inicializar SQLAlchemy
db = SQLAlchemy(app)

# Inicializar extensões de segurança
# csrf = CSRFProtect(app)  — desativado
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri="memory://"
)

# ── Garantir colunas antes do primeiro request ──────────────────
_colunas_garantidas = False

@app.before_request
def _garantir_colunas_antes_de_tudo():
    global _colunas_garantidas
    if not _colunas_garantidas:
        garantir_colunas_necessarias()
        _colunas_garantidas = True

# Configurar PRAGMAs de performance no SQLite a cada nova conexão
from sqlalchemy import event as _sa_event
from sqlalchemy.engine import Engine as _Engine

@_sa_event.listens_for(_Engine, "connect")
def _set_sqlite_pragmas(dbapi_conn, _connection_record):
    cursor = dbapi_conn.cursor()
    cursor.execute("PRAGMA journal_mode=WAL")    # leitura simultânea sem travar escrita
    cursor.execute("PRAGMA cache_size=-32000")   # 32 MB de cache em RAM
    cursor.execute("PRAGMA synchronous=NORMAL")  # mais rápido, ainda seguro contra crash
    cursor.execute("PRAGMA temp_store=MEMORY")   # tabelas temporárias na RAM
    cursor.close()

# Modelo para Kanban
class Kanban(db.Model):
    __tablename__ = 'Kanban'
    id = db.Column(db.Integer, primary_key=True)
    task_name = db.Column(db.Text, nullable=False)
    status = db.Column(db.Text, nullable=False)

# Modelo para Materiais
class Materiais(db.Model):
    __tablename__ = 'Materiais'
    CodMaterial = db.Column(db.Integer, primary_key=True, autoincrement=True)
    DescricaoMaterial = db.Column(db.Text, nullable=False)
    Empresa = db.Column(db.Text, nullable=False)
    Aplicacao = db.Column(db.Text, nullable=False)
    QuantidadeEstoque = db.Column(db.Integer, default=0)
    Fornecedor = db.Column(db.Text)
    NumeroNF = db.Column(db.Text)
    FatorConsumo = db.Column(db.Float, nullable=True, default=0.0)
    Ativo = db.Column(db.Boolean, default=False)

    def to_dict(self):
        return {
            'CodMaterial': self.CodMaterial,
            'DescricaoMaterial': self.DescricaoMaterial,
            'Empresa': self.Empresa,
            'Aplicacao': self.Aplicacao,
            'QuantidadeEstoque': self.QuantidadeEstoque,
            'Fornecedor': self.Fornecedor,
            'NumeroNF': self.NumeroNF,
            'FatorConsumo': self.FatorConsumo,
            'Ativo': self.Ativo
        }

# Modelo para Solicitações de Compra
class SolicitacoesCompra(db.Model):
    __tablename__ = 'SolicitacoesCompra'
    id = db.Column(db.Integer, primary_key=True)
    cod_material = db.Column(db.Integer, db.ForeignKey('Materiais.CodMaterial'), nullable=False)
    especificacao = db.Column(db.Text, nullable=False)
    quantidade = db.Column(db.String(20), nullable=False)
    unidade_medida = db.Column(db.String(20), nullable=False, default='Unidade')
    aplicacao = db.Column(db.Text, nullable=True)  # Aplicação específica do item
    aplicacao_geral = db.Column(db.Text, nullable=True)  # NOVA COLUNA: Aplicação geral do formulário
    empresa = db.Column(db.Text, nullable=False)
    data_solicitacao = db.Column(db.DateTime, nullable=False, default=get_local_time)
    usuario = db.Column(db.Text, nullable=False)
    foto_path = db.Column(db.Text, nullable=True)
    marca = db.Column(db.Text, nullable=True)
    ativo = db.Column(db.String(3), nullable=False)
    nome_ativo = db.Column(db.Text, nullable=True)
    prioridade = db.Column(db.String(20), nullable=False, default='Programado')
    status_aprovacao = db.Column(db.String(20), nullable=True, default=None)
    observacoes_col = db.Column(db.Text, nullable=True)
    material = db.relationship('Materiais', backref='solicitacoes')
    comprador_atribuido = db.Column(db.Text, nullable=True)  # Novo campo
    data_envio_aprovacao = db.Column(db.DateTime, nullable=True)   # Data que o comprador levou p/ aprovação
    data_retorno_aprovacao = db.Column(db.DateTime, nullable=True)  # Data que retornou aprovado/reprovado

    def to_dict(self):
        return {
            'id': self.id,
            'cod_material': self.cod_material,
            'especificacao': self.especificacao,
            'quantidade': self.quantidade,
            'unidade_medida': self.unidade_medida,
            'aplicacao': self.aplicacao,
            'aplicacao_geral': self.aplicacao_geral,  # Adicione esta linha
            'empresa': self.empresa,
            'data_solicitacao': self.data_solicitacao.isoformat() if self.data_solicitacao else None,
            'usuario': self.usuario,
            'foto_path': self.foto_path,
            'marca': self.marca,
            'ativo': self.ativo,
            'nome_ativo': self.nome_ativo,
            'prioridade': self.prioridade,
            'status_aprovacao': self.status_aprovacao,
            'comprador_atribuido': self.comprador_atribuido
        }
    
# Modelo para Solicitações Preenchidas
class SolicitacoesPreenchidas(db.Model):
    __tablename__ = 'SolicitacoesPreenchidas'
    id = db.Column(db.Integer, primary_key=True)
    solicitacao_id = db.Column(db.Integer, db.ForeignKey('SolicitacoesCompra.id'), nullable=False)
    fornecedor_id = db.Column(db.Integer, nullable=False)
    valor_unitario = db.Column(db.Float, nullable=False)
    valor_total = db.Column(db.Float, nullable=False)
    valor_frete = db.Column(db.Float, nullable=True)
    prazo_entrega = db.Column(db.Text, nullable=False)
    condicao_pagamento = db.Column(db.Text, nullable=False)
    data_preenchimento = db.Column(db.DateTime, nullable=False, default=get_local_time)
    usuario = db.Column(db.Text, nullable=False)
    status = db.Column(db.Text, nullable=True, default='Rascunho')
    pdf_path = db.Column(db.Text, nullable=True)
    observacoes = db.Column(db.Text, nullable=True)  # Novo campo para observações
    marca = db.Column(db.Text, nullable=True)  # 🔴 NOVO CAMPO ADICIONADO AQUI

    
    solicitacao = db.relationship('SolicitacoesCompra', backref='preenchimentos_fornecidos')
    
    historico_descontos = db.relationship(
        "HistoricoDescontos", 
        backref="solicitacao_preenchida",
        lazy="dynamic",
        cascade="all, delete-orphan"
    )

    def to_dict(self):
        return {
            'id': self.id,
            'solicitacao_id': self.solicitacao_id,
            'fornecedor_id': self.fornecedor_id,
            'valor_unitario': self.valor_unitario,
            'valor_total': self.valor_total,
            'valor_frete': self.valor_frete,
            'prazo_entrega': self.prazo_entrega,
            'condicao_pagamento': self.condicao_pagamento,
            'data_preenchimento': self.data_preenchimento.isoformat() if self.data_preenchimento else None,
            'usuario': self.usuario,
            'status': self.status,
            'pdf_path': self.pdf_path,
            'observacoes': self.observacoes,
            'marca': self.marca,  # 🔴 ADICIONADO AQUI
            'historico_descontos': [h.to_dict() for h in self.historico_descontos]
        }
# add_marca_to_preenchidas() removida — era redundante com garantir_colunas_necessarias()

def executar_todas_migracoes():
    """Executa todas as migrações necessárias em ordem correta"""
    print("\n" + "=" * 60)
    print("🚀 INICIANDO MIGRAÇÕES DO BANCO DE DADOS")
    print("=" * 60)
    
    # ============================================
    # 1. GARANTIR QUE O BANCO DE DADOS EXISTE
    # ============================================
    print("\n📁 Verificando banco de dados...")
    if not os.path.exists(DATABASE):
        print("   ⚠️ Banco de dados não encontrado. Será criado automaticamente.")
    
    # ============================================
    # 2. MIGRAÇÃO CRÍTICA: COLUNA 'marca' EM SolicitacoesPreenchidas
    # ============================================
    print("\n🔧 VERIFICANDO COLUNA 'marca' EM SolicitacoesPreenchidas...")
    
    try:
        # Conectar ao banco de dados
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # Verificar se a tabela SolicitacoesPreenchidas existe
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='SolicitacoesPreenchidas'")
        tabela_existe = cursor.fetchone()
        
        if tabela_existe:
            # Verificar se a coluna 'marca' já existe
            cursor.execute("PRAGMA table_info(SolicitacoesPreenchidas)")
            colunas_existentes = [col[1] for col in cursor.fetchall()]
            
            if 'marca' not in colunas_existentes:
                print("   ➕ Adicionando coluna 'marca' à tabela SolicitacoesPreenchidas...")
                
                try:
                    # Método 1: Tentar ALTER TABLE diretamente
                    cursor.execute("ALTER TABLE SolicitacoesPreenchidas ADD COLUMN marca TEXT")
                    conn.commit()
                    print("   ✅ Coluna 'marca' adicionada com sucesso (ALTER TABLE)!")
                    
                    # Verificar se funcionou
                    cursor.execute("PRAGMA table_info(SolicitacoesPreenchidas)")
                    colunas_apos = [col[1] for col in cursor.fetchall()]
                    if 'marca' in colunas_apos:
                        print("   ✅ Confirmação: Coluna 'marca' existe agora!")
                    else:
                        print("   ⚠️ Não foi possível confirmar a existência da coluna")
                        
                except sqlite3.OperationalError as e:
                    if "duplicate column name" in str(e).lower():
                        print("   ℹ️ Coluna 'marca' já existe (ignorando erro)")
                    else:
                        print(f"   ❌ Erro ao adicionar coluna via ALTER TABLE: {e}")
                        
                        # Método 2: Recriar a tabela se necessário (caso mais complexo)
                        print("   🔄 Tentando método alternativo: recriar tabela...")
                        try:
                            conn.execute("BEGIN")
                            try:
                                cursor2 = conn.cursor()
                                # Obter estrutura atual
                                cursor2.execute("PRAGMA table_info(SolicitacoesPreenchidas)")
                                colunas_atuais = cursor2.fetchall()
                                nomes_colunas = [col[1] for col in colunas_atuais]

                                if 'marca' not in nomes_colunas:
                                    # Criar tabela temporária com nova coluna
                                    colunas_sql = ', '.join([f'"{col[1]}" {col[2]}' for col in colunas_atuais])
                                    cursor2.execute(
                                        f"CREATE TABLE SolicitacoesPreenchidas_temp ({colunas_sql}, marca TEXT)"
                                    )

                                    # Copiar dados — DROP só após INSERT bem-sucedido
                                    colunas_copy = ', '.join([f'"{col}"' for col in nomes_colunas])
                                    cursor2.execute(
                                        f"INSERT INTO SolicitacoesPreenchidas_temp ({colunas_copy}) "
                                        f"SELECT {colunas_copy} FROM SolicitacoesPreenchidas"
                                    )
                                    cursor2.execute("DROP TABLE SolicitacoesPreenchidas")
                                    cursor2.execute(
                                        "ALTER TABLE SolicitacoesPreenchidas_temp "
                                        "RENAME TO SolicitacoesPreenchidas"
                                    )

                                conn.execute("COMMIT")
                                print("   ✅ Coluna 'marca' adicionada com sucesso (método alternativo)!")
                            except Exception as e2:
                                conn.execute("ROLLBACK")
                                print(f"   ❌ Rollback executado. Nenhum dado foi perdido. Erro: {e2}")
                        except Exception as e2:
                            print(f"   ❌ Método alternativo também falhou: {e2}")
            else:
                print("   ✅ Coluna 'marca' já existe na tabela SolicitacoesPreenchidas")
        else:
            print("   ℹ️ Tabela 'SolicitacoesPreenchidas' ainda não existe - será criada pelo SQLAlchemy")
        
        conn.close()
        
    except sqlite3.Error as e:
        logging.error(f"Erro ao adicionar coluna marca: {str(e)}")
        print(f"   ❌ Erro: {str(e)}")
    
    # ============================================
    # 3. MIGRAÇÃO: COLUNA 'observacoes' EM SolicitacoesPreenchidas
    # ============================================
    print("\n🔧 Verificando coluna 'observacoes' em SolicitacoesPreenchidas...")
    
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute("PRAGMA table_info(SolicitacoesPreenchidas)")
        colunas_existentes = [col[1] for col in cursor.fetchall()]
        
        if 'observacoes' not in colunas_existentes:
            cursor.execute("ALTER TABLE SolicitacoesPreenchidas ADD COLUMN observacoes TEXT")
            conn.commit()
            print("   ✅ Coluna 'observacoes' adicionada com sucesso!")
        else:
            print("   ✅ Coluna 'observacoes' já existe")
        
        conn.close()
    except Exception as e:
        print(f"   ⚠️ Erro ao verificar coluna observacoes: {e}")
    
    # ============================================
    # 4. MIGRAÇÃO: COLUNA 'pdf_path' EM SolicitacoesPreenchidas
    # ============================================
    print("\n🔧 Verificando coluna 'pdf_path' em SolicitacoesPreenchidas...")
    
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute("PRAGMA table_info(SolicitacoesPreenchidas)")
        colunas_existentes = [col[1] for col in cursor.fetchall()]
        
        if 'pdf_path' not in colunas_existentes:
            cursor.execute("ALTER TABLE SolicitacoesPreenchidas ADD COLUMN pdf_path TEXT")
            conn.commit()
            print("   ✅ Coluna 'pdf_path' adicionada com sucesso!")
        else:
            print("   ✅ Coluna 'pdf_path' já existe")
        
        conn.close()
    except Exception as e:
        print(f"   ⚠️ Erro ao verificar coluna pdf_path: {e}")
    
    # ============================================
    # 5. MIGRAÇÃO: COLUNA 'status' EM SolicitacoesPreenchidas
    # ============================================
    print("\n🔧 Verificando coluna 'status' em SolicitacoesPreenchidas...")
    
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute("PRAGMA table_info(SolicitacoesPreenchidas)")
        colunas_existentes = [col[1] for col in cursor.fetchall()]
        
        if 'status' not in colunas_existentes:
            cursor.execute("ALTER TABLE SolicitacoesPreenchidas ADD COLUMN status TEXT DEFAULT 'Rascunho'")
            conn.commit()
            print("   ✅ Coluna 'status' adicionada com sucesso!")
        else:
            print("   ✅ Coluna 'status' já existe")
        
        conn.close()
    except Exception as e:
        print(f"   ⚠️ Erro ao verificar coluna status: {e}")
    
    # ============================================
    # 6. MIGRAÇÃO: TABELA SolicitacoesCompra
    # ============================================
    print("\n🔧 Verificando colunas da tabela SolicitacoesCompra...")
    
    colunas_solicitacoes = [
        ('status_aprovacao', "ALTER TABLE SolicitacoesCompra ADD COLUMN status_aprovacao TEXT"),
        ('observacoes_col', "ALTER TABLE SolicitacoesCompra ADD COLUMN observacoes_col TEXT"),
        ('aplicacao_geral', "ALTER TABLE SolicitacoesCompra ADD COLUMN aplicacao_geral TEXT"),
        ('comprador_atribuido', "ALTER TABLE SolicitacoesCompra ADD COLUMN comprador_atribuido TEXT"),
        ('prioridade', "ALTER TABLE SolicitacoesCompra ADD COLUMN prioridade TEXT DEFAULT 'Programado'")
    ]
    
    for coluna, sql in colunas_solicitacoes:
        try:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            cursor.execute("PRAGMA table_info(SolicitacoesCompra)")
            colunas_existentes = [col[1] for col in cursor.fetchall()]
            
            if coluna not in colunas_existentes:
                cursor.execute(sql)
                conn.commit()
                print(f"   ✅ Coluna '{coluna}' adicionada com sucesso!")
            else:
                print(f"   ✅ Coluna '{coluna}' já existe")
            
            conn.close()
        except Exception as e:
            print(f"   ⚠️ Erro ao verificar coluna {coluna}: {e}")
    
    # ============================================
    # 7. MIGRAÇÃO: TABELA Materiais
    # ============================================
    print("\n🔧 Verificando colunas da tabela Materiais...")
    
    colunas_materiais = [
        ('Ativo', "ALTER TABLE Materiais ADD COLUMN Ativo BOOLEAN DEFAULT 0"),
        ('FatorConsumo', "ALTER TABLE Materiais ADD COLUMN FatorConsumo REAL DEFAULT 0.0")
    ]
    
    for coluna, sql in colunas_materiais:
        try:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            cursor.execute("PRAGMA table_info(Materiais)")
            colunas_existentes = [col[1] for col in cursor.fetchall()]
            
            if coluna not in colunas_existentes:
                cursor.execute(sql)
                conn.commit()
                print(f"   ✅ Coluna '{coluna}' adicionada com sucesso!")
            else:
                print(f"   ✅ Coluna '{coluna}' já existe")
            
            conn.close()
        except Exception as e:
            print(f"   ⚠️ Erro ao verificar coluna {coluna}: {e}")
    
    # ============================================
    # 8. MIGRAÇÃO: TABELA PedidosCompra
    # ============================================
    print("\n🔧 Verificando colunas da tabela PedidosCompra...")
    
    colunas_pedidos = [
        ('observacoes', "ALTER TABLE PedidosCompra ADD COLUMN observacoes TEXT"),
        ('comprovante_pagamento', "ALTER TABLE PedidosCompra ADD COLUMN comprovante_pagamento TEXT"),
        ('valor_frete', "ALTER TABLE PedidosCompra ADD COLUMN valor_frete REAL DEFAULT 0.0"),
        ('valor_liquido', "ALTER TABLE PedidosCompra ADD COLUMN valor_liquido REAL DEFAULT 0.0"),
        ('forma_pagamento', "ALTER TABLE PedidosCompra ADD COLUMN forma_pagamento TEXT")
    ]
    
    for coluna, sql in colunas_pedidos:
        try:
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            
            cursor.execute("PRAGMA table_info(PedidosCompra)")
            colunas_existentes = [col[1] for col in cursor.fetchall()]
            
            if coluna not in colunas_existentes:
                cursor.execute(sql)
                conn.commit()
                print(f"   ✅ Coluna '{coluna}' adicionada com sucesso!")
            else:
                print(f"   ✅ Coluna '{coluna}' já existe")
            
            conn.close()
        except Exception as e:
            print(f"   ⚠️ Erro ao verificar coluna {coluna}: {e}")
    
    # ============================================
    # 9. MIGRAÇÃO: QUANTIDADE PARA TEXT (transação atômica)
    # ============================================
    print("\n🔧 Verificando tipo da coluna 'quantidade' em SolicitacoesCompra...")

    def _migracao_quantidade_para_text(conn):
        """Executa a migração dentro de uma transação já aberta pelo chamador."""
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(SolicitacoesCompra)")
        colunas_info = cursor.fetchall()
        quantidade_col = next((col for col in colunas_info if col[1] == 'quantidade'), None)

        if quantidade_col and quantidade_col[2].upper() in ('INTEGER', 'REAL', 'FLOAT', 'NUMERIC'):
            print("   🔄 Migrando coluna quantidade para TEXT...")

            # Remover tabela temporária residual, se existir
            cursor.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='SolicitacoesCompra_temp'"
            )
            if cursor.fetchone():
                cursor.execute("DROP TABLE SolicitacoesCompra_temp")

            # Criar tabela temporária com coluna TEXT e incluindo colunas de data de aprovação
            cursor.execute("""
                CREATE TABLE SolicitacoesCompra_temp (
                    id                     INTEGER PRIMARY KEY AUTOINCREMENT,
                    cod_material           INTEGER  NOT NULL,
                    especificacao          TEXT     NOT NULL,
                    quantidade             TEXT     NOT NULL,
                    unidade_medida         TEXT     NOT NULL DEFAULT 'Unidade',
                    aplicacao              TEXT,
                    aplicacao_geral        TEXT,
                    empresa                TEXT     NOT NULL,
                    data_solicitacao       DATETIME NOT NULL,
                    usuario                TEXT     NOT NULL,
                    foto_path              TEXT,
                    marca                  TEXT,
                    ativo                  TEXT     NOT NULL,
                    nome_ativo             TEXT,
                    prioridade             TEXT     NOT NULL DEFAULT 'Programado',
                    status_aprovacao       TEXT,
                    observacoes_col        TEXT,
                    comprador_atribuido    TEXT,
                    data_envio_aprovacao   DATETIME,
                    data_retorno_aprovacao DATETIME
                )
            """)

            # Copiar dados — CAST garante que NULLs ou não-numéricos virem "0"
            cursor.execute("""
                INSERT INTO SolicitacoesCompra_temp (
                    id, cod_material, especificacao, quantidade, unidade_medida,
                    aplicacao, aplicacao_geral, empresa, data_solicitacao, usuario,
                    foto_path, marca, ativo, nome_ativo, prioridade,
                    status_aprovacao, observacoes_col, comprador_atribuido,
                    data_envio_aprovacao, data_retorno_aprovacao
                )
                SELECT
                    id, cod_material, especificacao,
                    CAST(COALESCE(quantidade, 0) AS TEXT),
                    unidade_medida,
                    aplicacao, aplicacao_geral, empresa, data_solicitacao, usuario,
                    foto_path, marca, ativo, nome_ativo,
                    COALESCE(prioridade, 'Programado'),
                    status_aprovacao, observacoes_col, comprador_atribuido,
                    data_envio_aprovacao, data_retorno_aprovacao
                FROM SolicitacoesCompra
            """)

            # DROP só após INSERT bem-sucedido
            cursor.execute("DROP TABLE SolicitacoesCompra")
            cursor.execute("ALTER TABLE SolicitacoesCompra_temp RENAME TO SolicitacoesCompra")
            print("   ✅ Coluna 'quantidade' migrada para TEXT com sucesso!")
        else:
            print("   ✅ Coluna 'quantidade' já está no tipo TEXT — nada a fazer.")

    try:
        conn = sqlite3.connect(DATABASE)
        conn.isolation_level = None   # controle manual de transação
        conn.execute("BEGIN")
        try:
            _migracao_quantidade_para_text(conn)
            conn.execute("COMMIT")
        except Exception as e:
            conn.execute("ROLLBACK")
            print(f"   ❌ Rollback executado. Nenhum dado foi perdido. Erro: {e}")
            raise
        finally:
            conn.close()
    except sqlite3.Error as e:
        print(f"   ⚠️ Erro de banco ao migrar quantidade: {e}")
    except Exception as e:
        print(f"   ⚠️ Erro inesperado ao migrar quantidade: {e}")
    
    # ============================================
    # 10. VERIFICAÇÃO FINAL
    # ============================================
    print("\n" + "=" * 60)
    print("🔍 VERIFICAÇÃO FINAL")
    print("=" * 60)
    
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # Verificar todas as colunas importantes
        tabelas_para_verificar = {
            'SolicitacoesPreenchidas': ['marca', 'observacoes', 'pdf_path', 'status'],
            'SolicitacoesCompra': ['status_aprovacao', 'observacoes_col', 'aplicacao_geral', 'comprador_atribuido', 'prioridade'],
            'Materiais': ['Ativo', 'FatorConsumo'],
            'PedidosCompra': ['observacoes', 'comprovante_pagamento', 'valor_frete', 'valor_liquido', 'forma_pagamento']
        }
        
        for tabela, colunas in tabelas_para_verificar.items():
            cursor.execute(f"PRAGMA table_info({tabela})")
            colunas_existentes = [col[1] for col in cursor.fetchall()]
            
            for coluna in colunas:
                if coluna in colunas_existentes:
                    print(f"   ✅ {tabela}.{coluna} - OK")
                else:
                    print(f"   ❌ {tabela}.{coluna} - NÃO ENCONTRADA!")
        
        conn.close()
        
    except Exception as e:
        print(f"   ⚠️ Erro na verificação final: {e}")
    
    print("\n" + "=" * 60)
    print("✅ TODAS AS MIGRAÇÕES FORAM CONCLUÍDAS!")
    print("=" * 60)

def migrate_observacoes_col():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # Verificar se a coluna já existe
        cursor.execute("PRAGMA table_info(SolicitacoesCompra)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'observacoes_col' not in columns:
            cursor.execute("ALTER TABLE SolicitacoesCompra ADD COLUMN observacoes_col TEXT")
            conn.commit()
            logging.info("Coluna observacoes_col adicionada à tabela SolicitacoesCompra")
            print("✓ Coluna observacoes_col adicionada com sucesso!")
        else:
            print("✓ Coluna observacoes_col já existe na tabela")
        
        conn.close()
        return True
    except sqlite3.Error as e:
        logging.error(f"Erro na migração observacoes_col: {str(e)}")
        print(f"✗ Erro na migração: {str(e)}")
        return False
    
# Modelo para Estoque
class Estoque(db.Model):
    __tablename__ = 'Estoque'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    preenchimento_id = db.Column(db.Integer, db.ForeignKey('SolicitacoesPreenchidas.id'), nullable=False)
    cod_material = db.Column(db.Integer, db.ForeignKey('Materiais.CodMaterial'), nullable=False)
    quantidade = db.Column(db.Integer, nullable=False)
    fornecedor = db.Column(db.Text, nullable=False)
    numero_nf = db.Column(db.Text, nullable=False)
    data_entrada = db.Column(db.DateTime, nullable=False, default=get_local_time)
    usuario = db.Column(db.Text, nullable=False)
    preenchimento = db.relationship('SolicitacoesPreenchidas', backref='estoque')
    material = db.relationship('Materiais', backref='estoque')

# Modelo para Auditoria
class Auditoria(db.Model):
    __tablename__ = 'Auditoria'
    id = db.Column(db.Integer, primary_key=True)
    solicitacao_id = db.Column(db.Integer, db.ForeignKey('SolicitacoesCompra.id'), nullable=False)
    data_validacao = db.Column(db.DateTime, nullable=False, default=get_local_time)
    colaborador_1 = db.Column(db.Text, nullable=False)
    colaborador_2 = db.Column(db.Text, nullable=True)
    status = db.Column(db.Text, nullable=False)  # 'Conforme' ou 'Não Conforme'
    observacao = db.Column(db.Text, nullable=True)
    solicitacao = db.relationship('SolicitacoesCompra', backref='auditorias')

# Modelo para Requisições
class Requisicoes(db.Model):
    __tablename__ = 'Requisicoes'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    preenchimento_id = db.Column(db.Integer, db.ForeignKey('SolicitacoesPreenchidas.id'), nullable=True)  # Alterado para nullable=True
    cod_material = db.Column(db.Integer, db.ForeignKey('Materiais.CodMaterial'), nullable=False)
    quantidade = db.Column(db.Integer, nullable=False)
    ticket = db.Column(db.Text, nullable=False)
    data_requisicao = db.Column(db.DateTime, nullable=False, default=get_local_time)
    usuario = db.Column(db.Text, nullable=False)
    preenchimento = db.relationship('SolicitacoesPreenchidas', backref='requisicoes')
    material = db.relationship('Materiais', backref='requisicoes')

# Modelo para Pedidos de Compra
class PedidosCompra(db.Model):
    __tablename__ = 'PedidosCompra'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    numero_pedido = db.Column(db.Text, nullable=False, unique=True)
    data_criacao = db.Column(db.DateTime, nullable=False, default=get_local_time)
    usuario = db.Column(db.Text, nullable=False)
    status = db.Column(db.Text, nullable=False, default='Gerado')
    pdf_path = db.Column(db.Text, nullable=True)
    valor_total = db.Column(db.Float, nullable=False)
    valor_frete = db.Column(db.Float, nullable=True, default=0.0)
    valor_liquido = db.Column(db.Float, nullable=False)
    forma_pagamento = db.Column(db.Text, nullable=True)
    # ADICIONE ESTA LINHA
    observacoes = db.Column(db.Text, nullable=True)
    comprovante_pagamento = db.Column(db.Text, nullable=True)
    preenchimentos = db.relationship('SolicitacoesPreenchidas', secondary='pedido_preenchimento_associacao')
# Tabela de associação para relacionar PedidosCompra e SolicitacoesPreenchidas
pedido_preenchimento_associacao = db.Table('pedido_preenchimento_associacao',
    db.Column('pedido_id', db.Integer, db.ForeignKey('PedidosCompra.id'), primary_key=True),
    db.Column('preenchimento_id', db.Integer, db.ForeignKey('SolicitacoesPreenchidas.id'), primary_key=True)
)

class HistoricoDescontos(db.Model):
    __tablename__ = 'HistoricoDescontos'
    id = db.Column(db.Integer, primary_key=True)
    preenchimento_id = db.Column(db.Integer, db.ForeignKey('SolicitacoesPreenchidas.id'), nullable=False)
    valor_unitario_anterior = db.Column(db.Float, nullable=False)
    valor_unitario_novo = db.Column(db.Float, nullable=False)
    valor_frete_anterior = db.Column(db.Float, nullable=True)
    valor_frete_novo = db.Column(db.Float, nullable=True)
    data_alteracao = db.Column(db.DateTime, nullable=False, default=get_local_time)
    usuario = db.Column(db.Text, nullable=False)

    def to_dict(self):
        return {
            'id': self.id,
            'preenchimento_id': self.preenchimento_id,
            'valor_unitario_anterior': self.valor_unitario_anterior,
            'valor_unitario_novo': self.valor_unitario_novo,
            'valor_frete_anterior': self.valor_frete_anterior,
            'valor_frete_novo': self.valor_frete_novo,
            'data_alteracao': self.data_alteracao.isoformat() if self.data_alteracao else None,
            'usuario': self.usuario
        }



def add_comprovante_pagamento_column():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(PedidosCompra)")
        columns = [col[1] for col in cursor.fetchall()]
        if 'comprovante_pagamento' not in columns:
            cursor.execute("ALTER TABLE PedidosCompra ADD COLUMN comprovante_pagamento TEXT")
            conn.commit()
            logging.info("Added comprovante_pagamento column to PedidosCompra")
        conn.close()
    except sqlite3.Error as e:
        logging.error(f"Error adding comprovante_pagamento column: {str(e)}")

# Funções de Banco de Dados
def init_db():
    with app.app_context():
        db.create_all()
        create_historico_descontos_table()

def create_database():
    try:
        conn = sqlite3.connect(DATABASE)
        conn.close()
        print(f"Database '{DATABASE}' criada ou já existe.")
        return True
    except sqlite3.Error as e:
        print(f"Erro ao criar a database: {str(e)}")
        return False

def create_materiais_table():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Materiais';")
        if cursor.fetchone():
            print("Tabela 'Materiais' já existe.")
            cursor.close()
            conn.close()
            return True
        
        create_table_query = """
        CREATE TABLE Materiais (
            CodMaterial INTEGER PRIMARY KEY AUTOINCREMENT,
            DescricaoMaterial TEXT NOT NULL,
            Empresa TEXT NOT NULL,
            Aplicacao TEXT NOT NULL,
            QuantidadeEstoque INTEGER DEFAULT 0,
            Fornecedor TEXT,
            NumeroNF TEXT,
            FatorConsumo REAL DEFAULT 0.0
        )
        """
        cursor.execute(create_table_query)
        conn.commit()
        print("Tabela 'Materiais' criada com sucesso.")
        cursor.close()
        conn.close()
        return True
    except sqlite3.Error as e:
        print(f"Erro ao criar a tabela Materiais: {str(e)}")
        log_error(f"Erro ao criar a tabela Materiais: {str(e)}")
        return False

def create_kanban_table():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        create_table_query = """
        CREATE TABLE IF NOT EXISTS Kanban (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_name TEXT NOT NULL,
            status TEXT NOT NULL
        )
        """
        cursor.execute(create_table_query)
        conn.commit()
        print("Tabela 'Kanban' criada com sucesso.")
        cursor.close()
        conn.close()
    except sqlite3.Error as e:
        print(f"Erro ao criar a tabela Kanban: {str(e)}")
        return False
    return True

def create_solicitacoes_compra_table():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        create_table_query = """
        CREATE TABLE IF NOT EXISTS SolicitacoesCompra (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cod_material INTEGER NOT NULL,
            especificacao TEXT NOT NULL,
            quantidade INTEGER NOT NULL,
            unidade_medida TEXT NOT NULL DEFAULT 'Unidade',
            aplicacao TEXT,
            empresa TEXT NOT NULL,
            data_solicitacao DATETIME NOT NULL,
            usuario TEXT NOT NULL,
            foto_path TEXT,
            marca TEXT,
            ativo TEXT NOT NULL,
            nome_ativo TEXT,
            prioridade TEXT NOT NULL DEFAULT 'Programado',  -- Novo campo
            FOREIGN KEY (cod_material) REFERENCES Materiais(CodMaterial)
        )
        """
        cursor.execute(create_table_query)
        conn.commit()
        print("Tabela 'SolicitacoesCompra' criada com sucesso.")
        cursor.close()
        conn.close()
    except sqlite3.Error as e:
        print(f"Erro ao criar a tabela SolicitacoesCompra: {str(e)}")
        return False
    return True

def create_solicitacoes_preenchidas_table():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='SolicitacoesPreenchidas';")
        if cursor.fetchone():
            # Verificar se as colunas necessárias existem
            cursor.execute("PRAGMA table_info(SolicitacoesPreenchidas)")
            columns = [col[1] for col in cursor.fetchall()]
            
            # Migrações para colunas existentes
            if 'fornecedor_id' not in columns:
                cursor.execute("ALTER TABLE SolicitacoesPreenchidas ADD COLUMN fornecedor_id INTEGER NOT NULL DEFAULT 0")
                logging.info("Added fornecedor_id column to SolicitacoesPreenchidas")
            
            if 'valor_frete' not in columns:
                cursor.execute("ALTER TABLE SolicitacoesPreenchidas ADD COLUMN valor_frete REAL")
                logging.info("Added valor_frete column to SolicitacoesPreenchidas")
            
            # Não precisamos adicionar marca/ativo aqui pois são propriedades derivadas
            cursor.close()
            conn.close()
            return True
        
        # Criação da tabela se não existir
        create_table_query = """
        CREATE TABLE SolicitacoesPreenchidas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            solicitacao_id INTEGER NOT NULL,
            fornecedor_id INTEGER NOT NULL,
            valor_unitario REAL NOT NULL,
            valor_total REAL NOT NULL,
            valor_frete REAL,
            prazo_entrega TEXT NOT NULL,
            condicao_pagamento TEXT NOT NULL,
            data_preenchimento DATETIME NOT NULL,
            usuario TEXT NOT NULL,
            status TEXT DEFAULT 'Aguardando Aprovacao',
            pdf_path TEXT,
            FOREIGN KEY (solicitacao_id) REFERENCES SolicitacoesCompra(id)
        )
        """
        cursor.execute(create_table_query)
        conn.commit()
        print("Tabela 'SolicitacoesPreenchidas' criada com sucesso.")
        cursor.close()
        conn.close()
    except sqlite3.Error as e:
        print(f"Erro ao criar a tabela SolicitacoesPreenchidas: {str(e)}")
        return False
    return True

def create_estoque_table():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        create_table_query = """
        CREATE TABLE IF NOT EXISTS Estoque (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            preenchimento_id INTEGER NOT NULL,
            cod_material INTEGER NOT NULL,
            quantidade INTEGER NOT NULL,
            fornecedor TEXT NOT NULL,
            numero_nf TEXT NOT NULL,
            data_entrada DATETIME NOT NULL,
            usuario TEXT NOT NULL,
            FOREIGN KEY (preenchimento_id) REFERENCES SolicitacoesPreenchidas(id),
            FOREIGN KEY (cod_material) REFERENCES Materiais(CodMaterial)
        )
        """
        cursor.execute(create_table_query)
        conn.commit()
        print("Tabela 'Estoque' criada com sucesso.")
        cursor.close()
        conn.close()
    except sqlite3.Error as e:
        print(f"Erro ao criar a tabela Estoque: {str(e)}")
        return False
    return True

def get_rastreabilidade_entries(solicitacao):
    """Gera entradas de rastreabilidade com hora local correta"""
    entries = []
    
    # As datas já estão em horário de Brasília (sem timezone)
    # Entrada de criação
    data_criacao = solicitacao.data_solicitacao
    
    entries.append({
        'data': data_criacao,
        'evento': 'Solicitação Criada',
        'descricao': f'Solicitação registrada no sistema por {solicitacao.usuario}'
    })
    
    # Entrada de aprovação/reprovação (se aplicável)
    if solicitacao.status_aprovacao:
        if solicitacao.status_aprovacao == 'Aprovado':
            entries.append({
                'data': data_criacao,  # Usar data de criação como fallback
                'evento': 'Aprovação',
                'descricao': f'Solicitação aprovada'
            })
        elif solicitacao.status_aprovacao == 'Reprovado':
            # Tentar obter data da auditoria
            auditoria = Auditoria.query.filter_by(solicitacao_id=solicitacao.id).first()
            data_reprovacao = auditoria.data_validacao if auditoria else data_criacao
            
            # Verificar observações para ver se foi reprovação individual
            motivo = ''
            if solicitacao.observacoes_col and 'REPROVADA INDIVIDUALMENTE' in solicitacao.observacoes_col:
                motivo = solicitacao.observacoes_col
            
            entries.append({
                'data': data_reprovacao,
                'evento': 'Reprovação',
                'descricao': f'Solicitação reprovada' + (f'. Motivo: {motivo}' if motivo else '')
            })
    
    # Entrada de cotação (se existir preenchimento)
    if solicitacao.preenchimentos_fornecidos:
        primeiro_preenchimento = solicitacao.preenchimentos_fornecidos[0]
        data_preenchimento = primeiro_preenchimento.data_preenchimento
        
        # Obter nome do fornecedor para a descrição
        fornecedor_nome = get_fornecedor_nome(primeiro_preenchimento.fornecedor_id)
        
        entries.append({
            'data': data_preenchimento,
            'evento': 'Cotação Enviada',
            'descricao': f'Cotação enviada para o fornecedor {fornecedor_nome}'
        })
        
        # Entrada de resposta do fornecedor (mesma data do preenchimento)
        entries.append({
            'data': data_preenchimento,
            'evento': 'Resposta do Fornecedor',
            'descricao': f'Fornecedor {fornecedor_nome} respondeu com os valores informados'
        })
    
    # Ordenar por data
    entries.sort(key=lambda x: x['data'] if x['data'] else datetime.min)
    return entries


@app.template_filter('format_brasil_time')
def format_brasil_time(dt):
    """Filtro Jinja2 para formatar datetimes no formato brasileiro"""
    if dt is None:
        return ""
    
    # Se é string, tentar converter para datetime
    if isinstance(dt, str):
        try:
            # Tentar diferentes formatos
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M:%S.%f']:
                try:
                    dt = datetime.strptime(dt, fmt)
                    break
                except ValueError:
                    continue
        except:
            return str(dt)
    
    # Se já é um objeto datetime, formatar diretamente
    if hasattr(dt, 'strftime'):
        return dt.strftime('%d/%m/%Y %H:%M')
    
    return str(dt)

def create_requisicoes_table():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        create_table_query = """
        CREATE TABLE IF NOT EXISTS Requisicoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            preenchimento_id INTEGER NOT NULL,
            cod_material INTEGER NOT NULL,
            quantidade INTEGER NOT NULL,
            ticket TEXT NOT NULL,
            data_requisicao DATETIME NOT NULL,
            usuario TEXT NOT NULL,
            FOREIGN KEY (preenchimento_id) REFERENCES SolicitacoesPreenchidas(id),
            FOREIGN KEY (cod_material) REFERENCES Materiais(CodMaterial)
        )
        """
        cursor.execute(create_table_query)
        conn.commit()
        print("Tabela 'Requisicoes' criada com sucesso.")
        cursor.close()
        conn.close()
    except sqlite3.Error as e:
        print(f"Erro ao criar a tabela Requisicoes: {str(e)}")
        return False
    return True

def create_pedidos_compra_table():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        create_table_query = """
        CREATE TABLE IF NOT EXISTS PedidosCompra (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_pedido TEXT NOT NULL UNIQUE,
            data_criacao DATETIME NOT NULL,
            usuario TEXT NOT NULL,
            status TEXT NOT NULL,
            pdf_path TEXT,
            valor_total REAL NOT NULL
        )
        """
        cursor.execute(create_table_query)
        create_associacao_table_query = """
        CREATE TABLE IF NOT EXISTS pedido_preenchimento_associacao (
            pedido_id INTEGER,
            preenchimento_id INTEGER,
            PRIMARY KEY (pedido_id, preenchimento_id),
            FOREIGN KEY (pedido_id) REFERENCES PedidosCompra(id),
            FOREIGN KEY (preenchimento_id) REFERENCES SolicitacoesPreenchidas(id)
        )
        """
        cursor.execute(create_associacao_table_query)
        conn.commit()
        print("Tabelas 'PedidosCompra' e 'pedido_preenchimento_associacao' criadas com sucesso.")
        cursor.close()
        conn.close()
    except sqlite3.Error as e:
        print(f"Erro ao criar tabelas PedidosCompra ou associação: {str(e)}")
        return False
    return True

def create_fornecedores_db():
    try:
        conn = sqlite3.connect(DB_PATH_FORNECEDORES)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS fornecedores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome_fantasia TEXT NOT NULL,
                cnpj TEXT NOT NULL UNIQUE,
                telefone TEXT NOT NULL,
                email TEXT NOT NULL,
                endereco TEXT NOT NULL,
                bairro TEXT NOT NULL,
                cidade TEXT NOT NULL,
                estado TEXT NOT NULL,
                contato TEXT NOT NULL,
                materiais TEXT NOT NULL
            )
        ''')
        conn.commit()
        conn.close()
        app.logger.info("Tabela 'fornecedores' criada com sucesso.")
        return True
    except sqlite3.Error as e:
        app.logger.error(f"Erro ao criar tabela fornecedores: {str(e)}")
        return False

class HistoricoEstoque(db.Model):
    __tablename__ = 'HistoricoEstoque'
    id = db.Column(db.Integer, primary_key=True)
    cod_material = db.Column(db.Integer, db.ForeignKey('Materiais.CodMaterial'))
    usuario = db.Column(db.Text, nullable=False)
    quantidade_anterior = db.Column(db.Integer, nullable=False)
    quantidade_nova = db.Column(db.Integer, nullable=False)
    data_alteracao = db.Column(db.DateTime, nullable=False, default=get_local_time)
    motivo = db.Column(db.Text, nullable=True)
    
    material = db.relationship('Materiais', backref='historico_estoque')

def get_db_connection(db_path):
    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        return conn
    except sqlite3.Error as e:
        app.logger.error(f"Erro ao conectar ao banco de fornecedores: {str(e)}")
        return None

def validate_cnpj(cnpj):
    cnpj = ''.join(filter(str.isdigit, cnpj))
    if len(cnpj) != 14:
        return False
    return True

def validate_email(email):
    return re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email) is not None

@app.template_filter('format_cnpj')
def format_cnpj(cnpj):
    if not cnpj:
        return ''
    cnpj = ''.join(filter(str.isdigit, cnpj))
    return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"

# Funções de Autenticação
def _ler_senhas_do_disco():
    """Lê o arquivo senhas.txt — chamada apenas pelo cache de ler_senhas()."""
    senhas = {}
    try:
        with open(SENHAS_FILE, "r", encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    partes = line.split("%")
                    if len(partes) >= 3:
                        usuario = partes[0].replace('@', '')  # Remove @ automaticamente
                        senha = partes[1]
                        pagina = partes[2]
                        empresa = partes[3] if len(partes) >= 4 else ""
                        senhas[usuario] = {
                            "senha": senha,
                            "pagina": pagina,
                            "empresa": empresa,
                            "usuario_base": usuario
                        }
    except Exception as e:
        logging.error(f"Erro ao ler senhas.txt. Tipo: {type(e).__name__}")
    return senhas

def ler_senhas():
    """Retorna senhas com cache de 60 s para reduzir I/O desnecessário."""
    return get_cached_or_execute('senhas_usuarios', _ler_senhas_do_disco, ttl_seconds=60)

def _invalidar_cache_senhas():
    """Invalida o cache de senhas após qualquer escrita no arquivo."""
    with _cache_lock:
        _cache.pop('senhas_usuarios', None)
        _cache_timestamp.pop('senhas_usuarios', None)

def salvar_senhas(senhas):
    try:
        with open(SENHAS_FILE, "w", encoding='utf-8') as f:
            for usuario, dados in senhas.items():
                empresa = dados.get('empresa', '')
                f.write(f"{usuario}%{dados['senha']}%{dados['pagina']}%{empresa}\n")
        _invalidar_cache_senhas()  # garante que próximo request lê dados atualizados
    except Exception as e:
        logging.error(f"Erro ao salvar senhas.txt. Tipo: {type(e).__name__}")


def registrar_log(usuario, tipo_acao, descricao=None, ip=None):
    """
    Registra ação no arquivo de log
    
    Args:
        usuario: Nome do usuário
        tipo_acao: Tipo de ação ('login', 'logout', 'reprovar_solicitacao_individual', etc.)
        descricao: Descrição adicional da ação (opcional)
        ip: Endereço IP do usuário (opcional)
    """
    acao_map = {
        'login': 'Acesso',
        'logout': 'Logout',
        'reprovar_solicitacao_individual': 'Reprovação Individual de Solicitação',
        'aprovar_solicitacao': 'Aprovação de Solicitação',
        'reprovar_solicitacao': 'Reprovação de Solicitação'
    }
    
    acao = acao_map.get(tipo_acao, tipo_acao)
    ip_info = f" - IP: {ip}" if ip else ''
    descricao_info = f" - {descricao}" if descricao else ''
    
    try:
        with open("arquivo.log", "a", encoding="utf-8") as log_file:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            log_file.write(f"{timestamp} - {acao} do usuário: {usuario}{ip_info}{descricao_info}\n")
    except Exception as e:
        logging.error(f"Erro ao registrar log: {str(e)}")
        
def ler_logs():
    try:
        if not os.path.exists('arquivo.log'):
            with open('arquivo.log', 'w', encoding='utf-8') as log_file:
                log_file.write("Log de início\n")
        
        with open('arquivo.log', 'rb') as log_file:
            raw_data = log_file.read()
            encoding = chardet.detect(raw_data)['encoding'] or 'utf-8'
        
        with open('arquivo.log', 'r', encoding=encoding) as log_file:
            return log_file.readlines()
    except Exception as e:
        logging.error(f"Erro ao ler logs: {str(e)}")
        return []

def validar_usuario(usuario):
    if not usuario or len(usuario) < 3:
        return False, "O usuário deve ter pelo menos 3 caracteres."
    if not re.match(r'^[a-zA-Z0-9._-]+$', usuario):
        return False, "O usuário só pode conter letras, números, pontos, hífens ou sublinhados."
    if '%' in usuario:
        return False, "O usuário não pode conter o caractere '%'."
    return True, ""

def validar_senha(senha):
    if len(senha) < 6:
        return False, "A senha deve ter pelo menos 6 caracteres."
    if not re.search(r"[A-Z]", senha):
        return False, "A senha deve conter pelo menos uma letra maiúscula."
    if not re.search(r"[0-9]", senha):
        return False, "A senha deve conter pelo menos um número."
    if not re.search(r"[\W_]", senha):
        return False, "A senha deve conter pelo menos um caractere especial."
    return True, ""

# Registrar funções e objetos no jinja_env.globals
app.jinja_env.globals.update(
    ler_senhas=ler_senhas,
    salvar_senhas=salvar_senhas,
    registrar_log=registrar_log,
    ler_logs=ler_logs,
    validar_usuario=validar_usuario,
    validar_senha=validar_senha,
    get_rastreabilidade_entries=get_rastreabilidade_entries,
    Kanban=Kanban,
    Materiais=Materiais,
    SolicitacoesCompra=SolicitacoesCompra,
    SolicitacoesPreenchidas=SolicitacoesPreenchidas,
    Estoque=Estoque,
    Requisicoes=Requisicoes,
    PedidosCompra=PedidosCompra,
    db=db
)

# Blueprint para rotas
routes_bp = Blueprint('routes_bp', __name__)

def requer_role(*roles_permitidas):
    """Decorator para verificar se o usuário tem a role necessária.
    Exemplo: @requer_role('menu_master.html', 'menu_aprovacao.html')
    """
    from functools import wraps
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'usuario' not in session:
                return jsonify({'success': False, 'error': 'Não autenticado'}), 401
            usuario_completo = session.get('usuario_completo', session['usuario'])
            senhas = ler_senhas()
            role = senhas.get(usuario_completo, {}).get('pagina', '')
            if not any(r in role for r in roles_permitidas):
                logging.warning(f"Acesso negado para role '{role}' em {request.endpoint}")
                return jsonify({'success': False, 'error': 'Sem permissão'}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@routes_bp.before_request
def verificar_sessao():
    """Verifica inatividade e controle de acesso por menu em um único hook."""
    endpoints_publicos = {'routes_bp.login', 'static'}
    if request.endpoint in endpoints_publicos:
        return

    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))

    # Verificar se o usuário ainda existe no arquivo de senhas — protege contra
    # sessões órfãs de contas excluídas ou desativadas.
    _uc = session.get('usuario_completo', session.get('usuario'))
    if _uc and _uc not in ler_senhas():
        # 🔒 Log de segurança obrigatório: sessão órfã / conta removida
        logging.warning(
            'AUTH_FAILURE | usuario=%s | ip=%s | session_id=%s | endpoint=%s | motivo=usuario_nao_encontrado',
            _uc, request.remote_addr, request.cookies.get('session', 'n/a')[:16], request.endpoint
        )
        try:
            registrar_log(_uc, 'auth_failure_sessao_orfan', f'Sessão inválida bloqueada em {request.endpoint}', request.remote_addr)
        except Exception:
            pass  # não deixar falha de log derrubar o redirect de segurança
        session.clear()
        flash('Sua conta foi alterada. Faça login novamente.', 'warning')
        return redirect(url_for('routes_bp.login'))

    # Verificar inatividade
    agora = datetime.now()
    ultima_atividade = session.get('ultima_atividade')
    if ultima_atividade:
        try:
            ultima_atividade = datetime.strptime(ultima_atividade, '%Y-%m-%d %H:%M:%S.%f')
            if agora - ultima_atividade > timedelta(minutes=30):
                usuario = session.get('usuario')
                registrar_log(usuario, 'logout (inatividade)', request.remote_addr)
                session.clear()
                flash('Você foi desconectado por inatividade.', 'warning')
                return redirect(url_for('routes_bp.login'))
        except (ValueError, TypeError):
            pass  # Formato inválido — ignora e reseta
    session['ultima_atividade'] = str(agora)

    # Verificar acesso ao menu correto
    endpoints_menus = {
        'routes_bp.menu_master': 'menu_master.html',
        'routes_bp.menu_supervisor': 'menu_supervisor.html',
        'routes_bp.menu_aprovacao': 'menu_aprovacao.html',
        'routes_bp.menu_comprador': 'menu_comprador.html',
        'routes_bp.menu_cadastro': 'menu_cadastro.html',
        'routes_bp.menu_solicitante': 'menu_solicitante.html',
        'routes_bp.menu_financeiro': 'menu_financeiro.html',
        'routes_bp.menu_estoquista': 'menu_estoquista.html',
        'routes_bp.menu_auditoria': 'menu_auditoria.html',
    }
    if request.endpoint in endpoints_menus:
        usuario_completo = session.get('usuario_completo')
        if usuario_completo:
            senhas = ler_senhas()
            if usuario_completo in senhas:
                menu_permitido = senhas[usuario_completo]['pagina']
                if menu_permitido != endpoints_menus[request.endpoint]:
                    session.clear()
                    flash('Acesso não permitido. Faça login novamente.', 'error')
                    return redirect(url_for('routes_bp.login'))

@routes_bp.route('/', methods=['GET'])
def home():
    return redirect(url_for('routes_bp.login'))

@routes_bp.route('/uploads/<path:filename>')
def uploads(filename):
    if 'usuario' not in session:
        abort(401)
    safe_path = safe_join(app.config['UPLOAD_FOLDER'], filename)
    if safe_path is None or not os.path.isfile(safe_path):
        abort(404)
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@routes_bp.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def login():
    if request.method == 'POST':
        usuario_completo = request.form.get('usuario', '').strip()
        senha = request.form.get('senha', '').strip()
        
        senhas = ler_senhas()
        usuario_dados = senhas.get(usuario_completo)
        autenticado = False

        if usuario_dados:
            hash_armazenado = usuario_dados["senha"]
            # Migração inline: se ainda for plaintext (sem prefixo werkzeug), migrar para hash
            if not hash_armazenado.startswith(('pbkdf2:', 'scrypt:', 'bcrypt')):
                if hash_armazenado == senha:
                    usuario_dados["senha"] = generate_password_hash(senha)
                    salvar_senhas(senhas)
                    autenticado = True
            else:
                autenticado = check_password_hash(hash_armazenado, senha)

        if autenticado:
            usuario_base = usuario_dados["usuario_base"]
            session['usuario'] = usuario_base
            session['usuario_completo'] = usuario_completo
            session['_fresh'] = True
            registrar_log(usuario_base, 'login', ip=request.remote_addr)
            pagina_destino = usuario_dados["pagina"].replace('.html', '')
            return redirect(url_for(f'routes_bp.{pagina_destino}'))
        
        flash('Credenciais inválidas', 'error')
    
    return render_template('login.html')

# verificar_acesso_menu foi incorporada em verificar_sessao acima

@routes_bp.route('/logout', methods=['GET'])
def logout():
    usuario = session.get('usuario')
    if usuario:
        registrar_log(usuario, 'logout', request.remote_addr)
        session.clear()
    return redirect(url_for('routes_bp.login'))

@routes_bp.route('/senhas', methods=['GET'])
def listar_senhas():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    senhas = ler_senhas()
    return render_template('senhas.html', senhas=senhas)

@routes_bp.route('/adicionar_senha', methods=['POST'])
def adicionar_senha():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))

    try:
        usuario = request.form['usuario'].strip()
        senha = request.form['senha'].strip()
        pagina = request.form['pagina']
        pagina_personalizada = request.form.get('pagina_personalizada', '').strip()
        empresa = request.form.get('empresa', '').strip()

        # Verifica se algum campo contém o caractere proibido
        for campo_nome, campo_valor in [("Usuário", usuario), ("Senha", senha), ("Empresa", empresa), ("Página personalizada", pagina_personalizada)]:
            if "%" in campo_valor:
                flash(f'O campo "{campo_nome}" não pode conter o caractere "%".', 'error')
                return redirect(url_for('routes_bp.listar_senhas'))

        # Validação básica do usuário
        is_valid_usuario, msg_usuario = validar_usuario(usuario)
        if not is_valid_usuario:
            flash(msg_usuario, 'error')
            return redirect(url_for('routes_bp.listar_senhas'))

        # Validação básica da senha
        is_valid_senha, msg_senha = validar_senha(senha)
        if not is_valid_senha:
            flash(msg_senha, 'error')
            return redirect(url_for('routes_bp.listar_senhas'))

        # Validação da página
        if not pagina:
            flash('O campo página é obrigatório.', 'error')
            return redirect(url_for('routes_bp.listar_senhas'))

        if pagina == 'outro':
            if not pagina_personalizada:
                flash('Página personalizada é obrigatória quando "outro" é selecionado.', 'error')
                return redirect(url_for('routes_bp.listar_senhas'))
            if not pagina_personalizada.endswith('.html'):
                flash('A página personalizada deve terminar com ".html".', 'error')
                return redirect(url_for('routes_bp.listar_senhas'))
            pagina = pagina_personalizada

        if not empresa:
            flash('O campo Empresa é obrigatório.', 'error')
            return redirect(url_for('routes_bp.listar_senhas'))

        senhas = ler_senhas()
        
        # Verifica se é uma atualização ou criação nova
        if usuario in senhas:
            mensagem = 'Senha atualizada com sucesso.'
        else:
            mensagem = 'Senha adicionada com sucesso.'

        senhas[usuario] = {
            "senha": generate_password_hash(senha),
            "pagina": pagina,
            "empresa": empresa
        }

        salvar_senhas(senhas)
        flash(mensagem, 'success')
        return redirect(url_for('routes_bp.listar_senhas'))

    except Exception as e:
        flash('Erro ao adicionar/atualizar senha.', 'error')
        return redirect(url_for('routes_bp.listar_senhas'))


@routes_bp.route('/excluir_senha/<usuario>', methods=['POST'])
def excluir_senha(usuario):
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        senhas = ler_senhas()
        if usuario in senhas:
            del senhas[usuario]
            salvar_senhas(senhas)
            flash(f'Senha para {usuario} excluída com sucesso.', 'success')
        else:
            flash(f'Usuário {usuario} não encontrado.', 'error')
        return redirect(url_for('routes_bp.listar_senhas'))
    except Exception as e:
        flash('Erro ao excluir senha.', 'error')
        return redirect(url_for('routes_bp.listar_senhas'))

@routes_bp.route('/kanban', methods=['GET'])
def kanban():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('kanban.html')

@routes_bp.route('/tasks', methods=['GET'])
def get_tasks():
    tasks = Kanban.query.all()
    return jsonify({
        'tasks': [{'id': t.id, 'task_name': t.task_name, 'status': t.status} for t in tasks]
    })

@routes_bp.route('/tasks', methods=['POST'])
def add_task():
    data = request.get_json()
    if not data or 'task_name' not in data or 'status' not in data:
        return jsonify({'message': 'Task name and status are required'}), 400
    
    valid_statuses = ['A Fazer', 'Em Progresso', 'Concluído']
    if data['status'] not in valid_statuses:
        return jsonify({'message': 'Invalid status'}), 400
    
    task = Kanban(task_name=data['task_name'], status=data['status'])
    db.session.add(task)
    db.session.commit()
    return jsonify({'message': 'Task added successfully'}), 201

@routes_bp.route('/tasks/<int:id>', methods=['PUT'])
def update_task(id):
    data = request.get_json()
    if not data or 'status' not in data:
        return jsonify({'message': 'Status is required'}), 400
    
    valid_statuses = ['A Fazer', 'Em Progresso', 'Concluído']
    if data['status'] not in valid_statuses:
        return jsonify({'message': 'Invalid status'}), 400
    
    task = Kanban.query.get_or_404(id)
    task.status = data['status']
    db.session.commit()
    return jsonify({'message': 'Task updated successfully'})

@routes_bp.route('/tasks/<int:id>', methods=['DELETE'])
def delete_task(id):
    task = Kanban.query.get_or_404(id)
    db.session.delete(task)
    db.session.commit()
    return jsonify({'message': 'Task deleted successfully'})

#Menu Master
@routes_bp.route('/menu_master', methods=['GET'])
def menu_master():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_master.html')

#Menu Master
@routes_bp.route('/menu_aprovacao', methods=['GET'])
def menu_aprovacao():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_aprovacao.html')

#Menu Estoquista
@routes_bp.route('/menu_estoquista', methods=['GET'])
def menu_estoquista():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_estoquista.html')

#Menu Solicitante
@routes_bp.route('/menu_solicitante', methods=['GET'])
def menu_solicitante():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_solicitante.html')

#Menu Comprado
@routes_bp.route('/menu_comprador', methods=['GET'])
def menu_comprador():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_comprador.html')

#Menu Supervisor
@routes_bp.route('/menu_supervisor', methods=['GET'])
def menu_supervisor():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_supervisor.html')

#Menu Auditoria
@routes_bp.route('/menu_auditoria', methods=['GET'])
def menu_auditoria():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_auditoria.html')

#Menu Financeiro
@routes_bp.route('/menu_financeiro', methods=['GET'])
def menu_financeiro():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_financeiro.html')

#Menu Cadastro
@routes_bp.route('/menu_cadastro', methods=['GET'])
def menu_cadastro():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_cadastro.html')

#Usuario Desativado
@routes_bp.route('/desativado', methods=['GET'])
def desativado():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('desativado.html')

@routes_bp.route('/menu_reenvio', methods=['GET'])
def menu_reenvio():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_reenvio.html')

@routes_bp.route('/formulario_custodia', methods=['GET'])
def formulario_custodia():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('formulario_custodia.html')


@routes_bp.route('/menu_logistica_auditoria', methods=['GET'])
def menu_logistica_auditoria():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_logistica_auditoria.html')

@routes_bp.route('/menu_logistica_supervisor', methods=['GET'])
def menu_logistica_supervisor():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_logistica_supervisor.html')

@routes_bp.route('/menu_logistica_analista1', methods=['GET'])
def menu_logistica_analista1():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_logistica_analista1.html')

@routes_bp.route('/menu_logistica_analista2', methods=['GET'])
def menu_logistica_analista2():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_logistica_analista2.html')

@routes_bp.route('/menu_logistica_analista3', methods=['GET'])
def menu_logistica_analista3():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('menu_logistica_analista3.html')

@routes_bp.route('/logs', methods=['GET'])
def logs():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    logs = ler_logs()
    return render_template('logs.html', logs=logs)

@routes_bp.route('/cadastrar_material', methods=['GET'])
def cadastrar_material():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('cadastrar_material.html')

@routes_bp.route('/materiais', methods=['GET'])
def listar_materiais():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    try:
        materiais = get_cached_or_execute(
            'todos_materiais',
            lambda: [m.to_dict() for m in Materiais.query.all()],
            ttl_seconds=300
        )
        return render_template('materiais.html', materiais=materiais)
    except Exception as e:
        _flash_erro_seguro('Erro ao carregar materiais.', e, app.logger)
        return render_template('materiais.html', materiais=[])

@routes_bp.route('/adicionar_material', methods=['POST'])
def adicionar_material():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    try:
        descricao = request.form.get('descricao', '').strip()
        empresa = request.form.get('empresa', '').strip()
        aplicacao = request.form.get('aplicacao', '').strip()

        if not descricao or not empresa or not aplicacao:
            flash('Todos os campos são obrigatórios.', 'error')
            return redirect(url_for('routes_bp.cadastrar_material'))

        if len(descricao) < 3:
            flash('A descrição deve ter pelo menos 3 caracteres.', 'error')
            return redirect(url_for('routes_bp.cadastrar_material'))

        if len(empresa) < 2:
            flash('A empresa deve ter pelo menos 2 caracteres.', 'error')
            return redirect(url_for('routes_bp.cadastrar_material'))

        if len(aplicacao) < 3:
            flash('A aplicação deve ter pelo menos 3 caracteres.', 'error')
            return redirect(url_for('routes_bp.cadastrar_material'))

        material = Materiais(
            DescricaoMaterial=descricao,
            Empresa=empresa,
            Aplicacao=aplicacao,
            QuantidadeEstoque=0,
            Fornecedor=None,
            NumeroNF=None,
            FatorConsumo=0.0
        )
        db.session.add(material)
        db.session.commit()

        flash('Material adicionado com sucesso.', 'success')
        return redirect(url_for('routes_bp.listar_materiais'))
    except Exception as e:
        db.session.rollback()
        _flash_erro_seguro('Erro ao adicionar material.', e, app.logger)
        return redirect(url_for('routes_bp.cadastrar_material'))

@routes_bp.route('/excluir_material/<int:cod>', methods=['POST'])
def excluir_material(cod):
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    try:
        material = Materiais.query.get_or_404(cod)
        db.session.delete(material)
        db.session.commit()
        flash(f'Material {material.DescricaoMaterial} excluído com sucesso.', 'success')
        return redirect(url_for('routes_bp.listar_materiais'))
    except Exception as e:
        _flash_erro_seguro('Erro ao excluir material.', e, app.logger)
        return redirect(url_for('routes_bp.listar_materiais'))

@routes_bp.route('/buscar_material', methods=['GET', 'POST'])
def buscar_material():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    materiais = []
    empresas_unicas = []
    aplicacoes_unicas = []
    
    try:
        # Obter valores únicos para os filtros
        empresas_unicas = db.session.query(
            Materiais.Empresa
        ).distinct().order_by(
            Materiais.Empresa
        ).all()
        
        aplicacoes_unicas = db.session.query(
            Materiais.Aplicacao
        ).distinct().order_by(
            Materiais.Aplicacao
        ).all()
        
        empresas_unicas = [e[0] for e in empresas_unicas if e[0]]
        aplicacoes_unicas = [a[0] for a in aplicacoes_unicas if a[0]]

        if request.method == 'POST':
            termo = request.form.get('termo', '').strip()
            empresa = request.form.get('empresa', '').strip()
            aplicacao = request.form.get('aplicacao', '').strip()
            status = request.form.get('status', '').strip()

            query = Materiais.query

            if termo:
                if termo.isdigit():
                    query = query.filter_by(CodMaterial=int(termo))
                else:
                    query = query.filter(
                        Materiais.DescricaoMaterial.ilike(f'%{termo}%')
                    )
            
            if empresa:
                query = query.filter_by(Empresa=empresa)
            
            if aplicacao:
                query = query.filter_by(Aplicacao=aplicacao)
            
            if status:
                query = query.filter_by(Ativo=(status == 'ativo'))

            materiais = query.all()
            
            if not materiais:
                flash('Nenhum material encontrado com os filtros aplicados.', 'info')
        
    except Exception as e:
        _flash_erro_seguro('Erro ao buscar materiais.', e, app.logger)
    
    return render_template('buscar_material.html', 
                         materiais=materiais,
                         empresas_unicas=empresas_unicas,
                         aplicacoes_unicas=aplicacoes_unicas)

@routes_bp.route('/solicitar_compra', defaults={'cod': None}, methods=['GET'])
@routes_bp.route('/solicitar_compra/<int:cod>', methods=['GET'])
def solicitar_compra(cod):
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))

    try:
        senhas = ler_senhas()
        usuario = session.get('usuario')

        if usuario not in senhas:
            flash('Usuário não encontrado.', 'error')
            return redirect(url_for('routes_bp.login'))

        empresa_usuario = senhas[usuario].get('empresa', '')
        
        # BUSCAR TODOS OS MATERIAIS — usando cache de 5 minutos (lista de dicts)
        todos_materiais = get_cached_or_execute(
            'todos_materiais',
            lambda: [m.to_dict() for m in Materiais.query.all()],
            ttl_seconds=300
        )

        # Se o código foi passado (botão de material)
        if cod:
            material = Materiais.query.get_or_404(cod)
            return render_template(
                'solicitar_compra.html',
                material=material,
                empresa_usuario=empresa_usuario,
                todos_materiais=todos_materiais  # ADICIONE ESTE PARÂMETRO
            )
        else:
            # Se o acesso veio do menu (sem material específico)
            return render_template(
                'solicitar_compra.html',
                material=None,
                empresa_usuario=empresa_usuario,
                todos_materiais=todos_materiais  # ADICIONE ESTE PARÂMETRO
            )

    except SQLAlchemyError as e:
        _flash_erro_seguro('Erro ao carregar material.', e, app.logger)
        return redirect(url_for('routes_bp.buscar_material'))
    except Exception as e:
        _flash_erro_seguro('Erro inesperado.', e, app.logger)
        return redirect(url_for('routes_bp.buscar_material'))
    

    
@routes_bp.route('/api/materiais_search', methods=['GET'])
def api_materiais_search():
    """API para buscar materiais com paginação para Select2"""
    try:
        search_term = request.args.get('q', '').strip()
        page = request.args.get('page', 1, type=int)
        per_page = 30
        
        # Query base
        query = Materiais.query
        
        # Aplicar filtro de busca se houver termo
        if search_term:
            if search_term.isdigit():
                # Busca por código
                query = query.filter(Materiais.CodMaterial == int(search_term))
            else:
                # Busca por descrição
                query = query.filter(Materiais.DescricaoMaterial.ilike(f'%{search_term}%'))
        
        # Paginação
        materiais_paginated = query.order_by(Materiais.DescricaoMaterial).paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        # Formatar resultados para Select2
        materiais_data = []
        for material in materiais_paginated.items:
            materiais_data.append({
                'id': material.CodMaterial,
                'text': f"{material.DescricaoMaterial} (Cód: {material.CodMaterial})",
                'CodMaterial': material.CodMaterial,
                'DescricaoMaterial': material.DescricaoMaterial
            })
        
        return jsonify({
            'success': True,
            'materiais': materiais_data,
            'total_count': materiais_paginated.total,
            'pagination': {
                'more': materiais_paginated.has_next
            }
        })
        
    except Exception as e:
        logging.error(f"Erro na API de busca de materiais: {type(e).__name__}")
        return jsonify({
            'success': False,
            'error': 'Erro interno. Tente novamente.',
            'materiais': []
        }), 500

@routes_bp.route('/api/materiais', methods=['GET'])
def api_materiais():
    """API para buscar todos os materiais"""
    try:
        print("🔍 Acessando API de materiais...")
        
        # Buscar todos os materiais — cache de 5 minutos (lista de dicts)
        materiais = get_cached_or_execute(
            'todos_materiais',
            lambda: [m.to_dict() for m in Materiais.query.all()],
            ttl_seconds=300
        )
        if not materiais:
            return jsonify({
                'success': False, 
                'error': 'Nenhum material cadastrado',
                'materiais': []
            })
        
        resultado = {
            'success': True,
            'materiais': [{
                'CodMaterial': m['CodMaterial'],
                'DescricaoMaterial': m['DescricaoMaterial']
            } for m in materiais]
        }
        
        return jsonify(resultado)
        
    except Exception as e:
        logging.error(f"Erro na API de materiais: {type(e).__name__}", exc_info=False)
        return jsonify({
            'success': False,
            'error': 'Erro interno. Tente novamente.',
            'materiais': []
        }), 500

@routes_bp.route('/abrir_solicitacao', methods=['POST'])
@limiter.limit("30 per minute")
def abrir_solicitacao():
    if 'usuario' not in session:
        flash('Você precisa estar logado para fazer uma solicitação.', 'error')
        return redirect(url_for('routes_bp.login'))
    
    try:
        # Obter arrays do formulário
        cod_materiais = request.form.getlist('cod_material[]')
        especificacoes = request.form.getlist('especificacao[]')
        quantidades = request.form.getlist('quantidade[]')  # Recebe como string
        unidades_medida = request.form.getlist('unidade_medida[]')
        aplicacoes = request.form.getlist('aplicacao[]')
        empresas = request.form.getlist('empresa[]')
        marcas = request.form.getlist('marca[]')
        ativos = request.form.getlist('ativo[]')
        nomes_ativos = request.form.getlist('nome_ativo[]')
        prioridades = request.form.getlist('prioridade[]')
        fotos = request.files.getlist('foto[]')

        aplicacao_geral = request.form.get('aplicacao', '').strip()

        solicitacoes_criadas = 0
        
        for i, cod_material in enumerate(cod_materiais):
            if not cod_material:
                continue
                
            try:
                cod_material = int(cod_material)
                if cod_material <= 0:
                    raise ValueError
            except ValueError:
                flash(f'Código do material inválido na solicitação {i+1}.', 'error')
                continue

            material = db.session.get(Materiais, cod_material)
            if not material:
                flash(f'Material com código {cod_material} não encontrado na solicitação {i+1}.', 'error')
                continue

            if (i >= len(especificacoes) or not especificacoes[i] or 
                i >= len(quantidades) or not quantidades[i] or
                i >= len(unidades_medida) or not unidades_medida[i] or
                i >= len(empresas) or not empresas[i] or
                i >= len(ativos) or not ativos[i] or
                i >= len(prioridades) or not prioridades[i]):
                flash(f'Campos obrigatórios não preenchidos na solicitação {i+1}.', 'error')
                continue

            # 🔴 CORREÇÃO: Converter quantidade para float e depois para string para armazenar
            quantidade_str = quantidades[i].strip()
            try:
                # Converte para float para validar, mas armazena como string
                quantidade_float = float(quantidade_str.replace(',', '.'))
                if quantidade_float <= 0:
                    flash(f'Quantidade deve ser maior que zero na solicitação {i+1}.', 'error')
                    continue
                # Armazena como string com ponto decimal
                quantidade_armazenar = str(quantidade_float).rstrip('0').rstrip('.') if '.' in str(quantidade_float) else str(quantidade_float)
            except ValueError:
                flash(f'Quantidade inválida na solicitação {i+1}. Use números como 6.2 ou 7.3', 'error')
                continue

            # Processar aplicação
            aplicacao = None
            if i < len(aplicacoes) and aplicacoes[i] and aplicacoes[i].strip():
                aplicacao = aplicacoes[i].strip()
            else:
                if aplicacao_geral:
                    aplicacao = aplicacao_geral
                else:
                    aplicacao = material.Aplicacao if material and material.Aplicacao else 'Aplicação não especificada'

            # Processar upload da foto
            foto_path = None
            if i < len(fotos) and fotos[i] and fotos[i].filename:
                foto = fotos[i]
                if allowed_file(foto.filename, {'jpg', 'jpeg', 'png', 'pdf'}):
                    filename = f"{uuid.uuid4()}_{secure_filename(foto.filename)}"
                    abs_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    foto.save(abs_path)
                    foto_path = filename  # armazena apenas o nome relativo, não o path absoluto

            # 🔴 CRIAR SOLICITAÇÃO - quantidade como string
            solicitacao = SolicitacoesCompra(
                cod_material=cod_material,
                especificacao=especificacoes[i],
                quantidade=quantidade_armazenar,  # Armazena como string
                unidade_medida=unidades_medida[i],
                aplicacao=aplicacao.strip().lower(),
                aplicacao_geral=aplicacao_geral,
                empresa=empresas[i],
                usuario=session['usuario'],
                foto_path=foto_path,
                marca=marcas[i] if i < len(marcas) and marcas[i] else None,
                ativo=ativos[i],
                nome_ativo=nomes_ativos[i] if i < len(nomes_ativos) and ativos[i] == 'Sim' and nomes_ativos[i] else None,
                prioridade=prioridades[i],
                status_aprovacao=None,
                comprador_atribuido=None  # será atribuído logo abaixo com tratamento individual
            )
            # Atribuir comprador com proteção individual para não cancelar as demais solicitações
            try:
                solicitacao.comprador_atribuido = get_next_comprador(aplicacao)
            except Exception as e_comprador:
                app.logger.warning(
                    f"Não foi possível atribuir comprador para solicitação {i+1}: {type(e_comprador).__name__}"
                )
                solicitacao.comprador_atribuido = None
            db.session.add(solicitacao)
            solicitacoes_criadas += 1

        if solicitacoes_criadas > 0:
            db.session.commit()
            invalidar_cache_filtros()  # empresas/aplicações podem ter mudado
            flash(f'{solicitacoes_criadas} solicitações de compra abertas com sucesso.', 'success')
        else:
            flash('Nenhuma solicitação válida para criar.', 'error')

        return redirect(url_for('routes_bp.solicitar_compra'))
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in abrir_solicitacao: {str(e)}")
        _flash_erro_seguro('Erro ao abrir solicitações.', e, app.logger)
        return redirect(url_for('routes_bp.buscar_material'))
    
def migrate_quantidade_to_text():
    """Migra a coluna quantidade de INTEGER para TEXT"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        cursor.execute("PRAGMA table_info(SolicitacoesCompra)")
        columns = cursor.fetchall()
        quantidade_col = next((col for col in columns if col[1] == 'quantidade'), None)
        
        if quantidade_col and quantidade_col[2] in ['INTEGER', 'REAL', 'FLOAT']:
            print("🔄 Migrando coluna quantidade para TEXT...")
            
            # Criar tabela temporária
            cursor.execute("""
                CREATE TABLE SolicitacoesCompra_temp (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    cod_material INTEGER NOT NULL,
                    especificacao TEXT NOT NULL,
                    quantidade TEXT NOT NULL,
                    unidade_medida TEXT NOT NULL DEFAULT 'Unidade',
                    aplicacao TEXT,
                    aplicacao_geral TEXT,
                    empresa TEXT NOT NULL,
                    data_solicitacao DATETIME NOT NULL,
                    usuario TEXT NOT NULL,
                    foto_path TEXT,
                    marca TEXT,
                    ativo TEXT NOT NULL,
                    nome_ativo TEXT,
                    prioridade TEXT NOT NULL DEFAULT 'Programado',
                    status_aprovacao TEXT,
                    observacoes_col TEXT,
                    comprador_atribuido TEXT
                )
            """)
            
            # Copiar dados
            cursor.execute("""
                INSERT INTO SolicitacoesCompra_temp (
                    id, cod_material, especificacao, quantidade, unidade_medida,
                    aplicacao, aplicacao_geral, empresa, data_solicitacao, usuario,
                    foto_path, marca, ativo, nome_ativo, prioridade,
                    status_aprovacao, observacoes_col, comprador_atribuido
                )
                SELECT 
                    id, cod_material, especificacao, CAST(quantidade AS TEXT), unidade_medida,
                    aplicacao, aplicacao_geral, empresa, data_solicitacao, usuario,
                    foto_path, marca, ativo, nome_ativo, prioridade,
                    status_aprovacao, observacoes_col, comprador_atribuido
                FROM SolicitacoesCompra
            """)
            
            cursor.execute("DROP TABLE SolicitacoesCompra")
            cursor.execute("ALTER TABLE SolicitacoesCompra_temp RENAME TO SolicitacoesCompra")
            
            conn.commit()
            print("✅ Coluna quantidade migrada para TEXT com sucesso!")
        else:
            print("✅ Coluna quantidade já é do tipo TEXT")
        
        conn.close()
        return True
        
    except sqlite3.Error as e:
        print(f"❌ Erro na migração: {str(e)}")
        return False

def _get_compradores_from_file():
    """Função interna para ler compradores do arquivo (usada apenas pelo cache)"""
    senhas = ler_senhas()
    compradores = []
    
    for usuario_completo, dados in senhas.items():
        if 'comprador' in dados['pagina'].lower():
            compradores.append(dados['usuario_base'])
    
    return sorted(set(compradores))

def get_compradores():
    """Retorna lista de nomes de compradores a partir do cache"""
    return get_cached_or_execute('compradores', _get_compradores_from_file, ttl_seconds=300)


@routes_bp.route('/listar_solicitacoes', methods=['GET'])
def listar_solicitacoes():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))

    try:
        # Parâmetros de paginação
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        
        # Parâmetros de filtro
        empresa = request.args.get('empresa')
        usuario = request.args.get('usuario')
        aplicacao_filter = request.args.get('aplicacao')
        status = request.args.get('status')
        comprador_filter = request.args.get('comprador')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')

        # 🔴 CORREÇÃO: Subquery para IDs das solicitações que já estão em pedidos de compra
        # Buscar preenchimentos que estão associados a pedidos
        from sqlalchemy import exists
        
        # Subquery: IDs das solicitações que têm preenchimentos associados a pedidos
        subquery_pedidos = db.session.query(
            SolicitacoesPreenchidas.solicitacao_id
        ).join(
            pedido_preenchimento_associacao,
            SolicitacoesPreenchidas.id == pedido_preenchimento_associacao.c.preenchimento_id
        ).distinct().subquery()

        # Query base: solicitações APROVADAS que NÃO estão na subquery (não geraram pedido)
        query = SolicitacoesCompra.query.filter(
            SolicitacoesCompra.status_aprovacao == "Aprovado",
            ~SolicitacoesCompra.id.in_(subquery_pedidos)  # 🔴 FILTRO CORRETO
        )

        # Aplicar filtros
        if empresa:
            query = query.filter(SolicitacoesCompra.empresa == empresa)
        if usuario:
            query = query.filter(SolicitacoesCompra.usuario == usuario)
        if aplicacao_filter:
            query = query.filter(SolicitacoesCompra.aplicacao == aplicacao_filter)
        if status:
            query = query.filter(SolicitacoesCompra.status_aprovacao == status)
        if comprador_filter:
            if comprador_filter == 'Não atribuído':
                query = query.filter(
                    (SolicitacoesCompra.comprador_atribuido == None) | 
                    (SolicitacoesCompra.comprador_atribuido == '')
                )
            else:
                query = query.filter(SolicitacoesCompra.comprador_atribuido == comprador_filter)
        if data_inicio:
            query = query.filter(SolicitacoesCompra.data_solicitacao >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        if data_fim:
            query = query.filter(SolicitacoesCompra.data_solicitacao <= datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1))

        # Obter todas as solicitações
        solicitacoes = query.all()
        
        # Agrupar por aplicação E data completa (com hora)
        grupos = {}
        for s in solicitacoes:
            if s.aplicacao:
                # Criar chave única com aplicação + data completa (com hora)
                data_hora_formatada = s.data_solicitacao.strftime('%Y-%m-%d %H:%M')
                chave_grupo = f"{s.aplicacao}|{data_hora_formatada}"
                
                if chave_grupo not in grupos:
                    grupos[chave_grupo] = []
                
                # Verificar se a solicitação pode ser reprovada individualmente
                preenchimentos = SolicitacoesPreenchidas.query.filter_by(
                    solicitacao_id=s.id
                ).filter(
                    SolicitacoesPreenchidas.status != 'Rascunho'
                ).all()
                
                s.pode_reprovar_individual = len(preenchimentos) == 0
                s.preenchimentos_count = len(preenchimentos)
                grupos[chave_grupo].append(s)

        # Ordenar grupos pela data mais recente
        grupos_ordenados = sorted(grupos.items(), key=lambda x: x[1][0].data_solicitacao if x[1] else datetime.min, reverse=True)
        
        # Calcular paginação
        total_grupos = len(grupos_ordenados)
        total_paginas = (total_grupos + per_page - 1) // per_page
        inicio = (page - 1) * per_page
        fim = inicio + per_page
        
        grupos_paginados = grupos_ordenados[inicio:fim]

        # Listas de filtros — centralizadas e com cache de 5 min
        _filtros = get_filtros_listagem(apenas_aprovadas=True)
        empresas   = _filtros['empresas']
        usuarios   = _filtros['usuarios']
        aplicacoes = _filtros['aplicacoes']
        compradores = get_compradores()

        return render_template(
            'listar_solicitacoes.html',
            grupos_paginados=grupos_paginados,
            empresas=empresas,
            usuarios=usuarios,
            aplicacoes=aplicacoes,
            compradores=compradores,
            page=page,
            per_page=per_page,
            total_paginas=total_paginas,
            total_grupos=total_grupos,
            request_args=request.args
        )

    except Exception as e:
        _flash_erro_seguro('Erro ao carregar solicitações.', e, app.logger)
        return render_template(
            'listar_solicitacoes.html',
            grupos_paginados=[],
            empresas=[],
            usuarios=[],
            aplicacoes=[],
            compradores=[],
            page=1,
            per_page=10,
            total_paginas=1,
            total_grupos=0,
            request_args={}
        )
    
@routes_bp.route('/api/editar_quantidade_solicitacao/<int:id>', methods=['POST'])
def editar_quantidade_solicitacao(id):
    """Edita a quantidade de uma solicitação individual (requer senha 122004)"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        data = request.get_json()
        nova_quantidade = data.get('quantidade')
        senha = data.get('senha')
        
        # Validar senha administrativa (definida em ADMIN_EDIT_PASSWORD no .env)
        senha_admin = os.getenv('ADMIN_EDIT_PASSWORD', '')
        if not senha_admin or senha != senha_admin:
            return jsonify({
                'success': False,
                'message': 'Senha administrativa inválida!'
            }), 403
        
        # Validar quantidade
        if nova_quantidade is None:
            return jsonify({
                'success': False, 
                'message': 'Quantidade não informada'
            }), 400
        
        try:
            # 👇 CORREÇÃO: Converter para float
            nova_quantidade = float(nova_quantidade)
            if nova_quantidade <= 0:
                return jsonify({
                    'success': False, 
                    'message': 'Quantidade deve ser maior que zero'
                }), 400
        except ValueError:
            return jsonify({
                'success': False, 
                'message': 'Quantidade inválida. Use números com ponto (ex: 10.5)'
            }), 400
        
        # Buscar a solicitação
        solicitacao = SolicitacoesCompra.query.get_or_404(id)
        usuario = session['usuario']
        
        # Guardar quantidade anterior para log
        quantidade_anterior = solicitacao.quantidade
        
        # Verificar se a solicitação já tem preenchimentos
        preenchimentos = SolicitacoesPreenchidas.query.filter_by(
            solicitacao_id=solicitacao.id
        ).filter(
            SolicitacoesPreenchidas.status != 'Rascunho'
        ).count()
        
        # Se já tem preenchimentos enviados, não permite editar
        if preenchimentos > 0:
            return jsonify({
                'success': False, 
                'message': 'Não é possível editar quantidade: já existem cotações enviadas para esta solicitação.'
            }), 400
        
        # Atualizar quantidade
        solicitacao.quantidade = nova_quantidade
        
        # Adicionar observação sobre a alteração
        observacao = f"QUANTIDADE ALTERADA de {quantidade_anterior} para {nova_quantidade} - Usuário: {usuario}"
        if solicitacao.observacoes_col:
            solicitacao.observacoes_col += f"\n{observacao}"
        else:
            solicitacao.observacoes_col = observacao
        
        # Registrar log
        ip = request.remote_addr
        registrar_log(usuario, 'editar_quantidade_solicitacao', 
                     f'Solicitação ID {id} quantidade alterada de {quantidade_anterior} para {nova_quantidade}', ip)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Quantidade atualizada com sucesso!',
            'nova_quantidade': nova_quantidade
        })
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Erro ao editar quantidade da solicitação {id}: {str(e)}")
        return jsonify({
            'success': False, 
            'message': 'Erro interno. Tente novamente.'
        }), 500
    
#Tela Comprado
@routes_bp.route('/listar_solicitacoes_comprador', methods=['GET'])
def listar_solicitacoes_comprador():
    """Rota do menu comprador para exibir solicitações APROVADAS para preenchimento."""
    
    if 'usuario' not in session:
        flash('Acesso não autorizado', 'error')
        return redirect(url_for('routes_bp.login'))

    try:
        from sqlalchemy.orm import joinedload
        from sqlalchemy import exists, and_, not_
        
        # 🔴 OBTER O COMPRADOR LOGADO
        comprador_logado = session.get('usuario')
        
        # 🔴 SUBQUERY 1: IDs das solicitações que JÁ GERARAM PEDIDO
        subquery_pedidos = db.session.query(
            SolicitacoesPreenchidas.solicitacao_id
        ).join(
            pedido_preenchimento_associacao,
            SolicitacoesPreenchidas.id == pedido_preenchimento_associacao.c.preenchimento_id
        ).distinct().subquery()
        
        # SUBQUERY 2: IDs das solicitações que o comprador logado já finalizou.
        # Filtramos pelo usuario do preenchimento para não esconder solicitações
        # que outro comprador tenha preenchido (cada comprador vê só as suas).
        subquery_finalizados = db.session.query(
            SolicitacoesPreenchidas.solicitacao_id
        ).filter(
            SolicitacoesPreenchidas.status != 'Rascunho',
            SolicitacoesPreenchidas.usuario == comprador_logado
        ).distinct().subquery()
        
        # BUSCAR APENAS SOLICITAÇÕES:
        # 1. Aprovadas pelo gestor (status_aprovacao == 'Aprovado')
        # 2. Atribuídas ao comprador logado
        # 3. Que NÃO geraram pedido ainda
        # 4. Que NÃO têm preenchimentos finalizados (se tem 'Aguardando Aprovacao' ou 'Aprovado', some da lista)
        solicitacoes = SolicitacoesCompra.query.filter(
            SolicitacoesCompra.status_aprovacao == 'Aprovado',
            SolicitacoesCompra.comprador_atribuido == comprador_logado,
            ~SolicitacoesCompra.id.in_(subquery_pedidos),
            ~SolicitacoesCompra.id.in_(subquery_finalizados)
        ).options(
            joinedload(SolicitacoesCompra.material),
            joinedload(SolicitacoesCompra.preenchimentos_fornecidos)
        ).all()


        # Agrupar por aplicação + data
        grupos = {}
        for s in solicitacoes:
            aplic = s.aplicacao or (s.material.Aplicacao if s.material else 'Sem Aplicação')
            data_hora_str = s.data_solicitacao.strftime('%Y-%m-%d %H:%M')
            chave = f"{aplic} | {data_hora_str}"
            
            if chave not in grupos:
                grupos[chave] = []
            grupos[chave].append(s)
        
        # Separar grupos com e sem rascunho
        grupos_com_rascunho = []
        grupos_sem_rascunho = []
        
        for chave, items in grupos.items():
            tem_rascunho = False
            for s in items:
                if s.preenchimentos_fornecidos:
                    for p in s.preenchimentos_fornecidos:
                        if p.status == 'Rascunho':
                            tem_rascunho = True
                            break
                if tem_rascunho:
                    break
            
            if tem_rascunho:
                grupos_com_rascunho.append((chave, items))
            else:
                grupos_sem_rascunho.append((chave, items))

        # 🔴 LISTA DE COMPRADORES: APENAS O LOGADO
        compradores = [comprador_logado]

        # Listas de filtros — centralizadas e com cache de 5 min (filtradas pelo comprador logado)
        _filtros = get_filtros_listagem(apenas_aprovadas=False, comprador=comprador_logado)
        empresas   = _filtros['empresas']
        usuarios   = _filtros['usuarios']
        aplicacoes = _filtros['aplicacoes']


        return render_template(
            'listar_solicitacoes_comprador.html',
            grupos_com_rascunho=grupos_com_rascunho,
            grupos_sem_rascunho=grupos_sem_rascunho,
            empresas=sorted(empresas),
            usuarios=sorted(usuarios),
            aplicacoes=sorted(aplicacoes),
            compradores=compradores,
            comprador_atual=comprador_logado,
            titulo_pagina="Solicitações Aprovadas - Preencher Cotação"
        )

    except SQLAlchemyError as e:
        _flash_erro_seguro('Erro ao carregar solicitações.', e, app.logger)
        import traceback; traceback.print_exc()
        return render_template(
            'listar_solicitacoes_comprador.html',
            grupos_com_rascunho=[],
            grupos_sem_rascunho=[],
            empresas=[],
            usuarios=[],
            aplicacoes=[],
            compradores=[session.get('usuario', '')],
            comprador_atual=session.get('usuario', ''),
            titulo_pagina="Solicitações Aprovadas - Preencher Cotação"
        )

@routes_bp.route('/aprovar_solicitacao', methods=['GET'])
def aprovar_solicitacao():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        # Importar or_ no início da função
        from sqlalchemy import or_
        
        # Obter parâmetros de filtro
        filtro_status = request.args.get('filtro_status', 'pendentes')
        aplicacao_filtro = request.args.get('aplicacao', '')
        
        
        # Importar joinedload para eager loading
        from sqlalchemy.orm import joinedload
        
        # Query base com eager loading do material
        query = SolicitacoesCompra.query.options(joinedload(SolicitacoesCompra.material))

        # ── Mapa de status → filtro SQL ──────────────────────────────────────
        # Adicionar novos status aqui, sem tocar no restante da lógica.
        _subq_com_preenchimento_finalizado = lambda: (
            db.session.query(SolicitacoesPreenchidas.solicitacao_id).filter(
                SolicitacoesPreenchidas.status.in_(
                    ['Aprovado', 'Reprovado', 'Entregue', 'Em Processamento']
                )
            ).distinct()
        )

        _STATUS_QUERY_MAP = {
            'pendentes': lambda q: q.filter(
                or_(
                    SolicitacoesCompra.status_aprovacao.is_(None),
                    SolicitacoesCompra.status_aprovacao == '',
                    SolicitacoesCompra.status_aprovacao == 'Pendente',
                )
            ),
            'abertas': lambda q: q.filter(
                ~SolicitacoesCompra.id.in_(_subq_com_preenchimento_finalizado())
            ),
            'aprovadas': lambda q: q.filter(
                SolicitacoesCompra.status_aprovacao == 'Aprovado'
            ),
            'reprovadas': lambda q: q.filter(
                SolicitacoesCompra.status_aprovacao == 'Reprovado'
            ),
            'preenchidas': lambda q: q.filter(
                SolicitacoesCompra.id.in_(_subq_com_preenchimento_finalizado())
            ),
        }

        aplicar_filtro = _STATUS_QUERY_MAP.get(filtro_status)
        solicitacoes = (aplicar_filtro(query) if aplicar_filtro else query).all()

        # Buscar aplicações em UMA ÚNICA consulta
        # Buscar aplicações de materiais
        aplicacoes_materiais = db.session.query(
            Materiais.Aplicacao
        ).filter(
            Materiais.Aplicacao.isnot(None),
            Materiais.Aplicacao != ''
        ).distinct().all()
        
        # Buscar aplicações de solicitações
        aplicacoes_solicitacoes = db.session.query(
            SolicitacoesCompra.aplicacao
        ).filter(
            SolicitacoesCompra.aplicacao.isnot(None),
            SolicitacoesCompra.aplicacao != ''
        ).distinct().all()
        
        # Buscar aplicações gerais
        aplicacoes_gerais = db.session.query(
            SolicitacoesCompra.aplicacao_geral
        ).filter(
            SolicitacoesCompra.aplicacao_geral.isnot(None),
            SolicitacoesCompra.aplicacao_geral != ''
        ).distinct().all()
        
        # Combinar todas as aplicações
        aplicacoes_unicas = set()
        for aplicacao in aplicacoes_materiais:
            if aplicacao[0]:
                aplicacoes_unicas.add(aplicacao[0].strip())
        for aplicacao in aplicacoes_solicitacoes:
            if aplicacao[0]:
                aplicacoes_unicas.add(aplicacao[0].strip())
        for aplicacao in aplicacoes_gerais:
            if aplicacao[0]:
                aplicacoes_unicas.add(aplicacao[0].strip())
        
        aplicacoes_unicas = sorted(aplicacoes_unicas)
        
        print(f"🔍 DEBUG - Total de aplicações únicas encontradas: {len(aplicacoes_unicas)}")

        # Aplicar filtro por aplicação se especificado
        if aplicacao_filtro:
            solicitacoes = [s for s in solicitacoes if 
                           (s.aplicacao and aplicacao_filtro.lower() in s.aplicacao.lower()) or
                           (s.aplicacao_geral and aplicacao_filtro.lower() in s.aplicacao_geral.lower()) or
                           (s.material and s.material.Aplicacao and aplicacao_filtro.lower() in s.material.Aplicacao.lower())]

        # Usuários e empresas — via ler_senhas() que já tem cache de 60 s
        _senhas_data = ler_senhas()
        usuarios  = sorted({d['usuario_base'] for d in _senhas_data.values()})
        empresas  = sorted({d['empresa'] for d in _senhas_data.values() if d.get('empresa')})
        
        if not solicitacoes:
            flash('Nenhuma solicitação encontrada com os filtros aplicados.', 'info')
        
        return render_template('aprovar_solicitacao.html', 
                            solicitacoes=solicitacoes,
                            empresas=empresas,
                            usuarios=usuarios,
                            aplicacoes_unicas=aplicacoes_unicas,
                            filtro_status=filtro_status,
                            aplicacao_filtro=aplicacao_filtro)
        
    except SQLAlchemyError as e:
        app.logger.error('aprovar_solicitacao – erro de banco', exc_info=True)
        _flash_erro_seguro('Erro ao carregar solicitações.', e, app.logger)
        return redirect(url_for('routes_bp.buscar_material'))
    except Exception as e:
        app.logger.error('aprovar_solicitacao – erro inesperado', exc_info=True)
        _flash_erro_seguro('Erro inesperado ao carregar solicitações.', e, app.logger)
        return redirect(url_for('routes_bp.buscar_material'))
    
    
@routes_bp.route('/debug_ultimas_solicitacoes')
def debug_ultimas_solicitacoes():
    """Debug das últimas solicitações criadas"""
    import os as _os
    if _os.getenv("FLASK_ENV", "production").lower() not in ("development", "dev"):
        abort(404)
    if 'usuario' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        # Buscar as últimas 5 solicitações
        solicitacoes = SolicitacoesCompra.query.order_by(SolicitacoesCompra.id.desc()).limit(5).all()
        
        resultado = []
        for sol in solicitacoes:
            resultado.append({
                'id': sol.id,
                'cod_material': sol.cod_material,
                'descricao_material': sol.material.DescricaoMaterial if sol.material else 'N/A',
                'especificacao': sol.especificacao,
                'quantidade': sol.quantidade,
                'empresa': sol.empresa,
                'usuario': sol.usuario,
                'data_solicitacao': sol.data_solicitacao.isoformat() if sol.data_solicitacao else None,
                'status_aprovacao': str(sol.status_aprovacao),
                'ativo': sol.ativo,
                'nome_ativo': sol.nome_ativo,
                'prioridade': sol.prioridade
            })
        
        return jsonify({
            'total_solicitacoes': len(solicitacoes),
            'solicitacoes': resultado
        })
    except Exception as e:
        app.logger.error(f'API error: {type(e).__name__}', exc_info=False)
        return jsonify({'error': 'Erro interno. Tente novamente.'}), 500
    
@routes_bp.route('/debug_solicitacoes_status')
def debug_solicitacoes_status():
    """Debug completo do status das solicitações"""
    import os as _os
    if _os.getenv("FLASK_ENV", "production").lower() not in ("development", "dev"):
        abort(404)
    if 'usuario' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        # Buscar as últimas 20 solicitações
        solicitacoes = SolicitacoesCompra.query.order_by(SolicitacoesCompra.id.desc()).limit(20).all()
        
        resultado = []
        for sol in solicitacoes:
            resultado.append({
                'id': sol.id,
                'cod_material': sol.cod_material,
                'especificacao': sol.especificacao[:50] + '...' if sol.especificacao else '',
                'status_aprovacao': str(sol.status_aprovacao),  # Converter para string para ver NULL
                'status_aprovacao_raw': sol.status_aprovacao,
                'data_solicitacao': sol.data_solicitacao.isoformat() if sol.data_solicitacao else None,
                'usuario': sol.usuario,
                'is_null': sol.status_aprovacao is None,
                'is_empty': sol.status_aprovacao == '',
                'is_pendente': sol.status_aprovacao == 'Pendente'
            })
        
        # Contar por status
        counts = {
            'total': len(solicitacoes),
            'null': len([s for s in resultado if s['is_null']]),
            'empty': len([s for s in resultado if s['is_empty']]),
            'pendente': len([s for s in resultado if s['is_pendente']]),
            'aprovado': len([s for s in resultado if s['status_aprovacao'] == 'Aprovado']),
            'reprovado': len([s for s in resultado if s['status_aprovacao'] == 'Reprovado']),
        }
        
        return jsonify({
            'counts': counts,
            'solicitacoes': resultado
        })
    except Exception as e:
        app.logger.error(f'API error: {type(e).__name__}', exc_info=False)
        return jsonify({'error': 'Erro interno. Tente novamente.'}), 500

@routes_bp.route('/debug_ultima_solicitacao')
def debug_ultima_solicitacao():
    """Debug da última solicitação criada"""
    import os as _os
    if _os.getenv("FLASK_ENV", "production").lower() not in ("development", "dev"):
        abort(404)
    if 'usuario' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        ultima_solicitacao = SolicitacoesCompra.query.order_by(SolicitacoesCompra.id.desc()).first()
        
        if not ultima_solicitacao:
            return jsonify({'message': 'Nenhuma solicitação encontrada'})
        
        resultado = {
            'id': ultima_solicitacao.id,
            'cod_material': ultima_solicitacao.cod_material,
            'especificacao': ultima_solicitacao.especificacao,
            'status_aprovacao': str(ultima_solicitacao.status_aprovacao),
            'status_aprovacao_is_none': ultima_solicitacao.status_aprovacao is None,
            'data_solicitacao': ultima_solicitacao.data_solicitacao.isoformat() if ultima_solicitacao.data_solicitacao else None,
            'usuario': ultima_solicitacao.usuario,
            'empresa': ultima_solicitacao.empresa
        }
        
        return jsonify(resultado)
    except Exception as e:
        app.logger.error(f'API error: {type(e).__name__}', exc_info=False)
        return jsonify({'error': 'Erro interno. Tente novamente.'}), 500
    
@routes_bp.route('/debug_solicitacoes_recentes')
def debug_solicitacoes_recentes():
    """Rota para debug - mostra as últimas 10 solicitações criadas"""
    import os as _os
    if _os.getenv("FLASK_ENV", "production").lower() not in ("development", "dev"):
        abort(404)
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        solicitacoes = SolicitacoesCompra.query.order_by(SolicitacoesCompra.id.desc()).limit(10).all()
        
        resultado = []
        for sol in solicitacoes:
            resultado.append({
                'id': sol.id,
                'cod_material': sol.cod_material,
                'especificacao': sol.especificacao,
                'status_aprovacao': sol.status_aprovacao,
                'data_solicitacao': sol.data_solicitacao.isoformat() if sol.data_solicitacao else None,
                'usuario': sol.usuario
            })
        
        return jsonify({
            'total': len(solicitacoes),
            'solicitacoes': resultado
        })
    except Exception as e:
        app.logger.error(f'API error: {type(e).__name__}', exc_info=False)
        return jsonify({'error': 'Erro interno. Tente novamente.'}), 500
    
@routes_bp.route('/debug_solicitacoes_detalhado')
def debug_solicitacoes_detalhado():
    """Debug detalhado de TODAS as solicitações"""
    import os as _os
    if _os.getenv("FLASK_ENV", "production").lower() not in ("development", "dev"):
        abort(404)
    if 'usuario' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        solicitacoes = SolicitacoesCompra.query.order_by(SolicitacoesCompra.id.desc()).limit(10).all()
        
        resultado = []
        for sol in solicitacoes:
            # Verificar valores exatos
            status_raw = sol.status_aprovacao
            resultado.append({
                'id': sol.id,
                'cod_material': sol.cod_material,
                'especificacao': sol.especificacao[:50] + '...' if sol.especificacao else '',
                'status_aprovacao_raw': status_raw,
                'status_aprovacao_repr': repr(status_raw),
                'status_aprovacao_type': type(status_raw).__name__,
                'is_none': status_raw is None,
                'is_empty_string': status_raw == '',
                'is_pendente': status_raw == 'Pendente',
                'data_solicitacao': sol.data_solicitacao.isoformat() if sol.data_solicitacao else None,
                'usuario': sol.usuario
            })
        
        return jsonify({
            'total': len(solicitacoes),
            'solicitacoes': resultado
        })
    except Exception as e:
        app.logger.error(f'API error: {type(e).__name__}', exc_info=False)
        return jsonify({'error': 'Erro interno. Tente novamente.'}), 500

@routes_bp.route('/debug_verificar_coluna_status')
def debug_verificar_coluna_status():
    """Verificar se a coluna existe e seus valores"""
    import os as _os
    if _os.getenv("FLASK_ENV", "production").lower() not in ("development", "dev"):
        abort(404)
    if 'usuario' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    try:
        # Verificar estrutura da tabela
        result = db.engine.execute("PRAGMA table_info(SolicitacoesCompra)").fetchall()
        colunas = [col[1] for col in result]
        
        # Verificar valores diretamente via SQL
        sql_result = db.engine.execute("""
            SELECT id, status_aprovacao, 
                   typeof(status_aprovacao) as tipo,
                   length(status_aprovacao) as tamanho
            FROM SolicitacoesCompra 
            ORDER BY id DESC 
            LIMIT 10
        """).fetchall()
        
        sql_data = []
        for row in sql_result:
            sql_data.append({
                'id': row[0],
                'status_aprovacao': row[1],
                'tipo': row[2],
                'tamanho': row[3] if row[3] is not None else 'NULL'
            })
        
        return jsonify({
            'colunas_tabela': colunas,
            'existe_status_aprovacao': 'status_aprovacao' in colunas,
            'dados_sql': sql_data
        })
    except Exception as e:
        app.logger.error(f'API error: {type(e).__name__}', exc_info=False)
        return jsonify({'error': 'Erro interno. Tente novamente.'}), 500
    
def migrate_solicitacoes_compra():
    try:
        # Usar o mesmo banco de dados do Flask-SQLAlchemy
        DATABASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ComparasDB.db')
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        # Verificar se a tabela SolicitacoesCompra existe
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='SolicitacoesCompra'")
        table_exists = cursor.fetchone()

        if not table_exists:
            # Criar a tabela com todas as colunas do modelo
            cursor.execute("""
                CREATE TABLE SolicitacoesCompra (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    cod_material INTEGER NOT NULL,
                    especificacao TEXT NOT NULL,
                    quantidade INTEGER NOT NULL,
                    unidade_medida TEXT NOT NULL DEFAULT 'Unidade',
                    aplicacao TEXT,
                    empresa TEXT NOT NULL,
                    data_solicitacao DATETIME NOT NULL,
                    usuario TEXT NOT NULL,
                    foto_path TEXT,
                    FOREIGN KEY (cod_material) REFERENCES Materiais(CodMaterial)
                )
            """)
            logging.info("Created SolicitacoesCompra table with all columns")
        else:
            # Verificar colunas atuais
            cursor.execute("PRAGMA table_info(SolicitacoesCompra)")
            columns = [col[1] for col in cursor.fetchall()]
            logging.info(f"Current columns in SolicitacoesCompra: {columns}")

            # Adicionar colunas faltantes, se necessário
            if 'unidade_medida' not in columns:
                cursor.execute("ALTER TABLE SolicitacoesCompra ADD COLUMN unidade_medida TEXT NOT NULL DEFAULT 'Unidade'")
                logging.info("Added unidade_medida column to SolicitacoesCompra")
            
            if 'aplicacao' not in columns:
                cursor.execute("ALTER TABLE SolicitacoesCompra ADD COLUMN aplicacao TEXT")
                logging.info("Added aplicacao column to SolicitacoesCompra")
            
            if 'foto_path' not in columns:
                cursor.execute("ALTER TABLE SolicitacoesCompra ADD COLUMN foto_path TEXT")
                logging.info("Added foto_path column to SolicitacoesCompra")

        conn.commit()
        logging.info("Database migration for SolicitacoesCompra completed successfully")

        # Verificar esquema atualizado
        cursor.execute("PRAGMA table_info(SolicitacoesCompra)")
        updated_columns = [col[1] for col in cursor.fetchall()]
        logging.info(f"Updated columns in SolicitacoesCompra: {updated_columns}")
    except sqlite3.Error as e:
        logging.error(f"Error during migration: {str(e)}")
        raise
    finally:
        conn.close()


def extrair_desconto_da_condicao(condicao_pagamento):
    """Extrai o valor do desconto da string condicao_pagamento"""
    if not condicao_pagamento:
        return 0.0
    
    import re
    # Procura padrão "DESCONTO: R$ 50,00"
    match = re.search(r'DESCONTO:\s*R?\$?\s*(\d+(?:[.,]\d+)?)', condicao_pagamento, re.IGNORECASE)
    if match:
        return float(match.group(1).replace(',', '.'))
    
    return 0.0

@routes_bp.route('/preencher_solicitacao/<int:id>', methods=['GET', 'POST'])
def preencher_solicitacao(id):
    if 'usuario' not in session:
        flash('Acesso não autorizado', 'error')
        return redirect(url_for('routes_bp.login'))
    
    # === FUNÇÃO AUXILIAR PARA CONVERTER QUANTIDADE ===
    def converter_quantidade_para_float(quantidade):
        """Converte quantidade para float, seja string ou número"""
        if quantidade is None:
            return 0.0
        if isinstance(quantidade, (int, float)):
            return float(quantidade)
        try:
            qtd_str = str(quantidade).strip().replace(',', '.')
            import re
            qtd_str = re.sub(r'[^\d.]', '', qtd_str)
            partes = qtd_str.split('.')
            if len(partes) > 2:
                qtd_str = partes[0] + '.' + ''.join(partes[1:])
            return float(qtd_str) if qtd_str else 0.0
        except (ValueError, TypeError):
            return 0.0
    
    # === FUNÇÃO AUXILIAR PARA CONVERSÃO DE VALORES ===
    def parse_br_currency_final(value_str):
        if not value_str or value_str == '':
            return 0.0
        
        valor = str(value_str)
        valor = valor.strip()
        valor = valor.replace('R$', '').replace('r$', '').strip()
        
        import re
        valor = re.sub(r'[^\d,\.\-]', '', valor)
        
        if not valor or valor == '-':
            return 0.0
        
        negativo = False
        if valor.startswith('-'):
            negativo = True
            valor = valor[1:]
        
        if valor == '' or valor == '0' or valor == '0,00' or valor == '0.00':
            return 0.0
        
        try:
            if ',' not in valor and '.' not in valor:
                resultado = float(valor)
            else:
                num_virgulas = valor.count(',')
                num_pontos = valor.count('.')
                
                if num_pontos >= 1 and num_virgulas == 1:
                    valor_sem_pontos = valor.replace('.', '')
                    valor_final = valor_sem_pontos.replace(',', '.')
                    resultado = float(valor_final)
                elif num_virgulas == 1 and num_pontos == 0:
                    partes = valor.split(',')
                    if len(partes) == 2:
                        resultado = float(f"{partes[0]}.{partes[1]}")
                    else:
                        resultado = float(valor.replace(',', '.'))
                elif num_pontos == 1 and num_virgulas == 0:
                    resultado = float(valor)
                elif num_virgulas >= 1 and num_pontos == 1:
                    valor_sem_virgulas = valor.replace(',', '')
                    resultado = float(valor_sem_virgulas)
                else:
                    valor_limpo = valor.replace('.', '').replace(',', '')
                    resultado = float(valor_limpo)
        except ValueError:
            resultado = 0.0
        
        if negativo:
            resultado = -resultado
        
        return resultado
    
    # === MODO GRUPO ===
    grupo_ids_param = request.args.get('grupo_ids', '')
    grupo_ids = []
    modo_grupo = False
    solicitacoes_grupo = []
    
    if grupo_ids_param:
        try:
            grupo_ids = [int(x) for x in grupo_ids_param.split(',') if x.strip().isdigit()]
            if grupo_ids:
                solicitacoes_grupo = SolicitacoesCompra.query.filter(
                    SolicitacoesCompra.id.in_(grupo_ids)
                ).all()
                modo_grupo = len(solicitacoes_grupo) > 1
        except:
            flash('IDs de grupo inválidos.', 'error')
            return redirect(url_for('routes_bp.listar_solicitacoes_comprador'))
    
    solicitacao = SolicitacoesCompra.query.get_or_404(id)
    
    if modo_grupo and id not in grupo_ids:
        if solicitacao not in solicitacoes_grupo:
            solicitacoes_grupo.append(solicitacao)
            grupo_ids.append(id)
    
    todas_solicitacoes = solicitacoes_grupo if modo_grupo else [solicitacao]
    todas_solicitacoes_ids = [s.id for s in todas_solicitacoes]
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if not action:
            if 'salvar_rascunho' in request.form:
                action = 'salvar_rascunho'
            elif 'salvar_finalizar' in request.form:
                action = 'salvar_finalizar'
        
        is_rascunho = action == 'salvar_rascunho'
        is_finalizar = action == 'salvar_finalizar'
        
        if not (is_rascunho or is_finalizar):
            flash('Ação inválida. Selecione "Salvar Rascunho" ou "Finalizar".', 'error')
            return redirect(request.url)
        
        try:
            usuario = session['usuario']
            
            # ============================================
            # PROCESSAR FORNECEDORES REMOVIDOS
            # ============================================
            fornecedores_removidos_json = request.form.get('fornecedores_removidos', '[]')
            fornecedores_ids_remover = []
            
            if fornecedores_removidos_json and fornecedores_removidos_json != '[]':
                try:
                    import json
                    fornecedores_ids_remover = json.loads(fornecedores_removidos_json)
                    
                    if fornecedores_ids_remover:
                        print(f"🚨 Processando exclusão de fornecedores: {fornecedores_ids_remover}")
                        
                        for fornecedor_id in fornecedores_ids_remover:
                            if fornecedor_id and fornecedor_id != '':
                                preenchimentos = SolicitacoesPreenchidas.query.filter(
                                    SolicitacoesPreenchidas.solicitacao_id.in_(todas_solicitacoes_ids),
                                    SolicitacoesPreenchidas.fornecedor_id == int(fornecedor_id),
                                    SolicitacoesPreenchidas.status == 'Rascunho'
                                ).all()
                                
                                for preenchimento in preenchimentos:
                                    HistoricoDescontos.query.filter_by(
                                        preenchimento_id=preenchimento.id
                                    ).delete()
                                    db.session.delete(preenchimento)
                                    print(f"✅ Preenchimento ID {preenchimento.id} do fornecedor {fornecedor_id} excluído")
                        
                        # NÃO commit aqui — será feito junto com os novos preenchimentos no final
                        print(f"✅ Remoções marcadas na sessão (commit no final do POST)")
                        
                except Exception as e:
                    db.session.rollback()
                    print(f"❌ Erro ao remover cotações: {str(e)}")
            
            # === Captura de dados em listas ===
            fornecedores_ids = request.form.getlist('fornecedor_id[]')
            valores_frete = request.form.getlist('valor_frete[]')
            prazos = request.form.getlist('prazo_entrega[]')
            condicoes = request.form.getlist('condicao_pagamento[]')
            observacoes = request.form.getlist('observacao[]')
            preenchimento_ids = request.form.getlist('preenchimento_id[]')
            pdf_files = request.files.getlist('pdf_file[]')
            
            # CAPTURAR DESCONTOS
            descontos = request.form.getlist('desconto[]')
            
            # === Captura de dados dos materiais ===
            todos_valores_unitarios = request.form.getlist('valor_unitario[]')
            todos_ids_solicitacao = request.form.getlist('solicitacao_id[]')
            todas_marcas = request.form.getlist('marca[]')
            
            logging.debug(f"Descontos recebidos: {len(descontos)} itens")
            print(f"📋 Total valores unitários: {len(todos_valores_unitarios)}")
            
            # === Processar cada cotação ===
            cotacoes_salvas = 0
            total_cotacoes = 0
            erros_validacao = []
            
            for idx, fornecedor_id in enumerate(fornecedores_ids):
                fornecedor_id = fornecedor_id.strip()
                if not fornecedor_id:
                    continue
                
                # PULAR se este fornecedor está na lista de removidos
                if fornecedor_id in fornecedores_ids_remover:
                    print(f"⏭️ Pulando fornecedor {fornecedor_id} (marcado para exclusão)")
                    continue
                
                total_cotacoes += 1
                
                # === Campos da cotação atual ===
                vf_str = valores_frete[idx] if idx < len(valores_frete) else '0'
                prazo = prazos[idx] if idx < len(prazos) else ''
                condicao_original = condicoes[idx] if idx < len(condicoes) else ''
                obs = observacoes[idx] if idx < len(observacoes) else ''
                preenchimento_id = preenchimento_ids[idx] if idx < len(preenchimento_ids) and preenchimento_ids[idx] else None
                
                # OBTER O DESCONTO
                desconto_str = descontos[idx] if idx < len(descontos) else '0'
                valor_desconto = parse_br_currency_final(desconto_str)
                
                # CONCATENAR O DESCONTO NA CONDIÇÃO DE PAGAMENTO
                condicao_final = condicao_original if condicao_original else ""
                if valor_desconto > 0:
                    if 'DESCONTO:' not in condicao_original.upper():
                        if condicao_original:
                            condicao_final = f"{condicao_original} | DESCONTO: R$ {valor_desconto:.2f}".replace('.', ',')
                        else:
                            condicao_final = f"DESCONTO: R$ {valor_desconto:.2f}".replace('.', ',')
                    else:
                        import re
                        condicao_final = re.sub(
                            r'DESCONTO:\s*R?\$?\s*\d+(?:[.,]\d+)?',
                            f'DESCONTO: R$ {valor_desconto:.2f}'.replace('.', ','),
                            condicao_original,
                            flags=re.IGNORECASE
                        )
                
                print(f"💰 Desconto: R$ {valor_desconto:.2f} -> Condição final: '{condicao_final}'")
                
                # 🔴 VALIDAÇÃO PARA FINALIZAR - REMOVIDA VALIDAÇÃO DA CONDIÇÃO DE PAGAMENTO
                if is_finalizar:
                    if not prazo:
                        erros_validacao.append(f'Cotação {idx+1}: Prazo de entrega é obrigatório.')
                    # Condição de pagamento NÃO é mais obrigatória - pode ser vazia
                    if not condicao_final:
                        condicao_final = ""  # Garante string vazia
                
                # CONVERSÃO DO FRETE
                valor_frete = parse_br_currency_final(vf_str)
                
                # === Processar cada material desta cotação ===
                num_materiais_por_cotacao = len(todas_solicitacoes)
                inicio_idx = idx * num_materiais_por_cotacao
                
                if inicio_idx >= len(todos_valores_unitarios):
                    erros_validacao.append(f'Cotação {idx+1}: Dados dos materiais incompletos.')
                    continue
                
                primeiro_material_desta_cotacao = True
                materiais_processados = 0
                
                for i, sol in enumerate(todas_solicitacoes):
                    material_idx = inicio_idx + i
                    
                    if material_idx >= len(todos_valores_unitarios):
                        erros_validacao.append(f'Cotação {idx+1}, Material {i+1}: Valor unitário faltando.')
                        continue
                    
                    valor_unitario_str = todos_valores_unitarios[material_idx].strip() if material_idx < len(todos_valores_unitarios) else ''
                    
                    # Tratar valor vazio como 0
                    if not valor_unitario_str or valor_unitario_str == '':
                        valor_unitario = 0.0
                    else:
                        valor_unitario = parse_br_currency_final(valor_unitario_str)
                    
                    # Converter quantidade para float
                    quantidade_float = converter_quantidade_para_float(sol.quantidade)
                    valor_total = round(valor_unitario * quantidade_float, 2)
                    
                    if valor_unitario <= 0:
                        print(f"   ⚠️ Material {i+1}: Valor unitário = {valor_unitario}")
                    
                    # CAPTURAR A MARCA PARA ESTE MATERIAL
                    marca = ''
                    if material_idx < len(todas_marcas):
                        marca = todas_marcas[material_idx].strip()
                    
                    print(f"   📝 Material {i+1}: Marca = '{marca}', Valor = {valor_unitario}, Qtd = {quantidade_float}")
                    
                    # BUSCAR PREENCHIMENTO
                    preenchimento = None
                    
                    if primeiro_material_desta_cotacao and preenchimento_id and preenchimento_id != '':
                        try:
                            preenchimento = SolicitacoesPreenchidas.query.get(int(preenchimento_id))
                            print(f"   🔍 Busca por ID {preenchimento_id}: {'Encontrado' if preenchimento else 'Não encontrado'}")
                        except:
                            pass
                    
                    if not preenchimento:
                        preenchimento = SolicitacoesPreenchidas.query.filter_by(
                            solicitacao_id=sol.id,
                            fornecedor_id=int(fornecedor_id),
                            status='Rascunho'
                        ).first()
                        print(f"   🔍 Busca por fornecedor {fornecedor_id} + solicitação {sol.id}: {'Encontrado' if preenchimento else 'Não encontrado'}")
                    
                    if not preenchimento:
                        # CRIAR NOVO PREENCHIMENTO
                        preenchimento = SolicitacoesPreenchidas(
                            solicitacao_id=sol.id,
                            fornecedor_id=int(fornecedor_id),
                            valor_unitario=valor_unitario,
                            valor_total=valor_total,
                            valor_frete=valor_frete if valor_frete > 0 else None,
                            prazo_entrega=prazo,
                            condicao_pagamento=condicao_final if condicao_final else "",  # Permite vazio
                            observacoes=obs,
                            data_preenchimento=get_local_time(),
                            usuario=usuario,
                            status='Rascunho' if is_rascunho else 'Aguardando Aprovacao',
                            marca=marca
                        )
                        db.session.add(preenchimento)
                        print(f"   ✅ Novo preenchimento criado para solicitação {sol.id} com marca '{marca}' e condição '{condicao_final}'")
                        
                    else:
                        # ATUALIZAR PREENCHIMENTO EXISTENTE
                        print(f"   🔄 Atualizando preenchimento ID {preenchimento.id}")
                        
                        # Verificar mudanças para histórico
                        valor_unitario_anterior = float(preenchimento.valor_unitario) if preenchimento.valor_unitario else 0.0
                        valor_frete_anterior = float(preenchimento.valor_frete) if preenchimento.valor_frete else 0.0
                        
                        valor_unitario_mudou = abs(valor_unitario_anterior - valor_unitario) > 0.001
                        
                        valor_frete_mudou = False
                        if valor_frete > 0:
                            if valor_frete_anterior is None or abs(valor_frete_anterior - valor_frete) > 0.001:
                                valor_frete_mudou = True
                        else:
                            if valor_frete_anterior is not None and valor_frete_anterior > 0:
                                valor_frete_mudou = True
                        
                        if valor_unitario_mudou or valor_frete_mudou:
                            historico = HistoricoDescontos(
                                preenchimento_id=preenchimento.id,
                                valor_unitario_anterior=valor_unitario_anterior,
                                valor_unitario_novo=valor_unitario,
                                valor_frete_anterior=valor_frete_anterior if valor_frete_anterior != 0 else None,
                                valor_frete_novo=valor_frete if valor_frete > 0 else None,
                                data_alteracao=get_local_time(),
                                usuario=usuario
                            )
                            db.session.add(historico)
                            print(f"   📝 Histórico de desconto registrado")
                        
                        # Atualizar dados
                        preenchimento.valor_unitario = valor_unitario
                        preenchimento.valor_total = valor_total
                        preenchimento.valor_frete = valor_frete if valor_frete > 0 else None
                        preenchimento.prazo_entrega = prazo
                        preenchimento.condicao_pagamento = condicao_final if condicao_final else ""  # Permite vazio
                        preenchimento.observacoes = obs
                        preenchimento.status = 'Rascunho' if is_rascunho else 'Aguardando Aprovacao'
                        preenchimento.data_preenchimento = get_local_time()
                        preenchimento.usuario = usuario
                        preenchimento.marca = marca
                        
                        print(f"   ✅ Marca '{marca}' e condição '{condicao_final}' atualizadas no preenchimento {preenchimento.id}")
                    
                    # PDF apenas para o primeiro material da cotação
                    if idx < len(pdf_files) and i == 0 and pdf_files[idx] and pdf_files[idx].filename:
                        pdf_file = pdf_files[idx]
                        if allowed_file(pdf_file.filename, {'pdf'}):
                            timestamp = get_local_time().strftime('%Y%m%d_%H%M%S')
                            safe_name = secure_filename(pdf_file.filename)
                            unique_name = f"cotacao_{preenchimento.id}_{timestamp}_{safe_name}"
                            path = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
                            pdf_file.save(path)
                            preenchimento.pdf_path = unique_name
                            print(f"   📎 PDF salvo: {unique_name}")
                    
                    primeiro_material_desta_cotacao = False
                    materiais_processados += 1
                
                cotacoes_salvas += 1
                print(f"✅ Cotação {idx+1} processada com sucesso - {materiais_processados} materiais")
            
            # Verificar se houve erros de validação
            if erros_validacao:
                for erro in erros_validacao:
                    flash(erro, 'error')
                return redirect(request.url)
            
            if total_cotacoes == 0:
                flash('É necessário preencher pelo menos uma cotação.', 'error')
                return redirect(request.url)
            
            db.session.commit()
            print(f"💾 COMMIT REALIZADO! {cotacoes_salvas} materiais em {total_cotacoes} cotações")
            
            if is_rascunho:
                flash(f'{cotacoes_salvas} material(is) em {total_cotacoes} cotação(ões) salvo(s) como rascunho!', 'success')
            else:
                flash(f'{cotacoes_salvas} material(is) em {total_cotacoes} cotação(ões) finalizado(s) e enviado(s) para aprovação!', 'success')
            
            return redirect(url_for('routes_bp.listar_solicitacoes_comprador'))
        
        except Exception as e:
            db.session.rollback()
            _flash_erro_seguro('Erro ao salvar.', e, app.logger)
            import traceback
            print(f"❌ Erro completo: {traceback.format_exc()}")
            return redirect(request.url)
    
    # === GET: Carregar rascunhos ===
    if modo_grupo:
        cotacoes_raw = SolicitacoesPreenchidas.query.filter(
            SolicitacoesPreenchidas.solicitacao_id.in_(grupo_ids),
            SolicitacoesPreenchidas.status == 'Rascunho'
        ).order_by(SolicitacoesPreenchidas.fornecedor_id, SolicitacoesPreenchidas.solicitacao_id).all()
    else:
        cotacoes_raw = SolicitacoesPreenchidas.query.filter_by(
            solicitacao_id=id,
            status='Rascunho'
        ).order_by(SolicitacoesPreenchidas.fornecedor_id).all()
    
    # Organizar cotações por fornecedor
    cotacoes_por_fornecedor = {}
    for c in cotacoes_raw:
        if c.fornecedor_id not in cotacoes_por_fornecedor:
            cotacoes_por_fornecedor[c.fornecedor_id] = []
        cotacoes_por_fornecedor[c.fornecedor_id].append(c)
    
    # Criar estrutura para o template
    cotacoes_estruturadas = []
    
    for fornecedor_id, cotacoes_fornecedor in cotacoes_por_fornecedor.items():
        if not cotacoes_fornecedor:
            continue
        
        cotacao_ref = cotacoes_fornecedor[0]
        
        # Buscar dados do fornecedor
        fornecedor_info = {
            'nome_fantasia': '',
            'cnpj': '',
            'telefone': '',
            'endereco': ''
        }
        
        try:
            conn = get_db_connection(DB_PATH_FORNECEDORES)
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT nome_fantasia, cnpj, telefone, endereco, bairro, cidade, estado
                    FROM fornecedores
                    WHERE id = ?
                """, (fornecedor_id,))
                row = cursor.fetchone()
                conn.close()
                
                if row:
                    fornecedor_info['nome_fantasia'] = row[0] or ''
                    fornecedor_info['cnpj'] = format_cnpj(row[1]) if row[1] else ''
                    fornecedor_info['telefone'] = row[2] or ''
                    endereco_parts = []
                    if row[3]:
                        endereco_parts.append(row[3])
                    if row[4]:
                        endereco_parts.append(row[4])
                    if row[5] and row[6]:
                        endereco_parts.append(f"{row[5]}/{row[6]}")
                    fornecedor_info['endereco'] = ', '.join(endereco_parts) if endereco_parts else 'Não informado'
        except Exception as e:
            print(f"Erro ao buscar fornecedor {fornecedor_id}: {str(e)}")
        
        # EXTRAIR DESCONTO DA CONDIÇÃO DE PAGAMENTO
        def extrair_desconto_da_condicao(condicao_pagamento):
            if not condicao_pagamento:
                return 0.0
            import re
            match = re.search(r'DESCONTO:\s*R?\$?\s*(\d+(?:[.,]\d+)?)', condicao_pagamento, re.IGNORECASE)
            if match:
                return float(match.group(1).replace(',', '.'))
            return 0.0
        
        valor_desconto = extrair_desconto_da_condicao(cotacao_ref.condicao_pagamento or '')
        
        cotacao_estruturada = {
            'fornecedor_id': fornecedor_id,
            'fornecedor_nome_fantasia': fornecedor_info['nome_fantasia'],
            'fornecedor_cnpj': fornecedor_info['cnpj'],
            'fornecedor_telefone': fornecedor_info['telefone'],
            'fornecedor_endereco': fornecedor_info['endereco'],
            'valor_frete': cotacao_ref.valor_frete,
            'prazo_entrega': cotacao_ref.prazo_entrega or '',
            'condicao_pagamento': cotacao_ref.condicao_pagamento or '',
            'observacoes': cotacao_ref.observacoes or '',
            'pdf_path': cotacao_ref.pdf_path,
            'desconto': valor_desconto,
            'materiais': []
        }
            
        for cotacao in cotacoes_fornecedor:
            solicitacao_match = next((s for s in todas_solicitacoes if s.id == cotacao.solicitacao_id), None)
            
            if solicitacao_match:
                historico_descontos = []
                try:
                    historico_query = HistoricoDescontos.query.filter_by(
                        preenchimento_id=cotacao.id
                    ).order_by(HistoricoDescontos.data_alteracao.desc()).all()
                    
                    for historico in historico_query:
                        historico_descontos.append({
                            'valor_unitario_anterior': historico.valor_unitario_anterior,
                            'valor_unitario_novo': historico.valor_unitario_novo,
                            'valor_frete_anterior': historico.valor_frete_anterior,
                            'valor_frete_novo': historico.valor_frete_novo,
                            'data_alteracao': historico.data_alteracao.strftime('%d/%m/%Y %H:%M:%S'),
                            'usuario': historico.usuario
                        })
                except Exception as e:
                    print(f"DEBUG - Erro ao buscar histórico de descontos: {str(e)}")
                
                material_info = {
                    'preenchimento_id': cotacao.id,
                    'solicitacao_id': cotacao.solicitacao_id,
                    'valor_unitario': float(cotacao.valor_unitario) if cotacao.valor_unitario else 0.0,
                    'valor_total': float(cotacao.valor_total) if cotacao.valor_total else 0.0,
                    'descricao': solicitacao_match.material.DescricaoMaterial if solicitacao_match.material else '',
                    'quantidade': converter_quantidade_para_float(solicitacao_match.quantidade),
                    'unidade': solicitacao_match.unidade_medida or 'un',
                    'marca': cotacao.marca or '',
                    'historico_descontos': historico_descontos
                }
                cotacao_estruturada['materiais'].append(material_info)
        
        cotacoes_estruturadas.append(cotacao_estruturada)
    
    if not cotacoes_estruturadas:
        cotacoes_estruturadas = [{
            'fornecedor_id': None,
            'fornecedor_nome_fantasia': '',
            'fornecedor_cnpj': '',
            'fornecedor_telefone': '',
            'fornecedor_endereco': '',
            'valor_frete': None,
            'prazo_entrega': '',
            'condicao_pagamento': '',
            'observacoes': '',
            'pdf_path': None,
            'desconto': 0.0,
            'materiais': []
        }]
    
    return render_template(
        'preencher_solicitacao.html',
        solicitacao=solicitacao,
        solicitacoes_grupo=todas_solicitacoes,
        cotacoes_salvas=cotacoes_estruturadas,
        modo_grupo=modo_grupo,
        grupo_ids=grupo_ids_param
    )

@app.template_filter('regex_replace')
def regex_replace(s, pattern, replacement):
    if s is None:
        return ''
    import re
    try:
        return re.sub(pattern, replacement, str(s))
    except:
        return str(s)
    
# Imprimir -------------------------------------------------------------------------
@routes_bp.route('/cotacao_imprimir/<int:id>', methods=['GET'])
def cotacao_imprimir(id):
    if 'usuario' not in session:
        flash('Acesso não autorizado', 'error')
        return redirect(url_for('routes_bp.login'))

    # === VERIFICAR SE EXISTE GRUPO ===
    grupo_ids_param = request.args.get('grupo_ids', '')
    grupo_ids = []
    solicitacoes_grupo = []

    if grupo_ids_param:
        try:
            grupo_ids = [int(x) for x in grupo_ids_param.split(',') if x.strip().isdigit()]
            if grupo_ids:
                solicitacoes_grupo = SolicitacoesCompra.query.filter(
                    SolicitacoesCompra.id.in_(grupo_ids),
                    SolicitacoesCompra.status_aprovacao == 'Aprovado'
                ).all()
        except Exception as e:
            print(f"Erro ao processar grupo_ids: {e}")
    
    if not solicitacoes_grupo:
        solicitacao_principal = SolicitacoesCompra.query.get_or_404(id)
        solicitacoes_grupo = [solicitacao_principal]
        grupo_ids = [id]

    # ============================================
    # 🔴 BUSCAR APENAS PREENCHIMENTOS DESTE GRUPO
    # ============================================
    
    # Buscar preenchimentos APENAS para as solicitações deste grupo
    preenchimentos = SolicitacoesPreenchidas.query.filter(
        SolicitacoesPreenchidas.solicitacao_id.in_(grupo_ids),
        SolicitacoesPreenchidas.status.in_(['Rascunho', 'Aguardando Aprovacao'])
    ).all()

    print(f"🔍 Total de preenchimentos encontrados: {len(preenchimentos)}")
    
    # 🔴 ORGANIZAR POR FORNECEDOR - APENAS OS QUE TÊM PREENCHIMENTOS REAIS
    cotacoes_por_fornecedor = {}
    
    for p in preenchimentos:
        fornecedor_id = p.fornecedor_id
        
        # Verificar se o fornecedor existe no banco
        conn = get_db_connection(DB_PATH_FORNECEDORES)
        fornecedor_nome = None
        fornecedor_telefone = None
        fornecedor_cnpj = None
        
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT nome_fantasia, cnpj, telefone FROM fornecedores WHERE id = ?", (fornecedor_id,))
            row = cursor.fetchone()
            conn.close()
            if row:
                fornecedor_nome = row[0] or 'Fornecedor não encontrado'
                fornecedor_cnpj = row[1] or ''
                fornecedor_telefone = row[2] or ''
            else:
                print(f"⚠️ Fornecedor ID {fornecedor_id} não encontrado no banco - ignorando")
                continue  # Pular este preenchimento se fornecedor não existe
        
        if fornecedor_nome is None:
            continue
            
        if fornecedor_id not in cotacoes_por_fornecedor:
            cotacoes_por_fornecedor[fornecedor_id] = {
                'fornecedor_id': fornecedor_id,
                'fornecedor_nome_fantasia': fornecedor_nome,
                'fornecedor_cnpj': fornecedor_cnpj,
                'fornecedor_telefone': fornecedor_telefone,
                'valor_frete': p.valor_frete or 0,
                'prazo_entrega': p.prazo_entrega or '',
                'condicao_pagamento': p.condicao_pagamento or '',
                'observacoes': p.observacoes or '',
                'materiais': []
            }
        
        # Buscar a solicitação correspondente
        solicitacao = next((s for s in solicitacoes_grupo if s.id == p.solicitacao_id), None)
        if solicitacao:
            # Converter quantidade
            try:
                quantidade = float(solicitacao.quantidade) if solicitacao.quantidade else 0
            except (ValueError, TypeError):
                quantidade = 0
            
            material_info = {
                'solicitacao_id': p.solicitacao_id,
                'valor_unitario': float(p.valor_unitario) if p.valor_unitario else 0,
                'valor_total': float(p.valor_total) if p.valor_total else 0,
                'descricao': solicitacao.material.DescricaoMaterial if solicitacao.material else 'Material não encontrado',
                'quantidade': quantidade,
                'unidade': solicitacao.unidade_medida or 'un',
                'marca': p.marca or solicitacao.marca or '',
                'especificacao': solicitacao.especificacao or '',
                'preenchimento_id': p.id
            }
            cotacoes_por_fornecedor[fornecedor_id]['materiais'].append(material_info)
    
    # Converter para lista e ordenar por nome do fornecedor
    cotacoes_estruturadas = list(cotacoes_por_fornecedor.values())
    cotacoes_estruturadas.sort(key=lambda x: x['fornecedor_nome_fantasia'])
    
    # 🔴 LOG PARA DEBUG - ver quais fornecedores estão sendo exibidos
    print(f"🔍 FORNECEDORES QUE SERÃO EXIBIDOS NA IMPRESSÃO ({len(cotacoes_estruturadas)}):")
    for c in cotacoes_estruturadas:
        print(f"   - {c['fornecedor_nome_fantasia']} (ID: {c['fornecedor_id']}) - Materiais: {len(c['materiais'])}")
    
    # Se não houver cotações, criar estrutura vazia
    if not cotacoes_estruturadas:
        cotacoes_estruturadas = [{
            'fornecedor_id': None,
            'fornecedor_nome_fantasia': '',
            'fornecedor_cnpj': '',
            'fornecedor_telefone': '',
            'valor_frete': 0,
            'prazo_entrega': '',
            'condicao_pagamento': '',
            'observacoes': '',
            'materiais': []
        }]

    # Criar lista de materiais únicos para o template
    todos_materiais_unicos = []
    for sol in solicitacoes_grupo:
        try:
            quantidade = float(sol.quantidade) if sol.quantidade else 0
        except (ValueError, TypeError):
            quantidade = 0
        
        material_completo = {
            'id': sol.id,
            'descricao': sol.material.DescricaoMaterial if sol.material else '',
            'quantidade': quantidade,
            'unidade': sol.unidade_medida or 'un',
            'especificacao': sol.especificacao or ''
        }
        todos_materiais_unicos.append(material_completo)
    
    todos_materiais_unicos.sort(key=lambda x: x['id'])

    return render_template(
        'cotacao_imprimir.html',
        solicitacao=solicitacoes_grupo[0],
        solicitacoes_grupo=solicitacoes_grupo,
        todos_materiais=todos_materiais_unicos,
        cotacoes_salvas=cotacoes_estruturadas,
        modo_grupo=len(solicitacoes_grupo) > 1,
        grupo_ids=','.join(str(gid) for gid in grupo_ids)
    )
#------------------------------------------------------------------

@routes_bp.route('/api/registrar_envio_aprovacao', methods=['POST'])
def registrar_envio_aprovacao():
    """Registra que o comprador levou as solicitações para aprovação presencial."""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401

    try:
        data = request.get_json()
        ids_str = data.get('ids', '')  # ex: "12,13,14"
        if not ids_str:
            return jsonify({'success': False, 'message': 'Nenhum ID informado'}), 400

        ids = [int(i.strip()) for i in ids_str.split(',') if i.strip().isdigit()]
        if not ids:
            return jsonify({'success': False, 'message': 'IDs inválidos'}), 400

        agora = get_local_time()
        atualizados = 0
        for sol_id in ids:
            sol = db.session.get(SolicitacoesCompra, sol_id)
            if sol:
                # Só registra se ainda não foi enviado (evita sobrescrever)
                if not sol.data_envio_aprovacao:
                    sol.data_envio_aprovacao = agora
                atualizados += 1

        db.session.commit()
        app.logger.info(f"Envio para aprovação registrado: IDs={ids}, usuário={session.get('usuario')}")
        return jsonify({'success': True, 'atualizados': atualizados, 'data': agora.strftime('%d/%m/%Y %H:%M')})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Erro em registrar_envio_aprovacao: {str(e)}', exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


@routes_bp.route('/listar_solicitacoes_preenchidas', methods=['GET'])
def listar_solicitacoes_preenchidas():
    """Lista solicitações preenchidas com status Aguardando Aprovacao — paginado (20 por página)"""

    if 'usuario' not in session:
        flash('🔒 Acesso não autorizado.', 'warning')
        return redirect(url_for('routes_bp.login'))

    try:
        # Parâmetros de paginação e filtro
        page = request.args.get('page', 1, type=int)
        per_page = 20

        empresa = request.args.get('empresa', '').strip()
        usuario = request.args.get('usuario', '').strip()
        data_inicio_str = request.args.get('data_inicio', '').strip()
        data_fim_str = request.args.get('data_fim', '').strip()

        filtros = {
            'empresa': empresa,
            'usuario': usuario,
            'data_inicio': data_inicio_str,
            'data_fim': data_fim_str
        }

        # Converter datas
        data_inicio = None
        data_fim = None

        if data_inicio_str:
            try:
                data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d')
            except ValueError:
                flash('⚠️ Data inicial inválida.', 'warning')

        if data_fim_str:
            try:
                data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d') + timedelta(days=1)
            except ValueError:
                flash('⚠️ Data final inválida.', 'warning')

        # Query base — somente 'Aguardando Aprovacao', com join para filtros por empresa
        query = SolicitacoesPreenchidas.query.filter(
            SolicitacoesPreenchidas.status == 'Aguardando Aprovacao'
        ).join(SolicitacoesCompra)

        if empresa:
            query = query.filter(SolicitacoesCompra.empresa == empresa)
        if usuario:
            query = query.filter(SolicitacoesPreenchidas.usuario == usuario)
        if data_inicio:
            query = query.filter(SolicitacoesPreenchidas.data_preenchimento >= data_inicio)
        if data_fim:
            query = query.filter(SolicitacoesPreenchidas.data_preenchimento <= data_fim)

        # Total real sem carregar tudo na memória
        total_cotacoes = query.count()

        # Paginar: busca só os 20 da página atual
        paginacao = query.order_by(
            SolicitacoesPreenchidas.data_preenchimento.desc()
        ).paginate(page=page, per_page=per_page, error_out=False)

        preenchimentos = paginacao.items

        # Estruturar dados — bulk fetch de fornecedores para evitar N+1
        forn_ids_preench = {p.fornecedor_id for p in preenchimentos if p.fornecedor_id}
        fornecedores_preench = get_fornecedores_bulk(forn_ids_preench)

        preenchimentos_por_material = {}
        for p in preenchimentos:
            if not p.solicitacao:
                continue
            material_nome = (
                p.solicitacao.material.DescricaoMaterial
                if p.solicitacao.material else 'Material não encontrado'
            )
            fornecedor_nome = (
                fornecedores_preench.get(p.fornecedor_id, {}).get('nome_fantasia', 'N/A')
                if p.fornecedor_id else 'N/A'
            )
            preenchimento_info = {
                'id': p.id,
                'fornecedor_nome': fornecedor_nome,
                'solicitacao': p.solicitacao,
                'valor_unitario': p.valor_unitario or 0,
                'valor_frete': p.valor_frete or 0,
                'valor_total': p.valor_total or 0,
                'prazo_entrega': p.prazo_entrega or 'Não informado',
                'condicao_pagamento': p.condicao_pagamento or 'Não informada',
                'status': p.status,
                'usuario': p.usuario,
                'pdf_path': p.pdf_path,
                'observacoes': p.observacoes or ''
            }
            if material_nome not in preenchimentos_por_material:
                preenchimentos_por_material[material_nome] = []
            preenchimentos_por_material[material_nome].append(preenchimento_info)

        # Listas para filtros (queries leves — só valores distintos)
        empresas = []
        try:
            empresas = sorted([
                e[0] for e in db.session.query(SolicitacoesCompra.empresa).distinct().all() if e[0]
            ])
        except Exception:
            pass

        usuarios = []
        try:
            usuarios = sorted([
                u[0] for u in db.session.query(SolicitacoesPreenchidas.usuario).distinct().all() if u[0]
            ])
        except Exception:
            pass

        return render_template(
            'listar_solicitacoes_preenchidas.html',
            preenchimentos_por_material=preenchimentos_por_material,
            empresas=empresas,
            usuarios=usuarios,
            filtros=filtros,
            total_cotacoes=total_cotacoes,
            paginacao=paginacao,
            page=page,
        )

    except Exception as e:
        app.logger.error(f'❌ Erro em listar_solicitacoes_preenchidas: {str(e)}', exc_info=True)
        _flash_erro_seguro('Erro ao carregar solicitações.', e, app.logger)
        return render_template(
            'listar_solicitacoes_preenchidas.html',
            preenchimentos_por_material={},
            empresas=[],
            usuarios=[],
            filtros={},
            total_cotacoes=0,
            paginacao=None,
            page=1,
        )
# Rota de debug - adicione temporariamente no app.py
@routes_bp.route('/debug_preenchimentos')
def debug_preenchimentos():
    import os as _os
    if _os.getenv("FLASK_ENV", "production").lower() not in ("development", "dev"):
        abort(404)
    if 'usuario' not in session:
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        # Contar por status
        from sqlalchemy import func
        
        status_counts = db.session.query(
            SolicitacoesPreenchidas.status, 
            func.count(SolicitacoesPreenchidas.id)
        ).group_by(SolicitacoesPreenchidas.status).all()
        
        # Buscar alguns registros com status 'Aguardando Aprovacao'
        aguardando = SolicitacoesPreenchidas.query.filter_by(
            status='Aguardando Aprovacao'
        ).limit(5).all()
        
        return jsonify({
            'status_counts': [{'status': s[0], 'count': s[1]} for s in status_counts],
            'aguardando_aprovacao': [
                {
                    'id': p.id,
                    'solicitacao_id': p.solicitacao_id,
                    'fornecedor_id': p.fornecedor_id,
                    'status': p.status,
                    'usuario': p.usuario,
                    'data_preenchimento': str(p.data_preenchimento)
                } for p in aguardando
            ]
        })
    except Exception as e:
        app.logger.error(f'API error: {type(e).__name__}', exc_info=False)
        return jsonify({'error': 'Erro interno. Tente novamente.'}), 500
    
def get_fornecedor_nome(fornecedor_id):
    """Retorna o nome do fornecedor pelo ID (com fallback)"""
    if not fornecedor_id:
        return 'Fornecedor não informado'
    
    fornecedor = get_fornecedor_completo(fornecedor_id)
    if fornecedor and fornecedor['nome_fantasia']:
        return fornecedor['nome_fantasia']
    
    # Fallback: tentar buscar apenas o nome diretamente
    try:
        conn = sqlite3.connect(DB_PATH_FORNECEDORES)
        cursor = conn.cursor()
        cursor.execute('SELECT nome_fantasia FROM fornecedores WHERE id = ?', (fornecedor_id,))
        result = cursor.fetchone()
        conn.close()
        if result and result[0]:
            return result[0]
    except:
        pass
    
    return f'Fornecedor ID {fornecedor_id} (não encontrado)'

def get_fornecedor_completo(fornecedor_id):
    """
    Busca todos os dados do fornecedor pelo ID.
    Retorna um dicionário com todas as informações ou None se não encontrado.
    """
    if not fornecedor_id:
        return None
    
    try:
        conn = sqlite3.connect(DB_PATH_FORNECEDORES)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, nome_fantasia, cnpj, telefone, email, endereco, bairro, cidade, estado, contato, materiais
            FROM fornecedores 
            WHERE id = ?
        ''', (fornecedor_id,))
        result = cursor.fetchone()
        conn.close()
        
        if result:
            # Construir endereço completo
            endereco_parts = []
            if result['endereco']:
                endereco_parts.append(result['endereco'])
            if result['bairro']:
                endereco_parts.append(result['bairro'])
            if result['cidade'] and result['estado']:
                endereco_parts.append(f"{result['cidade']}/{result['estado']}")
            elif result['cidade']:
                endereco_parts.append(result['cidade'])
            elif result['estado']:
                endereco_parts.append(result['estado'])
            
            return {
                'id': result['id'],
                'nome_fantasia': result['nome_fantasia'] or '',
                'cnpj': result['cnpj'] or '',
                'telefone': result['telefone'] or '',
                'endereco': ', '.join(endereco_parts) if endereco_parts else 'Não informado',
                'email': result['email'] or '',
                'contato': result['contato'] or '',
                'materiais': result['materiais'] or ''
            }
        return None
    except sqlite3.Error as e:
        logging.error(f"Erro ao buscar fornecedor completo {fornecedor_id}: {str(e)}")
        return None
    except Exception as e:
        logging.error(f"Erro inesperado ao buscar fornecedor {fornecedor_id}: {str(e)}")
        return None


@routes_bp.route('/download_pdf/<int:preenchimento_id>', methods=['GET'])
def download_pdf(preenchimento_id):
    """Endpoint para download seguro de PDFs de cotações"""
    
    # Verificação de autenticação
    if 'usuario' not in session:
        flash('🔒 Acesso restrito. Por favor, faça login.', 'error')
        return redirect(url_for('routes_bp.login'))

    try:
        # Busca o preenchimento
        preenchimento = SolicitacoesPreenchidas.query.get_or_404(preenchimento_id)
        
        # Verifica se existe caminho do PDF
        if not preenchimento.pdf_path:
            flash('📄 Este orçamento não possui arquivo anexado.', 'info')
            return redirect(url_for('routes_bp.listar_pedidos_compra'))

        # Configurações de caminho
        uploads_dir = os.path.abspath(app.config['UPLOAD_FOLDER'])
        file_to_download = None

        # Tentativa 1: Se o caminho é absoluto
        if os.path.isabs(preenchimento.pdf_path):
            candidate_path = preenchimento.pdf_path
        else:
            # Tentativa 2: Caminho relativo à pasta Uploads
            candidate_path = os.path.join(uploads_dir, preenchimento.pdf_path)

        # Verifica se o arquivo existe
        if os.path.isfile(candidate_path):
            file_to_download = candidate_path
        else:
            # Tentativa 3: Busca por padrão (se o nome estiver incompleto)
            from glob import glob
            pattern = os.path.join(uploads_dir, f'*{preenchimento_id}*')
            matching_files = glob(pattern)
            
            if matching_files:
                file_to_download = matching_files[0]

        # Validação final
        if not file_to_download or not os.path.isfile(file_to_download):
            flash('⚠️ Arquivo PDF não encontrado no servidor.', 'warning')
            # Em vez de redirecionar para listar_solicitacoes_preenchidas, volta para a lista de pedidos
            return redirect(url_for('routes_bp.listar_pedidos_compra'))

        # Verificação de segurança
        if not os.path.abspath(file_to_download).startswith(uploads_dir):
            logging.warning(f'Tentativa de acesso a caminho não autorizado: {file_to_download}')
            flash('⛔ Caminho do arquivo inválido.', 'danger')
            return redirect(url_for('routes_bp.listar_pedidos_compra'))

        # Preparação do download
        fornecedor_nome = get_fornecedor_nome(preenchimento.fornecedor_id)
        filename = f"ORCAMENTO_{fornecedor_nome}_{preenchimento_id}_{os.path.basename(file_to_download)}"
        
        return send_file(
            file_to_download,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf',
            conditional=True
        )
        
    except Exception as e:
        logging.error(f'❌ Erro ao baixar PDF do preenchimento {preenchimento_id}: {str(e)}', exc_info=True)
        flash('⛔ Erro inesperado ao baixar orçamento.', 'danger')
        return redirect(url_for('routes_bp.listar_pedidos_compra'))

@routes_bp.route('/atualizar_status_preenchimento/<int:id>', methods=['POST'])
def atualizar_status_preenchimento(id):
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401

    try:
        preenchimento = SolicitacoesPreenchidas.query.get_or_404(id)
        novo_status = request.form.get('status')

        print(f"DEBUG >> Atualizando status para: {novo_status}")

        # Lógica para reverter status (status especial 'Reaberto')
        if novo_status == 'Reaberto':
            # Verificar se o status atual permite reverter
            if preenchimento.status in ['Aprovado', 'Reprovado']:
                preenchimento.status = 'Aguardando Aprovacao'  # Volta para o status inicial
                db.session.commit()
                print(f"DEBUG >> Status revertido para: Aguardando Aprovacao")
                
                return jsonify({ 
                    'success': True, 
                    'message': 'Status revertido com sucesso!'
                })
            else:
                return jsonify({
                    'success': False, 
                    'message': 'Não é possível reverter este status'
                }), 400

        # Lógica normal de atualização de status
        if novo_status:
            preenchimento.status = novo_status

        db.session.commit()
        
        # Log para debug
        print(f"DEBUG >> Status atualizado: {preenchimento.status}")
        
        return jsonify({
            'success': True, 
            'message': 'Status atualizado com sucesso!'
        })

    except Exception as e:
        db.session.rollback()
        print(f"ERRO >> {str(e)}")
        return jsonify({
            'success': False, 
            'message': 'Erro interno. Tente novamente.'
        }), 500
    
@routes_bp.route('/gerar_pedido_compra', methods=['GET', 'POST'])
def gerar_pedido_compra():
    if 'usuario' not in session:
        flash('Acesso não autorizado', 'error')
        return redirect(url_for('routes_bp.login'))

    try:
        # ==========================
        # POST
        # ==========================
        if request.method == 'POST':
            import time
            time.sleep(0.3)
            
            # 🔴 CORREÇÃO: Múltiplas estratégias para capturar os IDs
            preenchimento_ids = []
            
            # Estratégia 1: Tentar getlist com colchetes
            if 'preenchimento_ids[]' in request.form:
                preenchimento_ids = request.form.getlist('preenchimento_ids[]')
                print(f"Estratégia 1 - getlist com []: {preenchimento_ids}")
            
            # Estratégia 2: Tentar getlist sem colchetes
            if not preenchimento_ids and 'preenchimento_ids' in request.form:
                preenchimento_ids = request.form.getlist('preenchimento_ids')
                print(f"Estratégia 2 - getlist sem []: {preenchimento_ids}")
            
            # Estratégia 3: Tentar get simples (se veio como string única)
            if not preenchimento_ids:
                single_value = request.form.get('preenchimento_ids[]')
                if single_value:
                    preenchimento_ids = [single_value]
                    print(f"Estratégia 3 - get simples: {preenchimento_ids}")
            
            # Estratégia 4: Procurar em todas as chaves do formulário
            if not preenchimento_ids:
                for key, value in request.form.items():
                    if 'preenchimento' in key.lower() and value:
                        if isinstance(value, list):
                            preenchimento_ids.extend(value)
                        else:
                            preenchimento_ids.append(value)
                print(f"Estratégia 4 - varredura total: {preenchimento_ids}")
            
            # Filtrar valores vazios e converter para inteiros
            preenchimento_ids = [pid for pid in preenchimento_ids if pid and str(pid).strip()]
            
            print(f"🔴 IDs finais após processamento: {preenchimento_ids}")
            
            forma_pagamento = request.form.get('forma_pagamento', '').strip()
            condicao_pagamento = request.form.get('condicao_pagamento', '').strip()
            observacoes = request.form.get('observacoes', '').strip()

            if not preenchimento_ids:
                flash('Nenhum item selecionado. Marque os checkboxes dos itens desejados.', 'error')
                print("❌ ERRO: Nenhum preenchimento ID encontrado!")
                return redirect(url_for('routes_bp.gerar_pedido_compra'))

            # Converter para inteiros
            try:
                preenchimento_ids_int = [int(pid) for pid in preenchimento_ids]
            except ValueError as e:
                _flash_erro_seguro('Erro ao processar IDs selecionados.', e, app.logger)
                return redirect(url_for('routes_bp.gerar_pedido_compra'))
            
            # 🔴 VERIFICAR SE OS ITENS JÁ ESTÃO EM ALGUM PEDIDO
            from sqlalchemy import exists
            
            for pid in preenchimento_ids_int:
                existe = db.session.query(
                    exists().where(
                        pedido_preenchimento_associacao.c.preenchimento_id == pid
                    )
                ).scalar()
                
                if existe:
                    flash(f'Item ID {pid} já foi adicionado a outro pedido. Recarregue a página.', 'error')
                    return redirect(url_for('routes_bp.gerar_pedido_compra'))

            # Buscar TODOS os preenchimentos selecionados
            preenchimentos = SolicitacoesPreenchidas.query.filter(
                SolicitacoesPreenchidas.id.in_(preenchimento_ids_int)
            ).all()
            
            print(f"🔴 {len(preenchimentos)} preenchimentos encontrados no banco")

            if not preenchimentos:
                flash('Nenhum preenchimento válido encontrado.', 'error')
                return redirect(url_for('routes_bp.gerar_pedido_compra'))

            # Validar status
            for preenchimento in preenchimentos:
                if preenchimento.status != 'Aprovado':
                    flash(f'Item do fornecedor {get_fornecedor_nome(preenchimento.fornecedor_id)} não está aprovado.', 'error')
                    return redirect(url_for('routes_bp.gerar_pedido_compra'))

            # Gerar número sequencial por ANO
            ano_atual = datetime.now().year
            
            # Buscar o maior número sequencial do ano atual
            from sqlalchemy import func
            ultimo_numero_registro = db.session.query(
                func.max(PedidosCompra.id)
            ).filter(
                PedidosCompra.numero_pedido.like(f'PC{ano_atual}%')
            ).scalar()
            
            if ultimo_numero_registro:
                ultimo_pedido = PedidosCompra.query.get(ultimo_numero_registro)
                import re
                match = re.search(rf'PC{ano_atual}(\d+)', ultimo_pedido.numero_pedido)
                if match:
                    proximo_numero = int(match.group(1)) + 1
                else:
                    proximo_numero = 1
            else:
                proximo_numero = 1
            
            numero_pedido = f"PC{ano_atual}{proximo_numero:04d}"
            
            print(f"🔴 Número do pedido: {numero_pedido}")

            # Totais
            valor_total = sum(p.valor_total or 0 for p in preenchimentos)
            valor_frete_total = sum(p.valor_frete or 0 for p in preenchimentos)
            valor_liquido = valor_total + valor_frete_total

            # Forma de pagamento
            forma_condicao_final = None
            if forma_pagamento and condicao_pagamento:
                forma_condicao_final = f"{forma_pagamento} - {condicao_pagamento}"
            elif forma_pagamento:
                forma_condicao_final = forma_pagamento
            elif condicao_pagamento:
                forma_condicao_final = condicao_pagamento

            # CRIAR UM ÚNICO PEDIDO
            pedido = PedidosCompra(
                numero_pedido=numero_pedido,
                usuario=session['usuario'],
                status='Gerado',
                valor_total=valor_total,
                valor_frete=valor_frete_total if valor_frete_total != 0 else None,
                valor_liquido=valor_liquido,
                forma_pagamento=forma_condicao_final,
                observacoes=observacoes if observacoes else None,
                data_criacao=datetime.now()
            )

            db.session.add(pedido)
            db.session.flush()

            # ASSOCIAR TODOS OS PREENCHIMENTOS
            for preenchimento in preenchimentos:
                existe_associacao = db.session.query(
                    exists().where(
                        pedido_preenchimento_associacao.c.preenchimento_id == preenchimento.id
                    )
                ).scalar()
                
                if not existe_associacao:
                    db.session.execute(
                        pedido_preenchimento_associacao.insert().values(
                            pedido_id=pedido.id,
                            preenchimento_id=preenchimento.id
                        )
                    )
                    preenchimento.status = 'Em Processamento'
                    print(f"🔴 Associado preenchimento {preenchimento.id}")

            db.session.commit()
            
            print(f"✅ Pedido {numero_pedido} criado com {len(preenchimentos)} itens")
            
            flash(f'✅ Pedido {numero_pedido} gerado com sucesso! {len(preenchimentos)} item(ns) incluído(s).', 'success')
            return redirect(url_for('routes_bp.auditoria_solicitacoes'))

        # ==========================
        # GET - Carregar página
        # ==========================
        preenchimentos = SolicitacoesPreenchidas.query.filter_by(status='Aprovado')\
            .join(SolicitacoesCompra)\
            .join(Materiais)\
            .order_by(Materiais.DescricaoMaterial)\
            .all()

        # Agrupar por material
        preenchimentos_por_material = {}
        for preenchimento in preenchimentos:
            if not preenchimento.solicitacao or not preenchimento.solicitacao.material:
                continue
            material_nome = preenchimento.solicitacao.material.DescricaoMaterial
            if material_nome not in preenchimentos_por_material:
                preenchimentos_por_material[material_nome] = []
            preenchimentos_por_material[material_nome].append(preenchimento)

        formas_pagamento = ['À Vista', 'A Prazo', 'Boleto', 'Cartão de Crédito', 'Transferência Bancária']
        
        condicao_pagamento_padrao = ''
        if preenchimentos:
            condicao_pagamento_padrao = preenchimentos[0].condicao_pagamento or ''

        return render_template(
            'gerar_pedido_compra.html',
            preenchimentos_por_material=preenchimentos_por_material,
            formas_pagamento=formas_pagamento,
            condicao_pagamento_padrao=condicao_pagamento_padrao
        )

    except Exception as e:
        db.session.rollback()
        logging.error(f"Erro em gerar_pedido_compra: {str(e)}", exc_info=True)
        _flash_erro_seguro('Erro ao gerar pedido.', e, app.logger)
        return redirect(url_for('routes_bp.listar_solicitacoes_preenchidas'))
    
#testeaqui
@routes_bp.route('/pedido/<int:pedido_id>/view')
def view_pedido_pdf(pedido_id):
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        # Buscar o pedido no banco de dados
        pedido = PedidosCompra.query.get_or_404(pedido_id)
        
        # Obter informações dos fornecedores e materiais (mesma lógica usada na geração do PDF)
        fornecedor_ids = {p.fornecedor_id for p in pedido.preenchimentos}
        fornecedores_info = {}
        
        conn = get_db_connection(DB_PATH_FORNECEDORES)
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute(f'''
                    SELECT id, nome_fantasia, cnpj, telefone, email, endereco, cidade, estado 
                    FROM fornecedores 
                    WHERE id IN ({",".join("?"*len(fornecedor_ids))})
                ''', list(fornecedor_ids))
                
                for row in cursor.fetchall():
                    fornecedores_info[row[0]] = {
                        'nome': row[1],
                        'cnpj': format_cnpj(row[2]) if row[2] else 'N/A',
                        'telefone': row[3],
                        'email': row[4],
                        'endereco': f"{row[5]}, {row[6]}/{row[7]}"
                    }
            finally:
                conn.close()

        # Organizar materiais por fornecedor
        materiais_por_fornecedor = {}
        for preenchimento in pedido.preenchimentos:
            fornecedor_id = preenchimento.fornecedor_id
            if fornecedor_id not in materiais_por_fornecedor:
                materiais_por_fornecedor[fornecedor_id] = {
                    'info': fornecedores_info.get(fornecedor_id, {
                        'nome': 'Fornecedor não encontrado',
                        'cnpj': 'N/A',
                        'telefone': 'N/A',
                        'email': 'N/A',
                        'endereco': 'N/A'
                    }),
                    'itens': []
                }
            
            materiais_por_fornecedor[fornecedor_id]['itens'].append({
                'material': preenchimento.solicitacao.material.DescricaoMaterial,
                'marca': preenchimento.solicitacao.marca or 'Não especificado',
                'especificacao': preenchimento.solicitacao.especificacao,
                'quantidade': preenchimento.solicitacao.quantidade,
                'unidade': preenchimento.solicitacao.unidade_medida,
                'valor_unitario': preenchimento.valor_unitario,
                'valor_total': preenchimento.valor_total,
                'prazo_entrega': preenchimento.prazo_entrega,
                'prioridade': preenchimento.solicitacao.prioridade
            })

        # Renderizar o template com os dados
        return render_template(
            'pedido_compra_pdf.html',
            pedido=pedido,
            materiais_por_fornecedor=materiais_por_fornecedor,
            data_criacao=pedido.data_criacao.strftime('%d/%m/%Y %H:%M:%S'),
            usuario=session['usuario'],
            total_itens=len(pedido.preenchimentos),
            fornecedores_count=len(materiais_por_fornecedor),
            observacoes=pedido.observacoes
        )
        
    except Exception as e:
        _flash_erro_seguro('Erro ao carregar pedido.', e, app.logger)
        logging.error(f"Erro em view_pedido_pdf: {str(e)}")
        return redirect(url_for('routes_bp.listar_pedidos_compra'))
def get_fornecedor_details_full(fornecedor_id):
    """Busca todos os detalhes do fornecedor pelo ID no banco fornecedores.db"""
    try:
        conn = sqlite3.connect(DB_PATH_FORNECEDORES)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, nome_fantasia, cnpj, telefone, email, endereco, bairro, cidade, estado, contato, materiais
            FROM fornecedores 
            WHERE id = ?
        ''', (fornecedor_id,))
        result = cursor.fetchone()
        conn.close()
        if result:
            return {
                'id': result[0],
                'nome_fantasia': result[1],
                'cnpj': result[2],
                'telefone': result[3],
                'email': result[4],
                'endereco': result[5],
                'bairro': result[6],
                'cidade': result[7],
                'estado': result[8],
                'contato': result[9],
                'materiais': result[10]
            }
        return None
    except sqlite3.Error as e:
        logging.error(f"Erro ao buscar fornecedor {fornecedor_id}: {str(e)}")
        return None

# Adicione esta nova rota no app.py (próximo às outras rotas do blueprint routes_bp)
@routes_bp.route('/visualizar_pedido/<int:pedido_id>', methods=['GET'])
def visualizar_pedido(pedido_id):
    if 'usuario' not in session:
        flash('Acesso não autorizado', 'error')
        return redirect(url_for('routes_bp.login'))
    
    try:
        pedido = PedidosCompra.query.get_or_404(pedido_id)
        
        # Agrupar materiais por fornecedor (lógica similar à geração de PDF, adaptada)
        materiais_por_fornecedor = {}
        for preench in pedido.preenchimentos:
            forn_id = str(preench.fornecedor_id)  # Usar string como chave para evitar issues
            if forn_id not in materiais_por_fornecedor:
                info = get_fornecedor_details_full(preench.fornecedor_id)
                materiais_por_fornecedor[forn_id] = {'info': info, 'itens': []}
            
            # Detalhes do item
            item = {
                'material': preench.solicitacao.material.DescricaoMaterial if preench.solicitacao.material else 'Não informado',
                'marca': preench.solicitacao.marca or 'Sem marca',
                'especificacao': preench.solicitacao.especificacao or 'N/A',
                'quantidade': preench.solicitacao.quantidade,
                'unidade': preench.solicitacao.unidade_medida or 'Un',
                'valor_unitario': preench.valor_unitario,
                'valor_total': preench.valor_total,
                'prazo_entrega': preench.prazo_entrega or 'N/A'
            }
            materiais_por_fornecedor[forn_id]['itens'].append(item)
        
        # Calcular total de itens
        total_itens = sum(len(dados['itens']) for dados in materiais_por_fornecedor.values())
        
        # Renderizar o template como HTML (sem conversão para PDF)
        return render_template(
            'pedido_compra_pdf.html',
            pedido=pedido,
            usuario=pedido.usuario or session['usuario'],  # Ajuste conforme necessário (pode ser pedido.usuario)
            total_itens=total_itens,
            materiais_por_fornecedor=materiais_por_fornecedor
        )
    
    except Exception as e:
        _flash_erro_seguro('Erro ao visualizar pedido.', e, app.logger)
        logging.error(f"Erro em visualizar_pedido para ID {pedido_id}: {str(e)}", exc_info=True)
        return redirect(url_for('routes_bp.listar_pedidos_compras'))  # Ajuste para a rota de listagem correta
    
@routes_bp.route('/download_comprovante/<int:pedido_id>', methods=['GET'])
def download_comprovante(pedido_id):
    """Endpoint para download seguro de comprovantes de pedidos"""
    
    # Verificação de autenticação
    if 'usuario' not in session:
        flash('🔒 Acesso restrito. Por favor, faça login.', 'error')
        return redirect(url_for('routes_bp.login'))

    try:
        # Busca o pedido com tratamento de 404
        pedido = PedidosCompra.query.get_or_404(pedido_id)
        
        # Verifica existência do comprovante
        if not pedido.comprovante_pagamento:
            flash('📄 Este pedido não possui comprovante cadastrado.', 'info')
            return redirect(url_for('routes_bp.listar_pedidos_compra'))

        # Configurações de caminho
        uploads_dir = os.path.abspath(app.config['UPLOAD_FOLDER'])
        file_to_download = None

        # Tentativa 1: Caminho absoluto/relativo do banco
        stored_path = os.path.normpath(pedido.comprovante_pagamento)
        if os.path.isabs(stored_path):
            candidate_path = stored_path
        else:
            candidate_path = os.path.join(uploads_dir, stored_path)

        if os.path.isfile(candidate_path):
            file_to_download = candidate_path

        # Tentativa 2: Busca por padrão conhecido (se necessário)
        if not file_to_download:
            from glob import glob
            pattern = os.path.join(uploads_dir, f'pedido_{pedido_id}_*')
            matching_files = glob(pattern)
            
            if matching_files:
                file_to_download = matching_files[0]

        # Validação final do arquivo
        if not file_to_download or not os.path.isfile(file_to_download):
            flash('⚠️ Comprovante não encontrado no servidor.', 'warning')
            return redirect(url_for('routes_bp.listar_pedidos_compra'))

        # Verificação de segurança
        if not os.path.abspath(file_to_download).startswith(uploads_dir):
            logging.warning(f'Tentativa de acesso a caminho não autorizado: {file_to_download}')
            flash('⛔ Caminho do arquivo inválido.', 'danger')
            return redirect(url_for('routes_bp.listar_pedidos_compra'))

        # Preparação do download
        filename = f"COMPROVANTE_{pedido.numero_pedido}_{os.path.basename(file_to_download)}"
        
        return send_file(
            file_to_download,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf',
            conditional=True
        )
        
    except Exception as e:
        logging.error(f'❌ Erro ao baixar comprovante {pedido_id}: {str(e)}', exc_info=True)
        flash('⛔ Erro inesperado ao baixar comprovante.', 'danger')
        return redirect(url_for('routes_bp.listar_pedidos_compra'))
    
@routes_bp.route('/listar_pedidos_compra', methods=['GET'])
def listar_pedidos_compra():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))

    try:
        # Parâmetros
        pagina       = request.args.get('pagina', 1, type=int)
        por_pagina   = request.args.get('por_pagina', 50, type=int)
        status       = request.args.get('status')
        empresa      = request.args.get('empresa')
        comprador_atribuido = request.args.get('comprador_atribuido')
        data_inicio  = request.args.get('data_inicio')
        data_fim     = request.args.get('data_fim')
        numero_pedido = request.args.get('numero_pedido')

        if not isinstance(por_pagina, int) or por_pagina not in [50, 100]:
            por_pagina = 50
        if not isinstance(pagina, int) or pagina < 1:
            pagina = 1

        # Query base com JOINs para evitar filtragem em memória
        query = db.session.query(PedidosCompra)

        if status:
            query = query.filter(PedidosCompra.status == status)

        if numero_pedido:
            query = query.filter(PedidosCompra.numero_pedido.ilike(f'%{numero_pedido}%'))

        if data_inicio:
            query = query.filter(PedidosCompra.data_criacao >= data_inicio)

        if data_fim:
            try:
                dt_fim = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(PedidosCompra.data_criacao < dt_fim)
            except (ValueError, TypeError):
                pass

        # Filtros empresa/comprador movidos para SQL via subquery
        if empresa or comprador_atribuido:
            subq = (
                db.session.query(pedido_preenchimento_associacao.c.pedido_id)
                .join(SolicitacoesPreenchidas,
                      SolicitacoesPreenchidas.id == pedido_preenchimento_associacao.c.preenchimento_id)
                .join(SolicitacoesCompra,
                      SolicitacoesCompra.id == SolicitacoesPreenchidas.solicitacao_id)
            )
            if empresa:
                subq = subq.filter(SolicitacoesCompra.empresa == empresa)
            if comprador_atribuido:
                subq = subq.filter(SolicitacoesCompra.comprador_atribuido == comprador_atribuido)
            query = query.filter(PedidosCompra.id.in_(subq.subquery()))

        # Contagem e paginação no banco (não em memória)
        total_itens  = query.count()
        total_paginas = max(1, (total_itens + por_pagina - 1) // por_pagina)
        if pagina > total_paginas:
            pagina = total_paginas
        pedidos_paginados = (
            query.order_by(PedidosCompra.data_criacao.desc())
            .offset((pagina - 1) * por_pagina)
            .limit(por_pagina)
            .all()
        )

        # Empresas e compradores via ler_senhas() cacheado — sem abrir senhas.txt diretamente
        _senhas = ler_senhas()
        empresas_unicas   = sorted({d['empresa'] for d in _senhas.values() if d.get('empresa')})
        compradores_unicos = sorted({
            u for u, d in _senhas.items() if 'comprador' in d.get('pagina', '').lower()
        })
        compradores_empresas = {
            u: d['empresa'] for u, d in _senhas.items()
            if 'comprador' in d.get('pagina', '').lower()
        }

        # Fornecedores em lote — UMA única conexão para toda a página
        fornecedor_ids = {
            p.fornecedor_id
            for pedido in pedidos_paginados
            for p in pedido.preenchimentos
            if p.fornecedor_id
        }
        fornecedores = get_fornecedores_bulk(fornecedor_ids)

        # Estruturar dados para o template
        pedidos_completos = []
        for pedido in pedidos_paginados:
            preenchimentos_info = []
            for preenchimento in pedido.preenchimentos:
                fornecedor_info = fornecedores.get(preenchimento.fornecedor_id, {
                    'nome_fantasia': 'Fornecedor não encontrado',
                    'cnpj': 'N/A'
                })
                
                solicitacao = preenchimento.solicitacao
                empresa_usuario = compradores_empresas.get(pedido.usuario, '')
                if not empresa_usuario and solicitacao:
                    empresa_usuario = solicitacao.empresa
                
                marca = solicitacao.marca if solicitacao and solicitacao.marca else 'Não informado'
                
                preenchimentos_info.append({
                    'id': preenchimento.id,
                    'marca': marca,
                    'fornecedor_nome': fornecedor_info['nome_fantasia'],
                    'fornecedor_cnpj': fornecedor_info.get('cnpj', 'N/A'),
                    'fornecedor_id': preenchimento.fornecedor_id,
                    'material': solicitacao.material.DescricaoMaterial if solicitacao and solicitacao.material else 'N/A',
                    'empresa': empresa_usuario,
                    'comprador_atribuido': solicitacao.comprador_atribuido if solicitacao else pedido.usuario,
                    'pdf_path': preenchimento.pdf_path
                })
            
            pedidos_completos.append({
                'pedido': pedido,
                'preenchimentos': preenchimentos_info,
                'observacoes': pedido.observacoes
            })

        return render_template(
            'listar_pedidos_compra.html', 
            pedidos_completos=pedidos_completos,
            empresas=sorted(empresas_unicas),
            compradores=sorted(compradores_unicos),
            filtros={
                'empresa': empresa,
                'comprador_atribuido': comprador_atribuido,
                'data_inicio': data_inicio,
                'data_fim': data_fim,
                'status': status,
                'numero_pedido': numero_pedido
            },
            pagina_atual=pagina,
            total_paginas=total_paginas,
            total_itens=total_itens,
            por_pagina=por_pagina,
            request=request
        )
        
    except Exception as e:
        logging.error(f"Erro ao listar pedidos de compra: {str(e)}", exc_info=True)
        _flash_erro_seguro('Erro ao listar pedidos de compra.', e, app.logger)
        return render_template(
            'listar_pedidos_compra.html', 
            pedidos_completos=[],
            empresas=[],
            compradores=[],
            filtros={},
            pagina_atual=1,
            total_paginas=1,
            total_itens=0,
            por_pagina=50,
            request=request
        )
    
#Nova Pagina pagamento
@routes_bp.route('/listar_pedidos_compras_pg', methods=['GET'])
def listar_pedidos_compra_pg():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        # Obter parâmetros de filtro (exceto status, pois queremos apenas AgPagamento)
        empresa_filtro = request.args.get('empresa')
        usuario_filtro = request.args.get('usuario')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')

        # Ler usuários e empresas do arquivo senhas.txt
        usuarios_empresas = {}
        empresas_unicas = set()
        usuarios_unicos = set()
        
        try:
            with open('senhas.txt', 'r', encoding='utf-8') as f:
                for line in f:
                    partes = line.strip().split('%')
                    if len(partes) >= 4:
                        usuario = partes[0]
                        empresa = partes[3]
                        usuarios_empresas[usuario] = empresa
                        empresas_unicas.add(empresa)
                        usuarios_unicos.add(usuario)
        except Exception as e:
            logging.error(f"Erro ao ler senhas.txt: {str(e)}")
            usuarios_empresas = {}

        # Consulta base - FILTRANDO APENAS POR AgPagamento
        query = db.session.query(PedidosCompra).filter(
            PedidosCompra.status == 'AgPagamento'  # Filtro fixo para AgPagamento
        ).join(
            pedido_preenchimento_associacao,
            PedidosCompra.id == pedido_preenchimento_associacao.c.pedido_id
        ).join(
            SolicitacoesPreenchidas,
            pedido_preenchimento_associacao.c.preenchimento_id == SolicitacoesPreenchidas.id
        ).join(
            SolicitacoesCompra,
            SolicitacoesPreenchidas.solicitacao_id == SolicitacoesCompra.id
        )

        # Aplicar outros filtros (exceto status)
        if empresa_filtro:
            query = query.filter(
                (SolicitacoesCompra.empresa == empresa_filtro) |
                (PedidosCompra.usuario.in_(
                    [usuario for usuario, empresa in usuarios_empresas.items() if empresa == empresa_filtro]
                ))
            )
        
        if usuario_filtro:
            query = query.filter(PedidosCompra.usuario == usuario_filtro)
        
        if data_inicio:
            query = query.filter(PedidosCompra.data_criacao >= data_inicio)
        
        if data_fim:
            data_fim_ajustada = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(PedidosCompra.data_criacao <= data_fim_ajustada)

        # Ordenar e executar a consulta
        pedidos = query.order_by(
            PedidosCompra.data_criacao.desc()
        ).distinct().all()

        # Restante do código permanece igual...
        fornecedor_ids = set()
        for pedido in pedidos:
            for preenchimento in pedido.preenchimentos:
                fornecedor_ids.add(preenchimento.fornecedor_id)
        
        fornecedores = {}
        conn = get_db_connection(DB_PATH_FORNECEDORES)
        if conn:
            try:
                cursor = conn.cursor()
                cursor.execute(f'SELECT id, nome_fantasia, cnpj FROM fornecedores WHERE id IN ({",".join("?"*len(fornecedor_ids))})', 
                             list(fornecedor_ids))
                for row in cursor.fetchall():
                    fornecedores[row[0]] = {
                        'nome_fantasia': row[1],
                        'cnpj': format_cnpj(row[2]) if row[2] else 'N/A'
                    }
            finally:
                conn.close()

        pedidos_completos = []
        for pedido in pedidos:
            preenchimentos_info = []
            for preenchimento in pedido.preenchimentos:
                fornecedor_info = fornecedores.get(preenchimento.fornecedor_id, {
                    'nome_fantasia': 'Fornecedor não encontrado',
                    'cnpj': 'N/A'
                })
                empresa_usuario = usuarios_empresas.get(pedido.usuario, preenchimento.solicitacao.empresa)
                
                # CORREÇÃO: A marca está na solicitação, não no preenchimento
                marca = preenchimento.solicitacao.marca if preenchimento.solicitacao.marca else 'Não informado'
                
                preenchimentos_info.append({
                    'id': preenchimento.id,
                    'marca': marca,  # Usando a marca da solicitação
                    'fornecedor_nome': fornecedor_info['nome_fantasia'],
                    'fornecedor_cnpj': fornecedor_info.get('cnpj', 'N/A'),
                    'material': preenchimento.solicitacao.material.DescricaoMaterial if preenchimento.solicitacao.material else 'N/A',
                    'empresa': empresa_usuario
                })
            pedidos_completos.append({
                'pedido': pedido,
                'preenchimentos': preenchimentos_info
            })

        return render_template(
            'listar_pedidos_compras_pg.html', 
            pedidos_completos=pedidos_completos,
            empresas=sorted(empresas_unicas),
            usuarios=sorted(usuarios_unicos),
            filtros={
                'empresa': empresa_filtro,
                'usuario': usuario_filtro,
                'data_inicio': data_inicio,
                'data_fim': data_fim,
                'status': 'AgPagamento'  # Definindo o status fixo para o template
            }
        )
    except Exception as e:
        _flash_erro_seguro('Erro ao listar pedidos de compra.', e, app.logger)
        return render_template(
            'listar_pedidos_compras_pg.html', 
            pedidos_completos=[],
            empresas=[],
            usuarios=[],
            filtros={}
        )
    
@routes_bp.route('/download_pedido_pdf/<int:pedido_id>', methods=['GET'])
def download_pedido_pdf(pedido_id):
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        pedido = app.jinja_env.globals['PedidosCompra'].query.get_or_404(pedido_id)
        if not pedido.pdf_path or not os.path.exists(pedido.pdf_path):
            flash('Arquivo PDF não encontrado.', 'error')
            return redirect(url_for('routes_bp.listar_pedidos_compra'))
        
        return send_file(pedido.pdf_path, as_attachment=True)
    except Exception as e:
        _flash_erro_seguro('Erro ao baixar PDF.', e, app.logger)
        return redirect(url_for('routes_bp.listar_pedidos_compra'))

@routes_bp.route('/atualizar_status_pedido/<int:id>', methods=['POST'])
def atualizar_status_pedido(id):
    if 'usuario' not in session:
        flash('Usuário não autenticado.', 'error')
        return jsonify({'flash': [['Usuário não autenticado.', 'error']]}), 401

    try:
        pedido = app.jinja_env.globals['PedidosCompra'].query.get_or_404(id)
        status = request.form.get('status', '').strip()
        fornecedor_id = request.form.get('fornecedor_id', '').strip()
        numero_nf = request.form.get('numero_nf', '').strip()
        comprovante_pagamento = request.form.get('comprovante', '').strip()
        pdf_file = request.files.get('pdf_file')

        valid_statuses = ['Gerado', 'Aprovado', 'AgPagamento', 'AgEntrega', 'Entregue', 'Reprovado']
        if not status or status not in valid_statuses:
            flash(f'Status inválido: {status}', 'error')
            return jsonify({'flash': [['Status inválido.', 'error']]}), 400

        # Lógica para pedidos à vista
        if pedido.forma_pagamento and 'À Vista' in pedido.forma_pagamento:
            if status == 'Aprovado':
                # Para pedidos à vista, aprovar muda automaticamente para AgPagamento
                pedido.status = 'AgPagamento'
                flash('Pedido aprovado e movido para Aguardando Pagamento.', 'success')
            elif status == 'AgPagamento':
                # Validar comprovante e PDF
                if not comprovante_pagamento:
                    flash('Número do comprovante é obrigatório para Aguardando Pagamento.', 'error')
                    return jsonify({'flash': [['Número do comprovante é obrigatório.', 'error']]}), 400
                if not pdf_file or not pdf_file.filename.endswith('.pdf'):
                    flash('Anexe um arquivo PDF válido para Aguardando Pagamento.', 'error')
                    return jsonify({'flash': [['Anexe um arquivo PDF válido.', 'error']]}), 400
                if pdf_file.content_length > 5 * 1024 * 1024:  # Limite de 5MB
                    flash('O arquivo PDF deve ter no máximo 5MB.', 'error')
                    return jsonify({'flash': [['O arquivo PDF deve ter no máximo 5MB.', 'error']]}), 400

                # Salvar PDF
                filename = f"pedido_{id}_{uuid.uuid4().hex[:8]}_{secure_filename(pdf_file.filename)}"
                pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                try:
                    pdf_file.save(pdf_path)
                except Exception as e:
                    _flash_erro_seguro('Erro ao salvar o PDF.', e, app.logger)
                    return jsonify({'flash': [['Erro ao salvar o PDF.', 'error']]}), 500

                pedido.pdf_path = pdf_path
                pedido.comprovante_pagamento = comprovante_pagamento
                pedido.status = 'AgEntrega'
                flash('Comprovante salvo. Pedido movido para Aguardando Entrega.', 'success')

        # Lógica para pedidos faturados a prazo
        else:
            if status == 'Aprovado':
                # Para pedidos a prazo, aprovar muda diretamente para AgEntrega
                pedido.status = 'AgEntrega'
                flash('Pedido aprovado e movido para Aguardando Entrega.', 'success')

        # Lógica para status Entregue (comum a ambos os tipos de pagamento)
        if status == 'Entregue':
            if not fornecedor_id or not numero_nf:
                flash('Fornecedor e número da NF são obrigatórios para o status Entregue.', 'error')
                return jsonify({'flash': [['Fornecedor e número da NF são obrigatórios.', 'error']]}), 400

            # Validar fornecedor
            conn = get_db_connection(DB_PATH_FORNECEDORES)
            if conn:
                try:
                    cursor = conn.cursor()
                    cursor.execute('SELECT id FROM fornecedores WHERE id = ?', (fornecedor_id,))
                    if not cursor.fetchone():
                        flash('Fornecedor inválido.', 'error')
                        return jsonify({'flash': [['Fornecedor inválido.', 'error']]}), 400
                finally:
                    conn.close()

            # Atualizar preenchimentos e estoque
            for preenchimento in pedido.preenchimentos:
                preenchimento.status = 'Entregue'
                preenchimento.fornecedor_id = fornecedor_id
                preenchimento.numero_nf = numero_nf

                material = Materiais.query.filter_by(
                    CodMaterial=preenchimento.solicitacao.cod_material
                ).first()

                if material:
                    material.QuantidadeEstoque += preenchimento.solicitacao.quantidade
                    material.Fornecedor = fornecedor_id
                    material.NumeroNF = numero_nf

                    estoque_existente = app.jinja_env.globals['Estoque'].query.filter_by(
                        preenchimento_id=preenchimento.id
                    ).first()

                    if not estoque_existente:
                        estoque = app.jinja_env.globals['Estoque'](
                            preenchimento_id=preenchimento.id,
                            cod_material=preenchimento.solicitacao.cod_material,
                            quantidade=preenchimento.solicitacao.quantidade,
                            fornecedor=fornecedor_id,
                            numero_nf=numero_nf,
                            usuario=session.get('usuario')
                        )
                        db.session.add(estoque)
                    else:
                        estoque_existente.fornecedor = fornecedor_id
                        estoque_existente.numero_nf = numero_nf
                        estoque_existente.quantidade = preenchimento.solicitacao.quantidade

            pedido.status = status
            flash('Pedido marcado como Entregue. Estoque atualizado.', 'success')

        # Atualizar status para outros casos
        if status not in ['Aprovado', 'AgPagamento', 'Entregue']:
            pedido.status = status
            flash('Status atualizado com sucesso.', 'success')

        db.session.commit()

        # Retornar mensagens flash como JSON
        messages = []
        for message, category in session.get('_flashes', []):
            messages.append([message, category])
        session['_flashes'] = []  # Limpar flashes após envio
        return jsonify({'flash': messages}), 204

    except Exception as e:
        db.session.rollback()
        _flash_erro_seguro('Erro interno ao atualizar status.', e, app.logger)
        return jsonify({'flash': [['Erro interno ao atualizar status.', 'error']]}), 500
    
@routes_bp.route('/adicionar_estoque/<int:preenchimento_id>', methods=['GET', 'POST'])
def adicionar_estoque(preenchimento_id):
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        preenchimento = app.jinja_env.globals['SolicitacoesPreenchidas'].query.get_or_404(preenchimento_id)
        if preenchimento.status != 'Entregue':
            flash('O status deve ser "Entregue" para adicionar ao estoque.', 'error')
            return redirect(url_for('routes_bp.listar_solicitacoes_preenchidas'))
        
        if request.method == 'POST':
            fornecedor = request.form.get('fornecedor', '').strip()
            numero_nf = request.form.get('numero_nf', '').strip()

            if not fornecedor or not numero_nf:
                flash('Fornecedor e número da NF são obrigatórios.', 'error')
                return render_template('adicionar_estoque.html', preenchimento=preenchimento)
            
            material = Materiais.query.filter_by(CodMaterial=preenchimento.solicitacao.cod_material).first()
            if material:
                material.QuantidadeEstoque += preenchimento.solicitacao.quantidade
                material.Fornecedor = fornecedor
                material.NumeroNF = numero_nf
            else:
                material = Materiais(
                    CodMaterial=preenchimento.solicitacao.cod_material,
                    DescricaoMaterial=preenchimento.solicitacao.material.DescricaoMaterial,
                    Empresa=preenchimento.solicitacao.empresa,
                    Aplicacao=preenchimento.solicitacao.material.Aplicacao,
                    QuantidadeEstoque=preenchimento.solicitacao.quantidade,
                    Fornecedor=fornecedor,
                    NumeroNF=numero_nf,
                    FatorConsumo=0.0
                )
                db.session.add(material)
            
            estoque = app.jinja_env.globals['Estoque'](
                preenchimento_id=preenchimento_id,
                cod_material=preenchimento.solicitacao.cod_material,
                quantidade=preenchimento.solicitacao.quantidade,
                fornecedor=fornecedor,
                numero_nf=numero_nf,
                usuario=session['usuario']
            )
            db.session.add(estoque)
            db.session.commit()

            flash('Material adicionado ao estoque com sucesso.', 'success')
            return redirect(url_for('routes_bp.listar_estoque'))
        
        return render_template('adicionar_estoque.html', preenchimento=preenchimento)
    except Exception as e:
        _flash_erro_seguro('Erro ao adicionar ao estoque.', e, app.logger)
        return redirect(url_for('routes_bp.listar_solicitacoes_preenchidas'))

@routes_bp.route('/requisitar_material/<int:preenchimento_id>', methods=['GET', 'POST'])
def requisitar_material(preenchimento_id):
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        preenchimento = app.jinja_env.globals['SolicitacoesPreenchidas'].query.get_or_404(preenchimento_id)
        estoque = app.jinja_env.globals['Estoque'].query.filter_by(preenchimento_id=preenchimento_id).first()
        
        if not estoque:
            flash('Material não está no estoque.', 'error')
            return redirect(url_for('routes_bp.listar_estoque'))
        
        if request.method == 'POST':
            quantidade = request.form.get('quantidade', '').strip()

            if not quantidade:
                flash('Quantidade é obrigatória.', 'error')
                return render_template('requisitar_material.html', preenchimento=preenchimento, estoque=estoque)
            
            try:
                quantidade = int(quantidade)
                if quantidade <= 0:
                    flash('Quantidade deve ser um número positivo.', 'error')
                    return render_template('requisitar_material.html', preenchimento=preenchimento, estoque=estoque)
                
                if quantidade > estoque.quantidade:
                    flash(f'Quantidade requisitada ({quantidade}) excede o estoque disponível ({estoque.quantidade}).', 'error')
                    return render_template('requisitar_material.html', preenchimento=preenchimento, estoque=estoque)
            except ValueError:
                flash('Quantidade deve ser um número válido.', 'error')
                return render_template('requisitar_material.html', preenchimento=preenchimento, estoque=estoque)
            
            material = Materiais.query.filter_by(CodMaterial=preenchimento.solicitacao.cod_material).first()
            if not material:
                flash(f'Material com código {preenchimento.solicitacao.cod_material} não encontrado.', 'error')
                return render_template('requisitar_material.html', preenchimento=preenchimento, estoque=estoque)
            
            if quantidade > material.QuantidadeEstoque:
                flash(f'Quantidade requisitada ({quantidade}) excede o estoque total do material ({material.QuantidadeEstoque}).', 'error')
                return render_template('requisitar_material.html', preenchimento=preenchimento, estoque=estoque)
            
            ticket = str(uuid.uuid4())[:8]
            requisicao = app.jinja_env.globals['Requisicoes'](
                preenchimento_id=preenchimento_id,
                cod_material=preenchimento.solicitacao.cod_material,
                quantidade=quantidade,
                ticket=ticket,
                usuario=session['usuario']
            )
            
            estoque.quantidade -= quantidade
            material.QuantidadeEstoque -= quantidade
            
            db.session.add(requisicao)
            db.session.commit()

            flash(f'Requisição realizada com sucesso. Ticket: {ticket}', 'success')
            return redirect(url_for('routes_bp.listar_requisicoes'))
        
        return render_template('requisitar_material.html', preenchimento=preenchimento, estoque=estoque)
    except Exception as e:
        db.session.rollback()
        _flash_erro_seguro('Erro ao requisitar material.', e, app.logger)
        return redirect(url_for('routes_bp.listar_estoque'))

@routes_bp.route('/listar_estoque', methods=['GET'])
def listar_estoque():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))

    try:
        # Uma única query com LEFT JOIN — sem loop de queries secundárias
        from sqlalchemy.orm import outerjoin
        rows = (
            db.session.query(Materiais, Estoque)
            .outerjoin(Estoque, Estoque.cod_material == Materiais.CodMaterial)
            .filter(Materiais.QuantidadeEstoque > 0)
            .all()
        )
        estoques = [
            {'material': m, 'preenchimento_id': e.preenchimento_id if e else None}
            for m, e in rows
        ]
        return render_template('listar_estoque.html', estoques=estoques)
    except Exception as e:
        app.logger.error(f'Erro em listar_estoque: {type(e).__name__}')
        flash('Erro ao carregar estoque.', 'error')
        return render_template('listar_estoque.html', estoques=[])
    
#Auditoria tela de listar estoque
@routes_bp.route('/listar_estoque_auditor', methods=['GET'])
def listar_estoque_auditor():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))

    try:
        rows = (
            db.session.query(Materiais, Estoque)
            .outerjoin(Estoque, Estoque.cod_material == Materiais.CodMaterial)
            .filter(Materiais.QuantidadeEstoque > 0)
            .all()
        )
        estoques = [
            {'material': m, 'preenchimento_id': e.preenchimento_id if e else None}
            for m, e in rows
        ]
        return render_template('listar_estoque_auditor.html', estoques=estoques)
    except Exception as e:
        app.logger.error(f'Erro em listar_estoque_auditor: {type(e).__name__}')
        flash('Erro ao carregar estoque.', 'error')
        return render_template('listar_estoque_auditor.html', estoques=[])

@routes_bp.route('/requisitar_manual/<int:cod_material>', methods=['GET', 'POST'])
def requisitar_manual(cod_material):
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        material = Materiais.query.get_or_404(cod_material)
        
        if request.method == 'POST':
            quantidade = request.form.get('quantidade', '').strip()
            
            if not quantidade:
                flash('Quantidade é obrigatória.', 'error')
                return render_template('requisitar_manual.html', material=material)
            
            try:
                quantidade = int(quantidade)
                if quantidade <= 0:
                    flash('Quantidade deve ser um número positivo.', 'error')
                    return render_template('requisitar_manual.html', material=material)
                
                if quantidade > material.QuantidadeEstoque:
                    flash(f'Quantidade requisitada ({quantidade}) excede o estoque disponível ({material.QuantidadeEstoque}).', 'error')
                    return render_template('requisitar_manual.html', material=material)
                
                # Cria um ticket único
                ticket = str(uuid.uuid4())[:8]
                
                # Cria a requisição com preenchimento_id como NULL
                requisicao = Requisicoes(
                    preenchimento_id=None,  # Explicitamente definido como NULL
                    cod_material=material.CodMaterial,
                    quantidade=quantidade,
                    ticket=ticket,
                    usuario=session['usuario']
                )
                
                # Atualiza o estoque
                material.QuantidadeEstoque -= quantidade
                
                db.session.add(requisicao)
                db.session.commit()

                flash(f'Requisição realizada com sucesso. Ticket: {ticket}', 'success')
                return redirect(url_for('routes_bp.listar_requisicoes'))
                
            except ValueError:
                flash('Quantidade deve ser um número válido.', 'error')
                return render_template('requisitar_manual.html', material=material)
        
        return render_template('requisitar_manual.html', material=material)
        
    except Exception as e:
        db.session.rollback()
        _flash_erro_seguro('Erro ao requisitar material.', e, app.logger)
        return redirect(url_for('routes_bp.listar_estoque'))

@routes_bp.route('/listar_requisicoes', methods=['GET'])
def listar_requisicoes():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        requisicoes = app.jinja_env.globals['Requisicoes'].query.all()
        for requisicao in requisicoes:
            estoque = app.jinja_env.globals['Estoque'].query.filter_by(cod_material=requisicao.cod_material).first()
            if estoque:
                requisicao.fornecedor = estoque.fornecedor
                requisicao.numero_nf = estoque.numero_nf
            else:
                requisicao.fornecedor = 'Não disponível'
                requisicao.numero_nf = 'Não disponível'
        return render_template('listar_requisicoes.html', requisicoes=requisicoes)
    except Exception as e:
        _flash_erro_seguro('Erro ao carregar requisições.', e, app.logger)
        return render_template('listar_requisicoes.html', requisicoes=[])
    
def atualizar_estrutura_requisicoes():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # Verifica se a coluna já permite NULL
        cursor.execute("PRAGMA table_info(Requisicoes)")
        columns = cursor.fetchall()
        preenchimento_id_col = next((col for col in columns if col[1] == 'preenchimento_id'), None)
        
        if preenchimento_id_col and preenchimento_id_col[3] == 1:  # 1 significa NOT NULL
            print("Atualizando estrutura da tabela Requisicoes...")
            
            # Executa cada comando separadamente
            try:
                # 1. Criar tabela temporária
                cursor.execute("CREATE TABLE temp_Requisicoes AS SELECT * FROM Requisicoes")
                
                # 2. Dropar tabela original
                cursor.execute("DROP TABLE Requisicoes")
                
                # 3. Criar nova tabela com estrutura atualizada
                cursor.execute("""
                    CREATE TABLE Requisicoes (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        preenchimento_id INTEGER,
                        cod_material INTEGER NOT NULL,
                        quantidade INTEGER NOT NULL,
                        ticket TEXT NOT NULL,
                        data_requisicao DATETIME NOT NULL,
                        usuario TEXT NOT NULL,
                        FOREIGN KEY (preenchimento_id) REFERENCES SolicitacoesPreenchidas(id),
                        FOREIGN KEY (cod_material) REFERENCES Materiais(CodMaterial)
                    )
                """)
                
                # 4. Copiar dados da tabela temporária
                cursor.execute("""
                    INSERT INTO Requisicoes (id, preenchimento_id, cod_material, quantidade, ticket, data_requisicao, usuario)
                    SELECT id, preenchimento_id, cod_material, quantidade, ticket, data_requisicao, usuario 
                    FROM temp_Requisicoes
                """)
                
                # 5. Dropar tabela temporária
                cursor.execute("DROP TABLE temp_Requisicoes")
                
                conn.commit()
                print("Estrutura da tabela Requisicoes atualizada com sucesso.")
            except sqlite3.Error as e:
                conn.rollback()
                print(f"Erro durante a migração: {str(e)}")
                return False
        else:
            print("A coluna preenchimento_id já permite valores NULL ou não foi encontrada.")
            
        conn.close()
        return True
    except Exception as e:
        print(f"Erro ao atualizar estrutura: {str(e)}")
        return False
    
@routes_bp.route('/pcp_material/<int:cod>', methods=['GET', 'POST'])
def pcp_material(cod):
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))

    material = Materiais.query.get_or_404(cod)

    if request.method == 'POST':
        status = request.form.get('status')  # 'pcp', 'pcm' ou 'inativo'
        fator_consumo = request.form.get('fator_consumo', '0').replace(',', '.')

        try:
            fator_consumo_float = float(fator_consumo) if fator_consumo else 0.0
            
            # Lógica das opções
            if status == 'pcp':
                material.Ativo = True
                material.FatorConsumo = fator_consumo_float
                redirect_to = 'listar_pcp'
            elif status == 'pcm':
                material.Ativo = False
                material.FatorConsumo = fator_consumo_float
                redirect_to = 'listar_ativos'
            else:  # inativo
                material.Ativo = False
                material.FatorConsumo = 0.0
                redirect_to = 'listar_estoque'
            
            db.session.commit()
            flash('Configurações salvas com sucesso!', 'success')
            return redirect(url_for(f'routes_bp.{redirect_to}'))

        except ValueError:
            flash('Valor do fator de consumo inválido', 'error')
        except Exception as e:
            db.session.rollback()
            _flash_erro_seguro('Erro ao salvar.', e, app.logger)

    # Calcula dias de estoque
    dias_estoque = material.QuantidadeEstoque / material.FatorConsumo if material.FatorConsumo > 0 else 0
    
    # Determina o status atual para seleção no formulário
    status_atual = 'inativo'
    if material.FatorConsumo > 0:
        status_atual = 'pcp' if material.Ativo else 'pcm'
    
    return render_template('pcp_material.html', 
                        material=material,
                        dias_estoque=dias_estoque,
                        status_atual=status_atual)


@routes_bp.route('/listar_pcp', methods=['GET'])
def listar_pcp():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        # Mostrar apenas materiais PCP (Ativo=True + FatorConsumo>0)
        materiais = Materiais.query.filter(
            Materiais.Ativo == True,
            Materiais.FatorConsumo > 0
        ).all()
        
        for material in materiais:
            material.dias_estoque = material.QuantidadeEstoque / material.FatorConsumo
            
        return render_template('listar_pcp.html', 
                            materiais=materiais,
                            total_materiais=len(materiais))
    except Exception as e:
        _flash_erro_seguro('Erro ao carregar PCP.', e, app.logger)
        return render_template('listar_pcp.html', 
                            materiais=[], 
                            total_materiais=0)

@routes_bp.route('/listar_ativos', methods=['GET'])
def listar_ativos():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        # Mostrar apenas materiais PCM (Ativo=False + FatorConsumo>0)
        materiais = Materiais.query.filter(
            Materiais.Ativo == False,
            Materiais.FatorConsumo > 0
        ).all()
        
        for material in materiais:
            material.dias_estoque = material.QuantidadeEstoque / material.FatorConsumo
            
        return render_template('ativos.html', 
                            materiais=materiais,
                            total_materiais=len(materiais))
    except Exception as e:
        _flash_erro_seguro('Erro ao carregar PCM.', e, app.logger)
        return render_template('ativos.html', 
                            materiais=[], 
                            total_materiais=0)
    
@routes_bp.route('/cadastro-fornecedores.html', methods=['GET'])
def cadastro_fornecedores():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    return render_template('cadastro-fornecedores.html')

@routes_bp.route('/lista_fornecedores.html', methods=['GET'])
def lista_fornecedores():
    if 'usuario' not in session:
        flash('Você precisa estar logado para acessar esta página.', 'warning')
        return redirect(url_for('routes_bp.login'))
    
    try:
        conn = get_db_connection(DB_PATH_FORNECEDORES)
        if not conn:
            flash('Erro ao conectar ao banco de dados de fornecedores.', 'error')
            app.logger.error('Falha ao conectar ao banco de fornecedores')
            return render_template('lista_fornecedores.html', fornecedores=[])
        
        try:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, nome_fantasia, cnpj, telefone, email, endereco, bairro, cidade, estado, contato, materiais
                FROM fornecedores
                ORDER BY nome_fantasia
            ''')
            fornecedores = [dict(row) for row in cursor.fetchall()]
            
            # Adicionar log para depuração
            app.logger.info(f"Fornecedores recuperados: {fornecedores}")
            
            for fornecedor in fornecedores:
                cnpj = fornecedor['cnpj']
                if cnpj and len(cnpj) == 14:
                    fornecedor['cnpj_formatado'] = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
                else:
                    fornecedor['cnpj_formatado'] = cnpj
            
            return render_template('lista_fornecedores.html', 
                                 fornecedores=fornecedores,
                                 total_fornecedores=len(fornecedores))
            
        except sqlite3.Error as db_error:
            app.logger.error(f'Erro no banco de dados: {str(db_error)}')
            flash('Erro ao acessar os dados dos fornecedores.', 'error')
            return render_template('lista_fornecedores.html', fornecedores=[])
            
        finally:
            if conn:
                conn.close()
                
    except Exception as e:
        app.logger.error(f'Erro inesperado em lista_fornecedores: {str(e)}', exc_info=True)
        flash('Ocorreu um erro inesperado ao listar os fornecedores.', 'error')
        return render_template('lista_fornecedores.html', fornecedores=[])
    
@routes_bp.route('/submit_form', methods=['POST'])
def submit_form():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        nome_fantasia = request.form.get('nomeFantasia', '').strip()
        cnpj = request.form.get('cnpj', '').strip()
        telefone = request.form.get('telefone', '').strip()
        email = request.form.get('email', '').strip()
        endereco = request.form.get('endereco', '').strip()
        bairro = request.form.get('bairro', '').strip()
        cidade = request.form.get('cidade', '').strip()
        estado = request.form.get('estado', '').strip()
        contato = request.form.get('contato', '').strip()
        materiais = request.form.get('materiais', '').strip()

        required_fields = {
            'Nome Fantasia': nome_fantasia,
            'CNPJ': cnpj,
            'Telefone': telefone,
            'Email': email,
            'Endereço': endereco,
            'Bairro': bairro,
            'Cidade': cidade,
            'Estado': estado,
            'Contato': contato,
            'Materiais': materiais
        }
        
        for field_name, field_value in required_fields.items():
            if not field_value:
                flash(f'O campo {field_name} é obrigatório.', 'error')
                return redirect(url_for('routes_bp.cadastro_fornecedores'))

        if len(nome_fantasia) < 3:
            flash('Nome Fantasia deve ter pelo menos 3 caracteres.', 'error')
            return redirect(url_for('routes_bp.cadastro_fornecedores'))

        if not validate_cnpj(cnpj):
            flash('CNPJ inválido. Deve conter 14 dígitos.', 'error')
            return redirect(url_for('routes_bp.cadastro_fornecedores'))

        if not validate_email(email):
            flash('Email inválido. Insira um email válido.', 'error')
            return redirect(url_for('routes_bp.cadastro_fornecedores'))

        if len(telefone) < 10:
            flash('Telefone inválido. Deve conter pelo menos 10 dígitos.', 'error')
            return redirect(url_for('routes_bp.cadastro_fornecedores'))

        if len(endereco) < 5:
            flash('Endereço deve ter pelo menos 5 caracteres.', 'error')
            return redirect(url_for('routes_bp.cadastro_fornecedores'))

        if len(bairro) < 3:
            flash('Bairro deve ter pelo menos 3 caracteres.', 'error')
            return redirect(url_for('routes_bp.cadastro_fornecedores'))

        if len(cidade) < 3:
            flash('Cidade deve ter pelo menos 3 caracteres.', 'error')
            return redirect(url_for('routes_bp.cadastro_fornecedores'))

        if len(estado) != 2:
            flash('Estado deve ter exatamente 2 caracteres (UF).', 'error')
            return redirect(url_for('routes_bp.cadastro_fornecedores'))

        if len(contato) < 3:
            flash('Contato deve ter pelo menos 3 caracteres.', 'error')
            return redirect(url_for('routes_bp.cadastro_fornecedores'))

        if len(materiais) < 3:
            flash('Materiais deve ter pelo menos 3 caracteres.', 'error')
            return redirect(url_for('routes_bp.cadastro_fornecedores'))

        # Verificar duplicidade de CNPJ
        conn = get_db_connection(DB_PATH_FORNECEDORES)
        if not conn:
            flash('Erro ao conectar ao banco de dados.', 'error')
            return redirect(url_for('routes_bp.cadastro_fornecedores'))

        cursor = conn.cursor()
        cursor.execute('SELECT cnpj FROM fornecedores WHERE cnpj = ?', (cnpj,))
        if cursor.fetchone():
            conn.close()
            flash('CNPJ já cadastrado.', 'error')
            return redirect(url_for('routes_bp.cadastro_fornecedores'))

        # Inserir fornecedor no banco
        cursor.execute('''
            INSERT INTO fornecedores (nome_fantasia, cnpj, telefone, email, endereco, bairro, cidade, estado, contato, materiais)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (nome_fantasia, cnpj, telefone, email, endereco, bairro, cidade, estado, contato, materiais))
        conn.commit()
        conn.close()
        flash('Fornecedor cadastrado com sucesso.', 'success')
        return redirect(url_for('routes_bp.lista_fornecedores'))

    except sqlite3.Error as e:
        if 'conn' in locals() and conn:
            conn.close()
        _flash_erro_seguro('Erro ao cadastrar fornecedor.', e, app.logger)
        return redirect(url_for('routes_bp.cadastro_fornecedores'))
    except Exception as e:
        if 'conn' in locals() and conn:
            conn.close()
        _flash_erro_seguro('Erro ao cadastrar fornecedor.', e, app.logger)
        return redirect(url_for('routes_bp.cadastro_fornecedores'))

# Definição da rota /search_fornecedores
@routes_bp.route('/search_fornecedores', methods=['GET'])
def search_fornecedores():
    try:
        query = request.args.get('q', '').strip()
        if not query:
            return jsonify([])   # 200 com lista vazia — evita quebrar o Select2
        
        conn = get_db_connection(DB_PATH_FORNECEDORES)
        if not conn:
            app.logger.error('Falha ao conectar ao banco de fornecedores')
            return jsonify([]), 500
        
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id, nome_fantasia, cnpj, telefone, email, endereco, 
                   bairro, cidade, estado, contato, materiais
            FROM fornecedores
            WHERE nome_fantasia LIKE ? OR cnpj LIKE ?
            ORDER BY nome_fantasia
        ''', (f'%{query}%', f'%{query}%'))
        
        fornecedores = []
        for row in cursor.fetchall():
            fornecedor = {
                'id': row['id'],
                'nome_fantasia': row['nome_fantasia'],
                'cnpj': row['cnpj'],
                'telefone': row['telefone'],
                'email': row['email'],
                'endereco': row['endereco'],
                'bairro': row['bairro'],
                'cidade': row['cidade'],
                'estado': row['estado'],
                'contato': row['contato'],
                'materiais': row['materiais']
            }
            fornecedores.append(fornecedor)
        
        conn.close()
        return jsonify(fornecedores)
        
    except sqlite3.Error as e:
        app.logger.error(f'Erro ao buscar fornecedores: {str(e)}')
        return jsonify([]), 500
    except Exception as e:
        app.logger.error(f'Erro inesperado em search_fornecedores: {str(e)}')
        return jsonify([]), 500
    
    
def verificar_coluna_observacoes():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(PedidosCompra)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'observacoes' not in columns:
            cursor.execute("ALTER TABLE PedidosCompra ADD COLUMN observacoes TEXT")
            conn.commit()
            logging.info("Coluna observacoes adicionada à tabela PedidosCompra")
            print("✓ Coluna observacoes adicionada com sucesso!")
        
        conn.close()
    except sqlite3.Error as e:
        logging.error(f"Erro ao verificar coluna observacoes: {str(e)}")

# Rota temporária para diagnóstico - pode remover depois
@routes_bp.route('/verificar_campo_ativo')
def verificar_campo_ativo():
    try:
        # Verifica se a coluna Ativo existe na tabela Materiais
        result = db.engine.execute("PRAGMA table_info(Materiais)").fetchall()
        colunas = [col[1] for col in result]
        existe = 'Ativo' in colunas
        
        return jsonify({
            'campo_ativo_existe': existe,
            'colunas': colunas,
            'materiais_ativos': Materiais.query.filter_by(Ativo=True).count()
        })
    except Exception as e:
        app.logger.error(f'API error: {type(e).__name__}', exc_info=False)
        return jsonify({'error': 'Erro interno. Tente novamente.'}), 500

#------------------------------------------Ativos 
# Adicione no início do arquivo, com as outras constantes
REGISTRO_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'registro.txt')

# Função para ler registros
def ler_registros():
    registros = []
    try:
        if not os.path.exists(REGISTRO_FILE):
            return registros
            
        with open(REGISTRO_FILE, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line:  # Ignora linhas vazias
                    partes = line.split('|')
                    if len(partes) >= 3:
                        registros.append({
                            'nome': partes[0],
                            'descricao': partes[1],
                            'empresa': partes[2],
                            'data': partes[3] if len(partes) > 3 else ''
                        })
    except Exception as e:
        logging.error(f"Erro ao ler registro.txt: {str(e)}")
    return registros

# Função para adicionar registro
def adicionar_registro(nome, descricao, empresa):
    try:
        data_atual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(REGISTRO_FILE, 'a', encoding='utf-8') as f:
            f.write(f"{nome}|{descricao}|{empresa}|{data_atual}\n")
        return True
    except Exception as e:
        logging.error(f"Erro ao adicionar registro: {str(e)}")
        return False

# Adicione ao jinja_env.globals
app.jinja_env.globals.update(
    ler_registros=ler_registros,
    adicionar_registro=adicionar_registro
)

@routes_bp.route('/registros', methods=['GET'], endpoint='registros')
def listar_registros():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    registros = ler_registros()
    return render_template('registros.html', registros=registros)


@routes_bp.route('/adicionar_registro', methods=['POST'])
def adicionar_registro_route():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    nome = request.form.get('nome', '').strip()
    descricao = request.form.get('descricao', '').strip()
    empresa = request.form.get('empresa', '').strip()
    
    # Validações básicas
    if not all([nome, descricao, empresa]):
        flash('Todos os campos são obrigatórios.', 'error')
        return redirect(url_for('routes_bp.registros'))
    
    # Verifica duplicados
    registros = ler_registros()
    for registro in registros:
        if (registro['nome'].lower() == nome.lower() and 
            registro['empresa'].lower() == empresa.lower()):
            flash('Já existe um registro com este nome e empresa!', 'error')
            return redirect(url_for('routes_bp.registros'))
    
    # Adiciona se não for duplicado
    if adicionar_registro(nome, descricao, empresa):
        flash('Registro adicionado com sucesso!', 'success')
    else:
        flash('Erro ao adicionar registro.', 'error')
    
    return redirect(url_for('routes_bp.registros'))


# Função para remover um registro específico
def remover_registro(linha_para_remover):
    try:
        # Lê todos os registros
        with open(REGISTRO_FILE, 'r', encoding='utf-8') as f:
            linhas = f.readlines()
        
        # Remove a linha específica
        linhas = [linha for linha in linhas if linha.strip() != linha_para_remover.strip()]
        
        # Reescreve o arquivo sem a linha removida
        with open(REGISTRO_FILE, 'w', encoding='utf-8') as f:
            f.writelines(linhas)
        
        return True
    except Exception as e:
        logging.error(f"Erro ao remover registro: {str(e)}")
        return False
    
@routes_bp.route('/excluir_registro', methods=['POST'])
def excluir_registro_route():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    # Obtém os dados do formulário
    nome = request.form.get('nome', '').strip()
    descricao = request.form.get('descricao', '').strip()
    empresa = request.form.get('empresa', '').strip()
    data = request.form.get('data', '').strip()
    
    # Reconstroi a linha no formato do arquivo
    linha_para_remover = f"{nome}|{descricao}|{empresa}|{data}\n"
    
    if remover_registro(linha_para_remover):
        flash('Registro excluído com sucesso!', 'success')
    else:
        flash('Erro ao excluir registro.', 'error')
    
    return redirect(url_for('routes_bp.registros'))

@routes_bp.route('/registros')
def get_registros():
    try:
        registros = ler_registros()
        
        # Verifica se há registros válidos
        if not registros:
            return jsonify({"error": "Nenhum registro válido encontrado"}), 404
            
        return jsonify({
            "success": True,
            "data": registros
        })
        
    except Exception as e:
        logging.error(f"Erro na rota /registros: {str(e)}")
        return jsonify({
            "error": "Erro interno ao processar registros",
            "details": str(e)
        }), 500
    
# Adicione esta função (se já não existir)
def ler_nomes_registros():
    nomes = set()
    try:
        if not os.path.exists(REGISTRO_FILE):
            return sorted(nomes)
            
        with open(REGISTRO_FILE, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line:  # Ignora linhas vazias
                    partes = line.split('|')
                    if len(partes) >= 1:
                        nomes.add(partes[0])  # Adiciona o primeiro campo (nome)
    except Exception as e:
        logging.error(f"Erro ao ler nomes do registro.txt: {str(e)}")
    return sorted(nomes)

# Adicione ao jinja_env.globals (procure por app.jinja_env.globals.update e adicione)
app.jinja_env.globals.update(
    ler_nomes_registros=ler_nomes_registros,
    # ... outros globais que já existam
)

@routes_bp.route('/buscar_ativos', methods=['GET'])
def buscar_ativos():
    termo = request.args.get('q', '').lower()
    registros = ler_registros()
    resultados = [{'nome': r['nome']} for r in registros if termo in r['nome'].lower()]
    return jsonify(resultados)

#Tela auditoria -----------------------------------------------------------
@routes_bp.route('/auditoria_solicitacoes', methods=['GET'])
def auditoria_solicitacoes():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        # Parâmetros de filtro
        empresa = request.args.get('empresa')
        usuario = request.args.get('usuario')
        ativo = request.args.get('ativo')
        nome_ativo = request.args.get('nome_ativo')
        status = request.args.get('status')
        prioridade_filtro = request.args.get('prioridade')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        # OBTER USUÁRIO LOGADO
        usuario_logado = session.get('usuario')
        
        # Se não foi especificado um usuário no filtro, usar o usuário logado por padrão
        if not usuario and usuario_logado:
            usuario = usuario_logado

        # Query base para buscar TODAS as solicitações
        query = db.session.query(SolicitacoesCompra)
        
        # Aplicar filtros básicos
        if empresa and empresa != 'Todas':
            query = query.filter(SolicitacoesCompra.empresa == empresa)
        if usuario and usuario != 'Todos':
            query = query.filter(SolicitacoesCompra.usuario == usuario)
        if ativo:
            query = query.filter(SolicitacoesCompra.ativo == ativo)
        if nome_ativo and ativo == 'Sim':
            query = query.filter(SolicitacoesCompra.nome_ativo.ilike(f'%{nome_ativo}%'))
        if data_inicio:
            query = query.filter(SolicitacoesCompra.data_solicitacao >= data_inicio)
        if data_fim:
            data_fim_ajustada = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(SolicitacoesCompra.data_solicitacao <= data_fim_ajustada)

        # Executar query — buscar todas as solicitações que passam nos filtros
        solicitacoes = query.options(
            joinedload(SolicitacoesCompra.material)
        ).order_by(SolicitacoesCompra.data_solicitacao.desc()).all()

        # ── Coletar IDs para bulk queries ────────────────────────────────────
        sol_ids = [s.id for s in solicitacoes]

        # Preenchimentos em bloco (excluindo rascunhos)
        preenchimentos_por_sol: dict[int, list] = {sid: [] for sid in sol_ids}
        todos_preenchimentos = (
            SolicitacoesPreenchidas.query
            .filter(
                SolicitacoesPreenchidas.solicitacao_id.in_(sol_ids),
                SolicitacoesPreenchidas.status != 'Rascunho'
            ).all()
        )
        preench_ids = [p.id for p in todos_preenchimentos]
        for p in todos_preenchimentos:
            preenchimentos_por_sol[p.solicitacao_id].append(p)

        # Pedidos associados em bloco
        pedido_por_preenchimento: dict[int, object] = {}
        if preench_ids:
            rows_pedidos = db.session.query(
                pedido_preenchimento_associacao.c.preenchimento_id,
                PedidosCompra
            ).join(
                PedidosCompra,
                PedidosCompra.id == pedido_preenchimento_associacao.c.pedido_id
            ).filter(
                pedido_preenchimento_associacao.c.preenchimento_id.in_(preench_ids)
            ).all()
            for pid, pedido_obj in rows_pedidos:
                pedido_por_preenchimento.setdefault(pid, pedido_obj)

        # Estoque em bloco
        estoque_por_preenchimento: dict[int, object] = {}
        if preench_ids:
            for est in Estoque.query.filter(Estoque.preenchimento_id.in_(preench_ids)).all():
                estoque_por_preenchimento[est.preenchimento_id] = est

        # Requisições em bloco
        requisicoes_por_preenchimento: dict[int, list] = {pid: [] for pid in preench_ids}
        if preench_ids:
            for req in Requisicoes.query.filter(Requisicoes.preenchimento_id.in_(preench_ids)).all():
                requisicoes_por_preenchimento[req.preenchimento_id].append(req)

        # Fornecedores em bloco — UMA única conexão, sem loop
        fornecedor_ids_bulk = {p.fornecedor_id for p in todos_preenchimentos if p.fornecedor_id}
        fornecedores_bulk = get_fornecedores_bulk(fornecedor_ids_bulk)

        # ── Mapa de prioridade por (status_aprovacao, preenchimento.status) ─
        _PRIO_SEM_PREENCH = {
            'Reprovado': 0, 'Aprovado': 1,
        }
        _PRIO_COM_PREENCH = {
            'Entregue': 3, ('Aprovado', True): 2, 'Aprovado': 1,
        }

        # ── Montar auditoria_data sem nenhuma query adicional ────────────────
        auditoria_data = []
        for solicitacao in solicitacoes:
            preenchimentos = preenchimentos_por_sol.get(solicitacao.id, [])
            material = solicitacao.material  # já eager-loaded

            if not preenchimentos:
                # Caso 1: sem preenchimentos
                status_sol = (solicitacao.status_aprovacao or
                              getattr(solicitacao, 'status', None) or 'Aberta')
                if status_sol not in ('Reprovado', 'Aprovado', 'Pendente'):
                    status_sol = 'Aberta'
                prioridade = _PRIO_SEM_PREENCH.get(status_sol, 0)

                if status and status != 'Todos' and status != status_sol:
                    continue
                if prioridade_filtro and prioridade_filtro != '':
                    try:
                        if int(prioridade_filtro) != prioridade:
                            continue
                    except (ValueError, TypeError):
                        pass

                auditoria_data.append({
                    'solicitacao': solicitacao, 'material': material,
                    'preenchimento': None, 'pedido': None, 'estoque': None,
                    'requisicoes': [], 'fornecedor': {}, 'prioridade': prioridade,
                    'status_determinado': status_sol,
                })

            else:
                # Caso 2: com preenchimentos
                for preenchimento in preenchimentos:
                    pedido   = pedido_por_preenchimento.get(preenchimento.id)
                    estoque  = estoque_por_preenchimento.get(preenchimento.id)
                    reqs     = requisicoes_por_preenchimento.get(preenchimento.id, [])
                    forn     = fornecedores_bulk.get(preenchimento.fornecedor_id, {})

                    # Filtro de status
                    if status and status != 'Todos':
                        status_match = (
                            preenchimento.status == status or
                            (solicitacao.status_aprovacao == status)
                        )
                        if not status_match:
                            continue

                    # Prioridade
                    if solicitacao.status_aprovacao == 'Reprovado':
                        prioridade = 0
                    elif preenchimento.status == 'Entregue':
                        prioridade = 3
                    elif preenchimento.status == 'Aprovado' and pedido:
                        prioridade = 2
                    elif preenchimento.status == 'Aprovado':
                        prioridade = 1
                    else:
                        prioridade = 0

                    if prioridade_filtro and prioridade_filtro != '':
                        try:
                            if int(prioridade_filtro) != prioridade:
                                continue
                        except (ValueError, TypeError):
                            pass

                    status_det = (solicitacao.status_aprovacao
                                  if solicitacao.status_aprovacao
                                  else preenchimento.status)
                    auditoria_data.append({
                        'solicitacao': solicitacao, 'material': material,
                        'preenchimento': preenchimento, 'pedido': pedido,
                        'estoque': estoque, 'requisicoes': reqs,
                        'fornecedor': forn, 'prioridade': prioridade,
                        'status_determinado': status_det,
                    })

        # Ordenar por prioridade (mais alta primeiro) e depois por data
        auditoria_data.sort(key=lambda x: (-x['prioridade'], x['solicitacao'].data_solicitacao), reverse=True)

        # Obter valores para filtros
        empresas = db.session.query(
            SolicitacoesCompra.empresa
        ).distinct().order_by(
            SolicitacoesCompra.empresa
        ).all()

        usuarios = db.session.query(
            SolicitacoesCompra.usuario
        ).distinct().order_by(
            SolicitacoesCompra.usuario
        ).all()

        nomes_ativos = db.session.query(
            SolicitacoesCompra.nome_ativo
        ).filter(
            SolicitacoesCompra.ativo == 'Sim',
            SolicitacoesCompra.nome_ativo.isnot(None)
        ).distinct().order_by(
            SolicitacoesCompra.nome_ativo
        ).all()

        total_registros = len(auditoria_data)

        # Passar o usuário logado para o template
        usuario_logado = session.get('usuario')

        return render_template(
            'auditoria_solicitacoes.html',
            auditoria=auditoria_data,
            empresas=[e[0] for e in empresas if e[0]],
            usuarios=[u[0] for u in usuarios if u[0]],
            nomes_ativos=[n[0] for n in nomes_ativos if n[0]],
            total_registros=total_registros,
            usuario_logado=usuario_logado,
            filtros={
                'empresa': empresa,
                'usuario': usuario,
                'ativo': ativo,
                'nome_ativo': nome_ativo,
                'status': status,
                'prioridade': prioridade_filtro,
                'data_inicio': data_inicio,
                'data_fim': data_fim
            }
        )
    except Exception as e:
        _flash_erro_seguro('Erro ao carregar página de auditoria.', e, app.logger)
        app.logger.error(f'Erro em auditoria_solicitacoes: {str(e)}')
        return render_template(
            'auditoria_solicitacoes.html',
            auditoria=[],
            empresas=[],
            usuarios=[],
            nomes_ativos=[],
            total_registros=0,
            usuario_logado=session.get('usuario'),
            filtros={}
        )

@routes_bp.route('/exportar_auditoria_xlsx', methods=['GET'])
def exportar_auditoria_xlsx():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        # ────────────────────────────────────────────────
        # Parâmetros vindos da URL (mesmos da tela de auditoria)
        # ────────────────────────────────────────────────
        empresa         = request.args.get('empresa')
        usuario         = request.args.get('usuario')
        ativo           = request.args.get('ativo')
        nome_ativo      = request.args.get('nome_ativo')
        status          = request.args.get('status')
        prioridade_filtro = request.args.get('prioridade')
        data_inicio     = request.args.get('data_inicio')
        data_fim        = request.args.get('data_fim')

        # Log para debug
        app.logger.info(f"=== INÍCIO EXPORTAÇÃO ===")
        app.logger.info(f"Parâmetro usuario recebido: '{usuario}'")
        app.logger.info(f"Tipo do parâmetro: {type(usuario)}")

        # CORREÇÃO: Verificar corretamente se é para exportar TODOS os usuários
        # Se usuario for vazio, None, string vazia ou 'Todos' (case insensitive)
        exportar_todos = (usuario is None or 
                         usuario == '' or 
                         usuario.lower() == 'todos' or 
                         usuario.lower() == 'todos os usuários')
        
        app.logger.info(f"exportar_todos: {exportar_todos}")

        # ────────────────────────────────────────────────
        # Query base
        # ────────────────────────────────────────────────
        query = db.session.query(SolicitacoesCompra).outerjoin(
            SolicitacoesPreenchidas,
            SolicitacoesCompra.id == SolicitacoesPreenchidas.solicitacao_id
        )

        # Filtros
        if empresa and empresa != '' and empresa != 'Todas':
            query = query.filter(SolicitacoesCompra.empresa == empresa)
            app.logger.info(f"Aplicando filtro empresa: {empresa}")

        # CORREÇÃO CRÍTICA: Aplicar filtro de usuário APENAS se NÃO for exportar todos
        if not exportar_todos:
            # Se não for exportar todos, filtra pelo usuário específico
            if usuario and usuario.strip():
                query = query.filter(SolicitacoesCompra.usuario == usuario)
                app.logger.info(f"Aplicando filtro para usuário específico: '{usuario}'")
            else:
                app.logger.warning(f"exportar_todos é False mas usuario está vazio: '{usuario}'")
        else:
            app.logger.info("Exportando TODOS os usuários (filtro de usuário ignorado)")

        if ativo and ativo != '':
            query = query.filter(SolicitacoesCompra.ativo == ativo)
            app.logger.info(f"Aplicando filtro ativo: {ativo}")

        if nome_ativo and nome_ativo.strip() and ativo == 'Sim':
            query = query.filter(SolicitacoesCompra.nome_ativo.ilike(f'%{nome_ativo}%'))
            app.logger.info(f"Aplicando filtro nome_ativo: {nome_ativo}")

        if data_inicio and data_inicio.strip():
            query = query.filter(SolicitacoesCompra.data_solicitacao >= data_inicio)
            app.logger.info(f"Aplicando filtro data_inicio: {data_inicio}")

        if data_fim and data_fim.strip():
            try:
                dt_fim = datetime.strptime(data_fim, '%Y-%m-%d')
                data_fim_ajustada = dt_fim + timedelta(days=1)
                query = query.filter(SolicitacoesCompra.data_solicitacao < data_fim_ajustada)
                app.logger.info(f"Aplicando filtro data_fim: {data_fim} (ajustada: {data_fim_ajustada})")
            except Exception as e:
                app.logger.error(f"Erro ao processar data_fim: {e}")

        # Busca os registros — limitado para evitar consumo excessivo de memória
        EXPORT_ROW_LIMIT = 5000
        solicitacoes = query.order_by(
            SolicitacoesCompra.data_solicitacao.desc()
        ).distinct().limit(EXPORT_ROW_LIMIT).all()

        app.logger.info(f"Export auditoria: {len(solicitacoes)} solicitações encontradas.")
        if len(solicitacoes) == EXPORT_ROW_LIMIT:
            app.logger.warning(f'Export de auditoria atingiu limite de {EXPORT_ROW_LIMIT} linhas.')

        # ────────────────────────────────────────────────
        # Preparar dados para o Excel
        # ────────────────────────────────────────────────
        dados = []
        contador_solicitacoes = 0
        contador_preenchimentos = 0

        for solicitacao in solicitacoes:
            contador_solicitacoes += 1
            material = db.session.get(Materiais, solicitacao.cod_material)

            # Buscar preenchimentos (excluindo rascunhos)
            preenchimentos_query = db.session.query(SolicitacoesPreenchidas).filter_by(
                solicitacao_id=solicitacao.id
            ).filter(
                SolicitacoesPreenchidas.status != 'Rascunho'
            )
            
            preenchimentos = preenchimentos_query.all()
            
            app.logger.debug(f"Solicitação {solicitacao.id} - {len(preenchimentos)} preenchimentos encontrados")

            # Caso não tenha preenchimento → trata como 'Aberta'
            if not preenchimentos:
                if status and status.strip() and status != 'Todos' and status != 'Aberta':
                    continue
                
                prioridade = 0
                nivel_prioridade = 'Baixa (Outras)'
                status_formatado = 'Aberta'
                fornecedor_nome = ''
                fornecedor_cnpj = ''
                pedido_numero = ''
                pedido_status = ''
                valor_unitario = 0
                valor_total = 0
                valor_frete = 0
                prazo_entrega = ''
                condicao_pagamento = ''
                observacoes = ''
                
                if prioridade_filtro and prioridade_filtro.strip():
                    if int(prioridade_filtro) != prioridade:
                        continue
                
                dados.append({
                    'ID_Solicitacao': solicitacao.id,
                    'ID_Cotacao': '',
                    'Material': material.DescricaoMaterial if material else 'Material não encontrado',
                    'Cod_Material': material.CodMaterial if material else '',
                    'Especificacao': solicitacao.especificacao,
                    'Marca': solicitacao.marca or '',
                    'Solicitante': solicitacao.usuario,
                    'Comprador_Atribuido': solicitacao.comprador_atribuido or '',
                    'Empresa': solicitacao.empresa,
                    'Ativo': 'Sim' if solicitacao.ativo == 'Sim' else 'Não',
                    'Nome_Ativo': solicitacao.nome_ativo if solicitacao.ativo == 'Sim' else '',
                    'Quantidade': solicitacao.quantidade,
                    'Unidade_Medida': solicitacao.unidade_medida,
                    'Prioridade_Original': solicitacao.prioridade,
                    'Nivel_Prioridade': nivel_prioridade,
                    'Valor_Prioridade': prioridade,
                    'Data_Solicitacao': solicitacao.data_solicitacao.strftime('%d/%m/%Y %H:%M') if solicitacao.data_solicitacao else '',
                    'Status': status_formatado,
                    'Fornecedor': fornecedor_nome,
                    'CNPJ_Fornecedor': fornecedor_cnpj,
                    'Valor_Unitario': valor_unitario,
                    'Valor_Total': valor_total,
                    'Valor_Frete': valor_frete,
                    'Prazo_Entrega': prazo_entrega,
                    'Condicao_Pagamento': condicao_pagamento,
                    'Pedido_Numero': pedido_numero,
                    'Pedido_Status': pedido_status,
                    'Observacoes': observacoes,
                    'Aprovacao': solicitacao.status_aprovacao or 'Pendente',
                    'Aplicacao': solicitacao.aplicacao or '',
                    'Aplicacao_Geral': solicitacao.aplicacao_geral or '',
                    'Data_Envio_Aprovacao': solicitacao.data_envio_aprovacao.strftime('%d/%m/%Y %H:%M') if solicitacao.data_envio_aprovacao else '',
                    'Dias_Aguardando_Aprovacao': (solicitacao.data_retorno_aprovacao - solicitacao.data_envio_aprovacao).days if (solicitacao.data_envio_aprovacao and solicitacao.data_retorno_aprovacao) else ''
                })
            for preenchimento in preenchimentos:
                contador_preenchimentos += 1
                
                if status and status.strip() and status != 'Todos':
                    if preenchimento.status != status:
                        continue
                
                # Lógica de prioridade (igual ao original)
                prioridade = 0
                if preenchimento.status == 'Entregue':
                    prioridade = 3
                elif preenchimento.status == 'Aprovado':
                    pedido = db.session.query(PedidosCompra).join(
                        pedido_preenchimento_associacao,
                        PedidosCompra.id == pedido_preenchimento_associacao.c.pedido_id
                    ).filter(
                        pedido_preenchimento_associacao.c.preenchimento_id == preenchimento.id
                    ).first()
                    prioridade = 2 if pedido else 1
                
                if prioridade_filtro and prioridade_filtro.strip():
                    if int(prioridade_filtro) != prioridade:
                        continue
                
                # Busca fornecedor
                fornecedor_nome = ''
                fornecedor_cnpj = ''
                if preenchimento.fornecedor_id:
                    conn = get_db_connection(DB_PATH_FORNECEDORES)
                    if conn:
                        cursor = conn.cursor()
                        cursor.execute('SELECT nome_fantasia, cnpj FROM fornecedores WHERE id = ?', (preenchimento.fornecedor_id,))
                        result = cursor.fetchone()
                        if result:
                            fornecedor_nome = result[0] or ''
                            fornecedor_cnpj = format_cnpj(result[1]) if result[1] else ''
                        conn.close()
                
                # Busca pedido associado
                pedido = None
                pedido_numero = ''
                pedido_status = ''
                if preenchimento:
                    pedido = db.session.query(PedidosCompra).join(
                        pedido_preenchimento_associacao,
                        PedidosCompra.id == pedido_preenchimento_associacao.c.pedido_id
                    ).filter(
                        pedido_preenchimento_associacao.c.preenchimento_id == preenchimento.id
                    ).first()
                    if pedido:
                        pedido_numero = pedido.numero_pedido
                        pedido_status = pedido.status or ''
                
                # Formata status
                status_formatado = preenchimento.status or 'Aberta'
                if status_formatado == 'Rascunho':
                    status_formatado = 'Rascunho'
                elif status_formatado == 'Aguardando Aprovação':
                    status_formatado = 'Aguardando Aprovação'
                elif status_formatado == 'Aprovado':
                    status_formatado = 'Aprovado'
                elif status_formatado == 'Reprovado':
                    status_formatado = 'Reprovado'
                elif status_formatado == 'Em Processamento':
                    status_formatado = 'Em Processamento'
                elif status_formatado == 'Entregue':
                    status_formatado = 'Entregue'
                
                nivel_prioridade = {
                    3: 'Máxima (Entregues)',
                    2: 'Alta (Aprovadas + Pedido)',
                    1: 'Média (Aprovadas)',
                    0: 'Baixa (Outras)'
                }.get(prioridade, 'Baixa (Outras)')
                
                dados.append({
                    'ID_Solicitacao': solicitacao.id,
                    'ID_Cotacao': preenchimento.id,
                    'Material': material.DescricaoMaterial if material else 'Material não encontrado',
                    'Cod_Material': material.CodMaterial if material else '',
                    'Especificacao': solicitacao.especificacao,
                    'Marca': solicitacao.marca or '',
                    'Solicitante': solicitacao.usuario,
                    'Comprador_Atribuido': solicitacao.comprador_atribuido or '',
                    'Empresa': solicitacao.empresa,
                    'Ativo': 'Sim' if solicitacao.ativo == 'Sim' else 'Não',
                    'Nome_Ativo': solicitacao.nome_ativo if solicitacao.ativo == 'Sim' else '',
                    'Quantidade': solicitacao.quantidade,
                    'Unidade_Medida': solicitacao.unidade_medida,
                    'Prioridade_Original': solicitacao.prioridade,
                    'Nivel_Prioridade': nivel_prioridade,
                    'Valor_Prioridade': prioridade,
                    'Data_Solicitacao': solicitacao.data_solicitacao.strftime('%d/%m/%Y %H:%M') if solicitacao.data_solicitacao else '',
                    'Status': status_formatado,
                    'Fornecedor': fornecedor_nome,
                    'CNPJ_Fornecedor': fornecedor_cnpj,
                    'Valor_Unitario': preenchimento.valor_unitario or 0,
                    'Valor_Total': preenchimento.valor_total or 0,
                    'Valor_Frete': preenchimento.valor_frete or 0,
                    'Prazo_Entrega': preenchimento.prazo_entrega or '',
                    'Condicao_Pagamento': preenchimento.condicao_pagamento or '',
                    'Pedido_Numero': pedido_numero,
                    'Pedido_Status': pedido_status,
                    'Observacoes': preenchimento.observacoes or '',
                    'Aprovacao': solicitacao.status_aprovacao or 'Pendente',
                    'Aplicacao': solicitacao.aplicacao or '',
                    'Aplicacao_Geral': solicitacao.aplicacao_geral or '',
                    'Data_Envio_Aprovacao': solicitacao.data_envio_aprovacao.strftime('%d/%m/%Y %H:%M') if solicitacao.data_envio_aprovacao else '',
                    'Dias_Aguardando_Aprovacao': (solicitacao.data_retorno_aprovacao - solicitacao.data_envio_aprovacao).days if (solicitacao.data_envio_aprovacao and solicitacao.data_retorno_aprovacao) else ''
                })

        app.logger.info(f"Total de solicitações processadas: {contador_solicitacoes}")
        app.logger.info(f"Total de preenchimentos processados: {contador_preenchimentos}")
        app.logger.info(f"Total de linhas geradas para Excel: {len(dados)}")

        if not dados:
            app.logger.warning("Nenhum dado encontrado para exportar!")
            flash('Nenhum dado encontrado para exportar com os filtros aplicados.', 'warning')
            return redirect(url_for('routes_bp.auditoria_solicitacoes'))

        # ────────────────────────────────────────────────
        # Geração do Excel
        # ────────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Auditoria Solicitações"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = [
            'ID Solicitação', 'ID Cotação', 'Material', 'Cód Material', 'Especificação', 'Marca',
            'Solicitante', 'Comprador Atribuído', 'Empresa', 'Ativo', 'Nome Ativo', 'Quantidade',
            'Unidade', 'Prioridade Original', 'Nível Prioridade', 'Valor Prioridade',
            'Data Solicitação', 'Status', 'Fornecedor', 'CNPJ Fornecedor', 'Valor Unitário',
            'Valor Total', 'Valor Frete', 'Prazo Entrega', 'Condição Pagamento',
            'Número Pedido', 'Status Pedido', 'Observações', 'Aprovação',
            'Aplicação', 'Aplicação Geral', 'Data Envio p/ Aprovação', 'Dias Aguardando Aprovação'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Dados
        for row_num, item in enumerate(dados, 2):
            row = [
                item['ID_Solicitacao'],
                item['ID_Cotacao'],
                item['Material'],
                item['Cod_Material'],
                item['Especificacao'],
                item['Marca'],
                item['Solicitante'],
                item['Comprador_Atribuido'],
                item['Empresa'],
                item['Ativo'],
                item['Nome_Ativo'],
                item['Quantidade'],
                item['Unidade_Medida'],
                item['Prioridade_Original'],
                item['Nivel_Prioridade'],
                item['Valor_Prioridade'],
                item['Data_Solicitacao'],
                item['Status'],
                item['Fornecedor'],
                item['CNPJ_Fornecedor'],
                item['Valor_Unitario'],
                item['Valor_Total'],
                item['Valor_Frete'],
                item['Prazo_Entrega'],
                item['Condicao_Pagamento'],
                item['Pedido_Numero'],
                item['Pedido_Status'],
                item['Observacoes'],
                item['Aprovacao'],
                item['Aplicacao'],
                item['Aplicacao_Geral'],
                item.get('Data_Envio_Aprovacao', ''),
                item.get('Dias_Aguardando_Aprovacao', '')
            ]
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Ajuste automático de largura
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Salvar em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Nome do arquivo indicando se é completo ou filtrado
        if exportar_todos:
            filename = f"auditoria_solicitacoes_TODOS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            # Sanitizar nome do usuário para evitar problemas com caracteres especiais
            usuario_sanitizado = usuario.replace(' ', '_').replace('\\', '_').replace('/', '_')
            filename = f"auditoria_solicitacoes_{usuario_sanitizado}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        from urllib.parse import quote as _url_quote
        response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{_url_quote(filename)}"

        app.logger.info(f"Arquivo gerado com sucesso: {filename}")
        return response

    except Exception as e:
        _flash_erro_seguro('Erro ao exportar relatório.', e, app.logger)
        app.logger.error(f'Erro em exportar_auditoria_xlsx: {str(e)}', exc_info=True)
        return redirect(url_for('routes_bp.auditoria_solicitacoes'))
    
def get_fornecedor_cnpj(fornecedor_id):
    conn = get_db_connection(DB_PATH_FORNECEDORES)
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT cnpj FROM fornecedores WHERE id = ?', (fornecedor_id,))
            result = cursor.fetchone()
            return format_cnpj(result['cnpj']) if result and result['cnpj'] else 'N/A'
        finally:
            conn.close()
    return 'N/A'

# 2° Tela Auditoria
@routes_bp.route('/tela_auditoria', methods=['GET', 'POST'])
def tela_auditoria():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        # Processar formulário de auditoria se for POST
        if request.method == 'POST':
            solicitacao_id = request.form.get('solicitacao_id')
            colaborador_1 = request.form.get('colaborador_1', '').strip()
            colaborador_2 = request.form.get('colaborador_2', '').strip()
            status = request.form.get('status', '').strip()
            observacao = request.form.get('observacao', '').strip()
            
            # Validações
            if not all([solicitacao_id, colaborador_1, status]):
                flash('Campos obrigatórios não preenchidos', 'error')
                return redirect(url_for('routes_bp.tela_auditoria'))
            
            if status not in ['Conforme', 'Não Conforme']:
                flash('Status inválido', 'error')
                return redirect(url_for('routes_bp.tela_auditoria'))
            
            # Criar registro de auditoria
            auditoria = Auditoria(
                solicitacao_id=solicitacao_id,
                colaborador_1=colaborador_1,
                colaborador_2=colaborador_2 if colaborador_2 else None,
                status=status,
                observacao=observacao if observacao else None
            )
            db.session.add(auditoria)
            db.session.commit()
            flash('Auditoria registrada com sucesso!', 'success')
            return redirect(url_for('routes_bp.tela_auditoria'))

        # Parâmetros de paginação e filtro
        page     = request.args.get('page', 1, type=int)
        per_page = 20

        empresa    = request.args.get('empresa', '')
        usuario    = request.args.get('usuario', '')
        ativo      = request.args.get('ativo', '')
        nome_ativo = request.args.get('nome_ativo', '')
        data_inicio = request.args.get('data_inicio', '')
        data_fim    = request.args.get('data_fim', '')
        status      = request.args.get('status', '')

        # ── Query base com filtros ──────────────────────────────────────────
        query = db.session.query(SolicitacoesCompra)

        if empresa:
            query = query.filter(SolicitacoesCompra.empresa == empresa)
        if usuario:
            query = query.filter(SolicitacoesCompra.usuario == usuario)
        if ativo:
            query = query.filter(SolicitacoesCompra.ativo == ativo)
        if nome_ativo and ativo == 'Sim':
            query = query.filter(SolicitacoesCompra.nome_ativo.ilike(f'%{nome_ativo}%'))
        if data_inicio:
            query = query.filter(SolicitacoesCompra.data_solicitacao >= data_inicio)
        if data_fim:
            data_fim_ajustada = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(SolicitacoesCompra.data_solicitacao <= data_fim_ajustada)

        # Filtro de status: precisa join com preenchimentos para filtrar no banco
        # (evita carregar tudo e descartar em Python)
        if status == 'Aberta':
            # solicitações SEM nenhum preenchimento
            query = query.filter(
                ~SolicitacoesCompra.id.in_(
                    db.session.query(SolicitacoesPreenchidas.solicitacao_id).distinct()
                )
            )
        elif status:
            # solicitações com preenchimento no status pedido
            query = query.filter(
                SolicitacoesCompra.id.in_(
                    db.session.query(SolicitacoesPreenchidas.solicitacao_id)
                    .filter(SolicitacoesPreenchidas.status == status)
                    .distinct()
                )
            )

        # ── Paginação: só busca os 20 desta página ─────────────────────────
        paginacao   = query.order_by(
            SolicitacoesCompra.data_solicitacao.desc()
        ).paginate(page=page, per_page=per_page, error_out=False)

        solicitacoes = paginacao.items
        sol_ids      = [s.id for s in solicitacoes]

        if not sol_ids:
            auditoria_data      = []
            empresas_lista      = []
            usuarios_lista      = []
            nomes_ativos_lista  = []
            usuarios_disponiveis = sorted(ler_senhas().keys())
        else:
            # ── Batch: uma query por tabela, indexada por solicitacao_id ───
            # Preenchimentos
            preenchimentos_map = {}
            for p in db.session.query(SolicitacoesPreenchidas).filter(
                SolicitacoesPreenchidas.solicitacao_id.in_(sol_ids)
            ).all():
                preenchimentos_map.setdefault(p.solicitacao_id, p)

            preenч_ids = [p.id for p in preenchimentos_map.values()]

            # Materiais
            mat_ids = list({s.cod_material for s in solicitacoes})
            materiais_map = {
                m.CodMaterial: m
                for m in db.session.query(Materiais).filter(
                    Materiais.CodMaterial.in_(mat_ids)
                ).all()
            }

            # Pedidos (via tabela de associação)
            pedidos_map = {}
            if preenч_ids:
                for pedido, assoc_pid in db.session.query(
                    PedidosCompra,
                    pedido_preenchimento_associacao.c.preenchimento_id
                ).join(
                    pedido_preenchimento_associacao,
                    PedidosCompra.id == pedido_preenchimento_associacao.c.pedido_id
                ).filter(
                    pedido_preenchimento_associacao.c.preenchimento_id.in_(preenч_ids)
                ).all():
                    pedidos_map[assoc_pid] = pedido

            # Estoques
            estoques_map = {}
            if preenч_ids:
                for e in db.session.query(Estoque).filter(
                    Estoque.preenchimento_id.in_(preenч_ids)
                ).all():
                    estoques_map[e.preenchimento_id] = e

            # Requisições
            requisicoes_map = {}
            if preenч_ids:
                for r in db.session.query(Requisicoes).filter(
                    Requisicoes.preenchimento_id.in_(preenч_ids)
                ).all():
                    requisicoes_map.setdefault(r.preenchimento_id, []).append(r)

            # Auditorias
            auditorias_map = {}
            for a in db.session.query(Auditoria).filter(
                Auditoria.solicitacao_id.in_(sol_ids)
            ).order_by(Auditoria.data_validacao.desc()).all():
                auditorias_map.setdefault(a.solicitacao_id, []).append(a)

            # Fornecedores (banco externo) — uma única query em batch
            fornecedor_ids = {
                p.fornecedor_id for p in preenchimentos_map.values()
                if p.fornecedor_id
            }
            fornecedores = {}
            if fornecedor_ids:
                conn_forn = get_db_connection(DB_PATH_FORNECEDORES)
                if conn_forn:
                    try:
                        cur = conn_forn.cursor()
                        cur.execute(
                            f'SELECT id, nome_fantasia, cnpj FROM fornecedores '
                            f'WHERE id IN ({",".join("?" * len(fornecedor_ids))})',
                            list(fornecedor_ids)
                        )
                        for row in cur.fetchall():
                            fornecedores[row[0]] = {
                                'nome_fantasia': row[1],
                                'cnpj': format_cnpj(row[2]) if row[2] else 'N/A'
                            }
                    finally:
                        conn_forn.close()

            # ── Montar lista final sem nenhuma query extra ─────────────────
            auditoria_data = []
            for solicitacao in solicitacoes:
                preenchimento = preenchimentos_map.get(solicitacao.id)
                pid           = preenchimento.id if preenchimento else None

                auditoria_data.append({
                    'solicitacao': solicitacao,
                    'material':    materiais_map.get(solicitacao.cod_material),
                    'preenchimento': preenchimento,
                    'pedido':      pedidos_map.get(pid) if pid else None,
                    'estoque':     estoques_map.get(pid) if pid else None,
                    'requisicoes': requisicoes_map.get(pid, []) if pid else [],
                    'fornecedor':  fornecedores.get(preenchimento.fornecedor_id, {
                        'nome_fantasia': 'Fornecedor não encontrado',
                        'cnpj': 'N/A'
                    }) if preenchimento and preenchimento.fornecedor_id else {},
                    'auditorias':  auditorias_map.get(solicitacao.id, []),
                })

            # Listas para filtros — queries leves, só valores distintos
            empresas_lista = [
                e[0] for e in db.session.query(SolicitacoesCompra.empresa)
                .distinct().order_by(SolicitacoesCompra.empresa).all() if e[0]
            ]
            usuarios_lista = [
                u[0] for u in db.session.query(SolicitacoesCompra.usuario)
                .distinct().order_by(SolicitacoesCompra.usuario).all() if u[0]
            ]
            nomes_ativos_lista = [
                n[0] for n in db.session.query(SolicitacoesCompra.nome_ativo)
                .filter(SolicitacoesCompra.ativo == 'Sim',
                        SolicitacoesCompra.nome_ativo.isnot(None))
                .distinct().order_by(SolicitacoesCompra.nome_ativo).all() if n[0]
            ]
            usuarios_disponiveis = sorted(ler_senhas().keys())

        return render_template(
            'tela_auditoria.html',
            auditoria=auditoria_data,
            empresas=empresas_lista,
            usuarios=usuarios_lista,
            nomes_ativos=nomes_ativos_lista,
            usuarios_disponiveis=usuarios_disponiveis,
            paginacao=paginacao,
            page=page,
            filtros={
                'empresa':      empresa,
                'usuario':      usuario,
                'ativo':        ativo,
                'nome_ativo':   nome_ativo,
                'data_inicio':  data_inicio,
                'data_fim':     data_fim,
                'status':       status,
            }
        )

    except Exception as e:
        db.session.rollback()
        _flash_erro_seguro('Erro ao carregar auditoria.', e, app.logger)
        app.logger.error(f'Erro em tela_auditoria: {str(e)}', exc_info=True)
        return render_template(
            'tela_auditoria.html',
            auditoria=[],
            empresas=[],
            usuarios=[],
            nomes_ativos=[],
            usuarios_disponiveis=[],
            paginacao=None,
            page=1,
            filtros={}
        )
    
def create_auditoria_table():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        create_table_query = """
        CREATE TABLE IF NOT EXISTS Auditoria (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            solicitacao_id INTEGER NOT NULL,
            data_validacao DATETIME NOT NULL,
            colaborador_1 TEXT NOT NULL,
            colaborador_2 TEXT,
            status TEXT NOT NULL,
            observacao TEXT,
            FOREIGN KEY (solicitacao_id) REFERENCES SolicitacoesCompra(id)
        )
        """
        cursor.execute(create_table_query)
        conn.commit()
        print("Tabela 'Auditoria' criada com sucesso.")
        cursor.close()
        conn.close()
        return True
    except sqlite3.Error as e:
        print(f"Erro ao criar a tabela Auditoria: {str(e)}")
        return False

@routes_bp.route('/excluir_auditoria/<int:id>', methods=['POST'])
def excluir_auditoria(id):
    if 'usuario' not in session:
        flash('Você precisa estar logado para realizar esta ação.', 'error')
        return redirect(url_for('routes_bp.login'))

    # Apenas master pode excluir auditorias
    usuario_completo = session.get('usuario_completo', session['usuario'])
    senhas = ler_senhas()
    role = senhas.get(usuario_completo, {}).get('pagina', '')
    if 'menu_master' not in role:
        flash('Sem permissão para excluir registros de auditoria.', 'error')
        return redirect(url_for('routes_bp.tela_auditoria'))
    
    try:
        auditoria = Auditoria.query.get_or_404(id)
        solicitacao_id = auditoria.solicitacao_id
        
        db.session.delete(auditoria)
        db.session.commit()
        
        flash('Registro de auditoria excluído com sucesso.', 'success')
        return redirect(url_for('routes_bp.tela_auditoria'))
    
    except Exception as e:
        db.session.rollback()
        _flash_erro_seguro('Erro ao excluir auditoria.', e, app.logger)
        return redirect(url_for('routes_bp.tela_auditoria'))


#Dasboard ----------------------------------------------
@routes_bp.route('/dashboard', methods=['GET'])
@limiter.limit("30 per minute")
def dashboard():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        # Importações necessárias
        from sqlalchemy.orm import joinedload
        from datetime import datetime, timedelta
        
        # Obter parâmetros de filtro
        empresa = request.args.get('empresa', 'Todas')
        periodo = request.args.get('periodo', 'Últimos 30 dias')
        status_filtro = request.args.get('status', 'Todos')
        comprador_filtro = request.args.get('comprador', 'Todos')
        
        # Calcular datas com base no período
        data_atual = get_local_time()
        data_inicio = None
        
        if periodo == 'Últimos 30 dias':
            data_inicio = data_atual - timedelta(days=30)
        elif periodo == 'Últimos 6 meses':
            data_inicio = data_atual - timedelta(days=180)
        elif periodo == '2025':
            data_inicio = datetime(2025, 1, 1)
        elif periodo == '2024':
            data_inicio = datetime(2024, 1, 1)
        
        # Query com eager loading
        query = SolicitacoesCompra.query.options(
            joinedload(SolicitacoesCompra.material),
            joinedload(SolicitacoesCompra.preenchimentos_fornecidos)
        )
        
        # Aplicar filtro de data
        if data_inicio:
            query = query.filter(SolicitacoesCompra.data_solicitacao >= data_inicio)
        
        # Executar consulta
        solicitacoes = query.all()
        
        # ============================================
        # FUNÇÃO AUXILIAR PARA CONVERTER QUANTIDADE
        # ============================================
        def converter_quantidade_para_float(quantidade):
            """Converte quantidade para float, seja string ou número"""
            if quantidade is None:
                return 0.0
            if isinstance(quantidade, (int, float)):
                return float(quantidade)
            try:
                # Substitui vírgula por ponto se necessário
                qtd_str = str(quantidade).strip().replace(',', '.')
                # Remove caracteres não numéricos exceto ponto
                import re
                qtd_str = re.sub(r'[^\d.]', '', qtd_str)
                # Remove múltiplos pontos (mantém apenas o primeiro)
                partes = qtd_str.split('.')
                if len(partes) > 2:
                    qtd_str = partes[0] + '.' + ''.join(partes[1:])
                return float(qtd_str) if qtd_str else 0.0
            except (ValueError, TypeError):
                return 0.0
        
        # ============================================
        # ESTRUTURAS DE DADOS
        # ============================================
        total_solicitacoes = len(solicitacoes)
        
        # Dicionários para agregação
        compradores_metrics = {}
        produtos_pendentes = {}
        volume_por_empresa = {}
        pendencias_por_empresa = {}
        gastos_por_veiculo = {}
        compras_por_veiculo = {}
        tempo_aberto_lista = []
        gastos_por_empresa = {}
        produtividade_por_empresa = {}
        
        # ============================================
        # PROCESSAMENTO DAS SOLICITAÇÕES
        # ============================================
        for solicitacao in solicitacoes:
            # Aplicar filtros em memória
            if empresa != 'Todas' and solicitacao.empresa != empresa:
                continue
            if comprador_filtro != 'Todos' and solicitacao.comprador_atribuido != comprador_filtro:
                continue
            
            # Status da solicitação
            status = solicitacao.status_aprovacao
            if status_filtro != 'Todos':
                if status_filtro == 'Pendente' and status not in [None, '', 'Pendente']:
                    continue
                elif status_filtro == 'Aprovado' and status != 'Aprovado':
                    continue
                elif status_filtro == 'Reprovado' and status != 'Reprovado':
                    continue
            
            # Converter quantidade para float
            quantidade_float = converter_quantidade_para_float(solicitacao.quantidade)
            
            # 1. PRODUTIVIDADE DOS COMPRADORES
            comprador = solicitacao.comprador_atribuido or 'Não Atribuído'
            if comprador not in compradores_metrics:
                compradores_metrics[comprador] = {
                    'pendentes': 0, 'aprovadas': 0, 'rejeitadas': 0,
                    'valor_total': 0, 'tempo_medio': 0, 'contador_tempo': 0,
                    'total': 0
                }
            
            # Incrementar contadores
            compradores_metrics[comprador]['total'] += 1
            
            if status == 'Aprovado':
                compradores_metrics[comprador]['aprovadas'] += 1
            elif status in [None, '', 'Pendente']:
                compradores_metrics[comprador]['pendentes'] += 1
            elif status == 'Reprovado':
                compradores_metrics[comprador]['rejeitadas'] += 1
            
            # Calcular valor total da solicitação
            valor_solicitacao = 0
            for preenchimento in solicitacao.preenchimentos_fornecidos:
                if preenchimento.status != 'Rascunho' and preenchimento.valor_total:
                    valor_solicitacao += preenchimento.valor_total
            
            compradores_metrics[comprador]['valor_total'] += valor_solicitacao
            
            # Tempo médio para aprovadas
            if status == 'Aprovado' and solicitacao.data_solicitacao:
                dias = (data_atual - solicitacao.data_solicitacao).days
                compradores_metrics[comprador]['tempo_medio'] += dias
                compradores_metrics[comprador]['contador_tempo'] += 1
            
            # 2. PRODUTOS PENDENTES
            if status in [None, '', 'Pendente'] and solicitacao.material:
                material = solicitacao.material
                produto_nome = material.DescricaoMaterial
                if len(produto_nome) > 40:
                    produto_nome = produto_nome[:40] + '...'
                
                chave = f"{material.CodMaterial}|{solicitacao.empresa}"
                
                if chave not in produtos_pendentes:
                    produtos_pendentes[chave] = {
                        'produto': produto_nome,
                        'quantidade': 0,
                        'valor_total': 0,
                        'empresa': solicitacao.empresa,
                        'dias_pendente': (data_atual - solicitacao.data_solicitacao).days if solicitacao.data_solicitacao else 0
                    }
                
                produtos_pendentes[chave]['quantidade'] += quantidade_float
                
                # Valor estimado
                if solicitacao.preenchimentos_fornecidos:
                    valores = [p.valor_unitario for p in solicitacao.preenchimentos_fornecidos 
                              if p.valor_unitario and p.status != 'Rascunho']
                    if valores:
                        valor_medio = sum(valores) / len(valores)
                        produtos_pendentes[chave]['valor_total'] += valor_medio * quantidade_float
            
            # 3. VOLUME POR EMPRESA
            empresa_nome = solicitacao.empresa
            if empresa_nome not in volume_por_empresa:
                volume_por_empresa[empresa_nome] = {'total': 0, 'valor_total': 0}
            
            volume_por_empresa[empresa_nome]['total'] += 1
            volume_por_empresa[empresa_nome]['valor_total'] += valor_solicitacao
            
            # 4. PENDÊNCIAS POR EMPRESA
            if status in [None, '', 'Pendente']:
                if empresa_nome not in pendencias_por_empresa:
                    pendencias_por_empresa[empresa_nome] = {'quantidade': 0, 'dias_total': 0, 'valor_pendente': 0}
                
                pendencias_por_empresa[empresa_nome]['quantidade'] += 1
                dias = (data_atual - solicitacao.data_solicitacao).days if solicitacao.data_solicitacao else 0
                pendencias_por_empresa[empresa_nome]['dias_total'] += dias
                pendencias_por_empresa[empresa_nome]['valor_pendente'] += valor_solicitacao
            
            # 5. GASTOS POR VEÍCULO
            if solicitacao.ativo == 'Sim' and solicitacao.nome_ativo:
                veiculo = solicitacao.nome_ativo
                if veiculo not in gastos_por_veiculo:
                    gastos_por_veiculo[veiculo] = {'valor_total': 0, 'solicitacoes': 0, 'itens_total': 0}
                
                gastos_por_veiculo[veiculo]['valor_total'] += valor_solicitacao
                gastos_por_veiculo[veiculo]['solicitacoes'] += 1
                gastos_por_veiculo[veiculo]['itens_total'] += quantidade_float
                
                # 6. COMPRAS POR VEÍCULO
                if veiculo not in compras_por_veiculo:
                    compras_por_veiculo[veiculo] = 0
                compras_por_veiculo[veiculo] += quantidade_float
            
            # 7. TEMPO EM ABERTO
            if status in [None, '', 'Pendente'] and solicitacao.data_solicitacao and solicitacao.material:
                dias_aberto = (data_atual - solicitacao.data_solicitacao).days
                if dias_aberto > 0:
                    material_nome = solicitacao.material.DescricaoMaterial
                    if len(material_nome) > 30:
                        material_nome = material_nome[:30] + '...'
                    
                    prioridade = dias_aberto * (1 + (valor_solicitacao / 10000))
                    
                    tempo_aberto_lista.append({
                        'id': solicitacao.id,
                        'dias_aberto': dias_aberto,
                        'empresa': solicitacao.empresa,
                        'material': material_nome,
                        'valor_pendente': valor_solicitacao,
                        'prioridade': round(prioridade, 2),
                        'comprador': solicitacao.comprador_atribuido or 'Não Atribuído'
                    })
            
            # 8. GASTOS POR EMPRESA
            if empresa_nome not in gastos_por_empresa:
                gastos_por_empresa[empresa_nome] = 0
            gastos_por_empresa[empresa_nome] += valor_solicitacao
            
            # 9. PRODUTIVIDADE POR EMPRESA
            if empresa_nome not in produtividade_por_empresa:
                produtividade_por_empresa[empresa_nome] = {
                    'compradores': set(), 'solicitacoes': 0, 'aprovadas': 0, 'valor_total': 0
                }
            
            produtividade_por_empresa[empresa_nome]['solicitacoes'] += 1
            if solicitacao.comprador_atribuido:
                produtividade_por_empresa[empresa_nome]['compradores'].add(solicitacao.comprador_atribuido)
            if status == 'Aprovado':
                produtividade_por_empresa[empresa_nome]['aprovadas'] += 1
            produtividade_por_empresa[empresa_nome]['valor_total'] += valor_solicitacao
        
        # ============================================
        # FORMATAÇÃO DOS DADOS PARA O TEMPLATE
        # ============================================
        
        # 1. Compradores
        compradores_lista = []
        for nome, dados in compradores_metrics.items():
            tempo_medio = dados['tempo_medio'] / dados['contador_tempo'] if dados['contador_tempo'] > 0 else 0
            total = dados['total']
            taxa_aprovacao = (dados['aprovadas'] / total * 100) if total > 0 else 0
            
            compradores_lista.append({
                'nome': nome,
                'pendentes': dados['pendentes'],
                'aprovadas': dados['aprovadas'],
                'rejeitadas': dados['rejeitadas'],
                'valor_total': dados['valor_total'],
                'tempo_medio_dias': round(tempo_medio, 1),
                'taxa_aprovacao': round(taxa_aprovacao, 1),
                'total': total
            })
        compradores_lista.sort(key=lambda x: x['total'], reverse=True)
        
        # 2. Produtos pendentes
        produtos_pendentes_lista = sorted(
            produtos_pendentes.values(),
            key=lambda x: x['dias_pendente'],
            reverse=True
        )[:8]
        
        # 3. Volume por empresa
        volume_total = sum(dados['total'] for dados in volume_por_empresa.values())
        volume_empresa_data = []
        for empresa_nome, dados in volume_por_empresa.items():
            porcentagem = round((dados['total'] / volume_total) * 100, 1) if volume_total > 0 else 0
            volume_empresa_data.append({
                'empresa': empresa_nome,
                'volume': dados['total'],
                'valor_total': dados['valor_total'],
                'crescimento': 0,
                'porcentagem': porcentagem
            })
        volume_empresa_data.sort(key=lambda x: x['volume'], reverse=True)
        
        # 4. Pendências por empresa
        pendencias_lista = []
        for empresa_nome, dados in pendencias_por_empresa.items():
            tempo_medio = dados['dias_total'] / dados['quantidade'] if dados['quantidade'] > 0 else 0
            pendencias_lista.append({
                'empresa': empresa_nome,
                'pendencias': dados['quantidade'],
                'tempo_medio_dias': round(tempo_medio, 1),
                'valor_pendente': dados['valor_pendente']
            })
        pendencias_lista.sort(key=lambda x: x['pendencias'], reverse=True)
        
        # 5. Gastos por veículo
        gastos_veiculo_lista = [
            {
                'veiculo': veiculo,
                'valor_total': dados['valor_total'],
                'solicitacoes': dados['solicitacoes'],
                'itens_total': dados['itens_total'],
                'valor_medio_por_item': dados['valor_total'] / dados['itens_total'] if dados['itens_total'] > 0 else 0
            }
            for veiculo, dados in gastos_por_veiculo.items()
        ]
        gastos_veiculo_lista.sort(key=lambda x: x['valor_total'], reverse=True)
        
        # 6. Compras por veículo
        compras_veiculo_lista = [
            {
                'veiculo': veiculo,
                'quantidade_total': quantidade,
                'solicitacoes': sum(1 for s in solicitacoes if s.nome_ativo == veiculo)
            }
            for veiculo, quantidade in compras_por_veiculo.items()
        ]
        compras_veiculo_lista.sort(key=lambda x: x['quantidade_total'], reverse=True)
        
        # 7. Tempo em aberto
        tempo_aberto_lista.sort(key=lambda x: x['prioridade'], reverse=True)
        tempo_aberto_lista = tempo_aberto_lista[:10]
        
        # 8. Gastos por empresa
        gastos_empresa_lista = []
        for empresa_nome, gastos in gastos_por_empresa.items():
            tendencia = 'alta' if gastos > 100000 else ('estável' if gastos > 50000 else 'baixa')
            simbolo = '↗' if gastos > 100000 else ('→' if gastos > 50000 else '↘')
            
            gastos_empresa_lista.append({
                'empresa': empresa_nome,
                'gastos_total': gastos,
                'variacao': 0,
                'tendencia': tendencia,
                'simbolo': simbolo
            })
        gastos_empresa_lista.sort(key=lambda x: x['gastos_total'], reverse=True)
        
        # 9. Produtividade por empresa
        produtividade_lista = []
        for empresa_nome, dados in produtividade_por_empresa.items():
            taxa_aprovacao = (dados['aprovadas'] / dados['solicitacoes'] * 100) if dados['solicitacoes'] > 0 else 0
            produtividade_lista.append({
                'empresa': empresa_nome,
                'compradores': len(dados['compradores']),
                'solicitacoes': dados['solicitacoes'],
                'aprovadas': dados['aprovadas'],
                'taxa_aprovacao': round(taxa_aprovacao, 1),
                'valor_total': dados['valor_total']
            })
        produtividade_lista.sort(key=lambda x: x['taxa_aprovacao'], reverse=True)
        
        # ============================================
        # DADOS BÁSICOS
        # ============================================
        aprovadas = sum(1 for s in solicitacoes if s.status_aprovacao == 'Aprovado')
        pendentes = sum(1 for s in solicitacoes if s.status_aprovacao in [None, '', 'Pendente'])
        rejeitadas = sum(1 for s in solicitacoes if s.status_aprovacao == 'Reprovado')
        
        # Valor total
        valor_total = sum(d['valor_total'] for d in volume_por_empresa.values())
        
        # Tempo médio de aprovação
        tempos_aprovacao = [s for s in solicitacoes if s.status_aprovacao == 'Aprovado' and s.data_solicitacao]
        tempo_medio_aprovacao = sum((data_atual - s.data_solicitacao).days for s in tempos_aprovacao) / len(tempos_aprovacao) if tempos_aprovacao else 0
        
        # Valor médio por solicitação
        valor_medio_por_solicitacao = valor_total / total_solicitacoes if total_solicitacoes > 0 else 0
        
        # ============================================
        # FUNÇÃO AUXILIAR DE FORMATAÇÃO
        # ============================================
        def formatar_valor(valor):
            try:
                valor = float(valor)
                if valor >= 1000000:
                    return f"R$ {valor/1000000:.1f}M"
                elif valor >= 1000:
                    return f"R$ {valor/1000:.1f}K"
                else:
                    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except:
                return "R$ 0,00"
        
        # ============================================
        # DADOS PARA EVOLUÇÃO (6 meses)
        # ============================================
        meses_nomes = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        mes_atual = data_atual.month
        meses_labels = []
        for i in range(6):
            mes_num = mes_atual - (5 - i)
            ano = data_atual.year
            if mes_num <= 0:
                mes_num += 12
                ano -= 1
            meses_labels.append(f"{meses_nomes[mes_num - 1]}/{str(ano)[-2:]}")
        
        # Evolução de compras (dados agregados por mês)
        evolucao_compras = {}
        empresas_existentes = list(set(s.empresa for s in solicitacoes if s.empresa))[:3]
        
        for empresa_nome in empresas_existentes:
            evolucao_compras[empresa_nome] = []
            for mes in meses_labels:
                # Aqui você pode implementar a lógica real de agregação mensal
                # Por enquanto, dados de exemplo
                evolucao_compras[empresa_nome].append(random.randint(50, 200))
        
        # ============================================
        # FILTROS DISPONÍVEIS
        # ============================================
        empresas_filtro = ['Todas'] + list(set(s.empresa for s in solicitacoes if s.empresa))
        compradores_filtro = ['Todos'] + list(set(s.comprador_atribuido for s in solicitacoes if s.comprador_atribuido))
        
        # Ordenar listas
        empresas_filtro.sort() if len(empresas_filtro) > 1 else None
        compradores_filtro.sort() if len(compradores_filtro) > 1 else None
        
        # ============================================
        # RENDERIZAR TEMPLATE
        # ============================================
        return render_template(
            'dashboard.html',
            # Dados básicos
            total_solicitacoes=total_solicitacoes,
            aprovadas=aprovadas,
            pendentes=pendentes,
            rejeitadas=rejeitadas,
            valor_total=formatar_valor(valor_total),
            tempo_medio_aprovacao=round(tempo_medio_aprovacao, 1),
            valor_medio_por_solicitacao=formatar_valor(valor_medio_por_solicitacao),
            
            # Dados para gráficos e tabelas
            compradores_produtividade=compradores_lista[:8],
            produtos_pendentes=produtos_pendentes_lista,
            volume_empresa=volume_empresa_data[:6],
            pendencias_empresa=pendencias_lista[:6],
            gastos_veiculo=gastos_veiculo_lista[:5],
            compras_veiculo=compras_veiculo_lista[:5],
            tempo_aberto=tempo_aberto_lista,
            gastos_empresa=gastos_empresa_lista[:6],
            evolucao_compras=evolucao_compras,
            evolucao_meses_labels=meses_labels,
            produtividade_geral=produtividade_lista[:6],
            
            # Filtros
            empresas_filtro=empresas_filtro,
            compradores_filtro=compradores_filtro,
            empresa_selecionada=empresa,
            periodo_selecionado=periodo,
            status_selecionado=status_filtro,
            comprador_selecionado=comprador_filtro,
            
            # Função auxiliar
            formatar_valor=formatar_valor
        )
        
    except Exception as e:
        _flash_erro_seguro('Erro ao carregar dashboard.', e, app.logger)
        app.logger.error(f'Erro no dashboard: {str(e)}', exc_info=True)
        
        # Função auxiliar de fallback
        def formatar_valor_fallback(valor):
            return "R$ 0,00"
        
        return render_template(
            'dashboard.html',
            total_solicitacoes=0,
            aprovadas=0,
            pendentes=0,
            rejeitadas=0,
            valor_total="R$ 0,00",
            tempo_medio_aprovacao=0,
            valor_medio_por_solicitacao="R$ 0,00",
            compradores_produtividade=[],
            produtos_pendentes=[],
            volume_empresa=[],
            pendencias_empresa=[],
            gastos_veiculo=[],
            compras_veiculo=[],
            tempo_aberto=[],
            gastos_empresa=[],
            evolucao_compras={},
            evolucao_meses_labels=[],
            produtividade_geral=[],
            empresas_filtro=['Todas'],
            compradores_filtro=['Todos'],
            empresa_selecionada='Todas',
            periodo_selecionado='Últimos 30 dias',
            status_selecionado='Todos',
            comprador_selecionado='Todos',
            formatar_valor=formatar_valor_fallback
        )
#------------------------------------------------------------------------------------
    
@routes_bp.route('/dashboard/data', methods=['GET'])
def dashboard_data():
    if 'usuario' not in session:
        return jsonify({'error': 'Não autenticado'}), 401
    
    try:
        # Dados resumidos
        total_materiais = Materiais.query.count()
        estoque_critico = Materiais.query.filter(
            Materiais.FatorConsumo > 0,
            (Materiais.QuantidadeEstoque / Materiais.FatorConsumo) < 5
        ).count()
        total_estoque = db.session.query(
            db.func.sum(Materiais.QuantidadeEstoque)
        ).scalar() or 0
        requisicoes_abertas = Requisicoes.query.count()
        
        # Status das solicitações
        status_solicitacoes = db.session.query(
            SolicitacoesPreenchidas.status,
            db.func.count(SolicitacoesPreenchidas.id)
        ).group_by(SolicitacoesPreenchidas.status).all()
        
        # Últimas solicitações
        ultimas_solicitacoes = SolicitacoesCompra.query.order_by(
            SolicitacoesCompra.data_solicitacao.desc()
        ).limit(5).all()
        
        # Materiais com maior estoque (top 5)
        materiais_estoque = Materiais.query.order_by(
            Materiais.QuantidadeEstoque.desc()
        ).limit(5).all()
        
        # Formatando os dados para o dashboard
        dados = {
            'totais': {
                'materiais': total_materiais,
                'estoque_critico': estoque_critico,
                'total_estoque': total_estoque,
                'requisicoes_abertas': requisicoes_abertas
            },
            'status_solicitacoes': [
                {'status': s[0], 'quantidade': s[1]} for s in status_solicitacoes
            ],
            'ultimas_solicitacoes': [
                {
                    'id': s.id,
                    'material': s.material.DescricaoMaterial if s.material else 'N/A',
                    'quantidade': s.quantidade,
                    'data': s.data_solicitacao.strftime('%d/%m/%Y')
                } for s in ultimas_solicitacoes
            ],
            'top_materiais': [
                {
                    'nome': m.DescricaoMaterial,
                    'quantidade': m.QuantidadeEstoque,
                    'dias_estoque': m.QuantidadeEstoque / m.FatorConsumo if m.FatorConsumo > 0 else 0
                } for m in materiais_estoque
            ]
        }
        
        return jsonify(dados)
    
    except Exception as e:
        app.logger.error(f'Erro no dashboard: {str(e)}')
        app.logger.error(f'API error: {type(e).__name__}', exc_info=False)
        return jsonify({'error': 'Erro interno. Tente novamente.'}), 500


#-------------------------------------------------
# Updated financeiro route
@routes_bp.route('/financeiro', methods=['GET'])
def financeiro():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        empresa = request.args.get('empresa')
        usuario = request.args.get('usuario')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')

        query = db.session.query(PedidosCompra).filter(
            PedidosCompra.status == 'AgPagamento'
        ).join(
            pedido_preenchimento_associacao,
            PedidosCompra.id == pedido_preenchimento_associacao.c.pedido_id
        ).join(
            SolicitacoesPreenchidas,
            pedido_preenchimento_associacao.c.preenchimento_id == SolicitacoesPreenchidas.id
        ).join(
            SolicitacoesCompra,
            SolicitacoesPreenchidas.solicitacao_id == SolicitacoesCompra.id
        )

        if empresa:
            query = query.filter(SolicitacoesCompra.empresa == empresa)
        if usuario:
            query = query.filter(PedidosCompra.usuario == usuario)
        if data_inicio:
            query = query.filter(PedidosCompra.data_criacao >= data_inicio)
        if data_fim:
            data_fim_ajustada = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(PedidosCompra.data_criacao <= data_fim_ajustada)

        pedidos = query.order_by(PedidosCompra.data_criacao.desc()).distinct().all()

        fornecedor_ids = set()
        for pedido in pedidos:
            for preenchimento in pedido.preenchimentos:
                fornecedor_ids.add(preenchimento.fornecedor_id)
        
        fornecedores = {}
        if fornecedor_ids:
            conn = get_db_connection(DB_PATH_FORNECEDORES)
            if conn:
                try:
                    cursor = conn.cursor()
                    cursor.execute(
                        f'SELECT id, nome_fantasia, cnpj FROM fornecedores WHERE id IN ({",".join("?"*len(fornecedor_ids))})',
                        list(fornecedor_ids)
                    )
                    for row in cursor.fetchall():
                        fornecedores[row[0]] = {
                            'nome_fantasia': row[1],
                            'cnpj': format_cnpj(row[2]) if row[2] else 'N/A'
                        }
                finally:
                    conn.close()

        pedidos_completos = []
        for pedido in pedidos:
            preenchimentos_info = []
            for preenchimento in pedido.preenchimentos:
                fornecedor_info = fornecedores.get(preenchimento.fornecedor_id, {
                    'nome_fantasia': 'Fornecedor não encontrado',
                    'cnpj': 'N/A'
                })
                
                # CORREÇÃO: A marca está na solicitação, não no preenchimento
                marca = preenchimento.solicitacao.marca if preenchimento.solicitacao.marca else 'Não informado'
                
                preenchimentos_info.append({
                    'id': preenchimento.id,
                    'marca': marca,  # Usando a marca da solicitação
                    'fornecedor_nome': fornecedor_info['nome_fantasia'],
                    'fornecedor_cnpj': fornecedor_info['cnpj'],
                    'material': preenchimento.solicitacao.material.DescricaoMaterial if preenchimento.solicitacao.material else 'N/A',
                    'empresa': preenchimento.solicitacao.empresa,
                    'valor_total': float(preenchimento.valor_total) if preenchimento.valor_total else 0.0
                })
            pedidos_completos.append({
                'pedido': pedido,
                'preenchimentos': preenchimentos_info
            })

        empresas = db.session.query(SolicitacoesCompra.empresa).distinct().order_by(SolicitacoesCompra.empresa).all()
        usuarios = db.session.query(PedidosCompra.usuario).distinct().order_by(PedidosCompra.usuario).all()

        return render_template(
            'financeiro.html',
            pedidos=pedidos_completos,
            empresas=[e[0] for e in empresas if e[0]],
            usuarios=[u[0] for u in usuarios if u[0]],
            filtros={
                'empresa': empresa,
                'usuario': usuario,
                'data_inicio': data_inicio,
                'data_fim': data_fim
            }
        )
    except Exception as e:
        _flash_erro_seguro('Erro ao carregar página financeiro.', e, app.logger)
        return render_template(
            'financeiro.html',
            pedidos=[],
            empresas=[],
            usuarios=[],
            filtros={}
        )

# Updated confirmar_pagamento route with explicit file size validation
@routes_bp.route('/api/pedidos/<int:pedido_id>/confirmar_pagamento', methods=['POST'])
def confirmar_pagamento(pedido_id):
    if 'usuario' not in session:
        return jsonify({'success': False, 'error': 'Usuário não autenticado'}), 401

    pedido = PedidosCompra.query.get_or_404(pedido_id)
    if pedido.status != 'AgPagamento':
        return jsonify({'success': False, 'error': 'Este pedido não está aguardando pagamento'}), 400

    if 'comprovante' not in request.files:
        return jsonify({'success': False, 'error': 'Nenhum arquivo enviado'}), 400

    comprovante = request.files['comprovante']
    if comprovante.filename == '':
        return jsonify({'success': False, 'error': 'Nenhum arquivo selecionado'}), 400

    if comprovante and allowed_file(comprovante.filename):
        try:
            # Validar tamanho ANTES de salvar no disco
            comprovante.seek(0, 2)  # seek to end
            file_size = comprovante.tell()
            comprovante.seek(0)     # reset
            if file_size > 5 * 1024 * 1024:
                return jsonify({'success': False, 'error': 'O arquivo PDF deve ter no máximo 5MB'}), 400

            filename = secure_filename(comprovante.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            unique_filename = f"{timestamp}_{filename}"
            comprovante_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            comprovante.save(comprovante_path)

            pedido.comprovante_pagamento = unique_filename  # armazena path relativo
            pedido.status = 'AgEntrega'
            db.session.commit()
            return jsonify({'success': True, 'message': 'Pagamento confirmado com sucesso'})
        except Exception as e:
            db.session.rollback()
            logging.error(f"Erro ao confirmar pagamento: {e}")
            return jsonify({'success': False, 'error': 'Erro interno. Tente novamente.'}), 500
    else:
        return jsonify({'success': False, 'error': 'Arquivo inválido. Apenas PDFs são permitidos'}), 400
    
#Excluir fonecedor  
@routes_bp.route('/delete_supplier', methods=['POST'])
def delete_supplier():
    if 'usuario' not in session:
        return jsonify({'status': 'error', 'message': 'Usuário não autenticado'}), 401
    
    try:
        data = request.get_json()
        supplier_id = data.get('id')
        
        if not supplier_id:
            return jsonify({'status': 'error', 'message': 'ID do fornecedor não fornecido'}), 400
        
        # Verificar se há solicitações associadas a este fornecedor no banco principal
        try:
            count = db.session.query(SolicitacoesPreenchidas).filter_by(fornecedor_id=supplier_id).count()
            
            if count > 0:
                return jsonify({
                    'status': 'error', 
                    'message': 'Não é possível excluir: fornecedor está associado a solicitações'
                }), 400
            
            # Se não houver associações, excluir o fornecedor do banco de fornecedores
            conn = get_db_connection(DB_PATH_FORNECEDORES)
            if not conn:
                return jsonify({'status': 'error', 'message': 'Erro ao conectar ao banco de dados'}), 500
                
            try:
                cursor = conn.cursor()
                cursor.execute('DELETE FROM fornecedores WHERE id = ?', (supplier_id,))
                conn.commit()
                return jsonify({'status': 'success', 'message': 'Fornecedor excluído com sucesso'})
            except sqlite3.Error as e:
                conn.rollback()
                return jsonify({'status': 'error', 'message': 'Erro interno. Tente novamente.'}), 500
            finally:
                if conn:
                    conn.close()
                    
        except SQLAlchemyError as e:
            return jsonify({'status': 'error', 'message': 'Erro interno. Tente novamente.'}), 500
                
    except Exception as e:
        return jsonify({'status': 'error', 'message': 'Erro interno. Tente novamente.'}), 500
    
#Inserir estoque manual caso necessario   
@routes_bp.route('/definir-quantidade/<int:cod>', methods=['POST'])
def definir_quantidade(cod):
    if 'usuario' not in session:
        flash('Acesso não autorizado', 'error')
        return redirect(url_for('routes_bp.login'))
    
    try:
        quantidade = request.form.get('quantidade', 0)
        motivo = request.form.get('motivo', 'Ajuste manual')  # Novo campo
        
        # Validações
        try:
            quantidade = int(quantidade)
        except ValueError:
            flash('Valor inválido para quantidade. Use apenas números inteiros.', 'error')
            return redirect(url_for('routes_bp.listar_materiais'))
        
        if quantidade < 0:
            flash('Quantidade não pode ser negativa', 'error')
            return redirect(url_for('routes_bp.listar_materiais'))
        
        material = Materiais.query.get_or_404(cod)
        
        # Registrar no histórico antes de alterar
        historico = HistoricoEstoque(
            cod_material=material.CodMaterial,
            usuario=session['usuario'],
            quantidade_anterior=material.QuantidadeEstoque,
            quantidade_nova=quantidade,
            motivo=motivo
        )
        db.session.add(historico)
        
        # Atualizar estoque
        material.QuantidadeEstoque = quantidade
        db.session.commit()
        
        flash(f'Estoque de {material.DescricaoMaterial} atualizado de {historico.quantidade_anterior} para {quantidade} unidades', 'success')
    
    except Exception as e:
        db.session.rollback()
        _flash_erro_seguro('Erro ao atualizar quantidade.', e, app.logger)
    
    return redirect(url_for('routes_bp.listar_materiais'))

@routes_bp.route('/relatorio-estoque')
def relatorio_estoque():
    if 'usuario' not in session:
        flash('Acesso não autorizado', 'error')
        return redirect(url_for('routes_bp.login'))

    # Obter parâmetros de filtro
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Itens por página
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    usuario = request.args.get('usuario')
    material = request.args.get('material')

    # Construir query base
    query = HistoricoEstoque.query.join(Materiais)

    # Aplicar filtros
    if data_inicio:
        query = query.filter(HistoricoEstoque.data_alteracao >= data_inicio)
    if data_fim:
        data_fim_ajustada = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(HistoricoEstoque.data_alteracao <= data_fim_ajustada)
    if usuario:
        query = query.filter(HistoricoEstoque.usuario.ilike(f'%{usuario}%'))
    if material:
        query = query.filter(or_(
            Materiais.DescricaoMaterial.ilike(f'%{material}%'),
            Materiais.CodMaterial.ilike(f'%{material}%')
        ))

    # Paginação
    historico_paginado = query.order_by(HistoricoEstoque.data_alteracao.desc()).paginate(
        page=page, 
        per_page=per_page, 
        error_out=False
    )

    return render_template('relatorio_estoque.html', historico=historico_paginado)

@routes_bp.route('/api/solicitacoes/<int:solicitacao_id>/status', methods=['POST'])
def atualizar_status_solicitacao(solicitacao_id):
    if 'usuario' not in session:
        return jsonify({'success': False, 'error': 'Usuário não autenticado'}), 401

    # Verificar role: apenas aprovação e master podem alterar status
    usuario_completo = session.get('usuario_completo', session['usuario'])
    senhas = ler_senhas()
    role = senhas.get(usuario_completo, {}).get('pagina', '')
    if not any(r in role for r in ('menu_aprovacao', 'menu_master', 'menu_supervisor')):
        logging.warning(
            'AUTHZ_FAILURE | usuario=%s | role=%s | ip=%s | endpoint=atualizar_status_solicitacao | recurso=solicitacao_%s',
            usuario_completo, role, request.remote_addr, solicitacao_id
        )
        registrar_log(usuario_completo, 'acesso_negado', f'Tentativa de alterar status solicitação ID {solicitacao_id} sem permissão', request.remote_addr)
        return jsonify({'success': False, 'error': 'Sem permissão para aprovar/reprovar'}), 403

    try:
        data = request.get_json()
        status = data.get('status')

        if not status:
            return jsonify({'success': False, 'error': 'Nenhum status fornecido'}), 400

        if status not in ['Aprovado', 'Reprovado']:
            return jsonify({'success': False, 'error': 'Status inválido'}), 400

        # Transição atômica com validação de estado anterior (Dirty Write prevention)
        # Regra: só é possível transitar de um estado "pendente" para Aprovado/Reprovado.
        _TRANSICOES_PERMITIDAS = {
            'Aprovado':  {None, '', 'Pendente'},
            'Reprovado': {None, '', 'Pendente', 'Aprovado'},
        }
        solicitacao = SolicitacoesCompra.query.with_for_update().get_or_404(solicitacao_id)
        estado_atual = solicitacao.status_aprovacao or ''
        if estado_atual not in _TRANSICOES_PERMITIDAS[status]:
            return jsonify({
                'success': False,
                'error': f'Transição inválida: {estado_atual or "Pendente"} → {status}'
            }), 409

        solicitacao.status_aprovacao = status
        registrar_log(
            usuario_completo,
            f'status_{status.lower()}',
            f'Solicitação ID {solicitacao_id}: {estado_atual or "Pendente"} → {status}',
            request.remote_addr
        )
        db.session.commit()
        invalidar_cache_filtros()

        return jsonify({
            'success': True,
            'message': f'Solicitação {status.lower()} com sucesso!'
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error updating status for solicitacao {solicitacao_id}: {type(e).__name__}")
        return jsonify({'success': False, 'error': 'Erro interno. Tente novamente.'}), 500
    
def migrate_solicitacoes_compra_status():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # Verificar se a coluna status_aprovacao já existe
        cursor.execute("PRAGMA table_info(SolicitacoesCompra)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'status_aprovacao' not in columns:
            cursor.execute("ALTER TABLE SolicitacoesCompra ADD COLUMN status_aprovacao TEXT")
            conn.commit()
            logging.info("Coluna status_aprovacao adicionada à tabela SolicitacoesCompra")
        
        conn.close()
    except sqlite3.Error as e:
        logging.error(f"Erro na migração: {str(e)}")

@routes_bp.route('/atualizar_valores_cotacao', methods=['POST'])
def atualizar_valores_cotacao():
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        data = request.get_json()
        preenchimento_id = data.get('preenchimento_id')
        valor_unitario_str = data.get('valor_unitario', '0')
        valor_frete_str = data.get('valor_frete', '0')
        valor_unitario_original_str = data.get('valor_unitario_original', '0')
        observacoes = data.get('observacoes')
        
        # Validações
        if not preenchimento_id:
            return jsonify({'success': False, 'message': 'ID do preenchimento é obrigatório'}), 400
        
        # 🔴 CORREÇÃO: Função robusta para conversão
        def safe_currency_convert(value_str):
            if not value_str or value_str == '':
                return 0.0
            try:
                # Remove caracteres não numéricos exceto vírgula e ponto
                import re
                cleaned = re.sub(r'[^\d,.]', '', str(value_str))
                if not cleaned:
                    return 0.0
                # Remove pontos de milhar (preserva apenas o último ponto como decimal)
                if ',' in cleaned and '.' in cleaned:
                    # Formato: 1.234,56 -> remove pontos de milhar
                    cleaned = cleaned.replace('.', '')
                    cleaned = cleaned.replace(',', '.')
                elif ',' in cleaned:
                    # Formato: 1234,56
                    cleaned = cleaned.replace(',', '.')
                # Se só tem ponto, assume que é decimal
                return float(cleaned) if cleaned else 0.0
            except (ValueError, TypeError):
                return 0.0
        
        valor_unitario = safe_currency_convert(valor_unitario_str)
        valor_frete = safe_currency_convert(valor_frete_str)
        valor_unitario_original = safe_currency_convert(valor_unitario_original_str)
        
        print(f"💰 Atualizando: ID={preenchimento_id}, Unitário={valor_unitario}, Frete={valor_frete}")
        
        # Buscar o preenchimento
        preenchimento = SolicitacoesPreenchidas.query.get_or_404(preenchimento_id)

        # Verificar propriedade: apenas o próprio usuário ou master pode editar
        usuario_completo = session.get('usuario_completo', session['usuario'])
        senhas_dict = ler_senhas()
        role = senhas_dict.get(usuario_completo, {}).get('pagina', '')
        if preenchimento.usuario != session['usuario'] and 'menu_master' not in role:
            return jsonify({'success': False, 'message': 'Sem permissão para editar este preenchimento'}), 403
        
        # Salvar valores anteriores
        valor_unitario_anterior = preenchimento.valor_unitario or 0
        valor_frete_anterior = preenchimento.valor_frete or 0
        
        # Calcular novo valor total
        try:
            quantidade = float(preenchimento.solicitacao.quantidade) if preenchimento.solicitacao.quantidade else 0
        except (ValueError, TypeError):
            quantidade = 0
        
        novo_valor_total = (valor_unitario * quantidade) + valor_frete
        
        # Atualizar os valores
        preenchimento.valor_unitario = valor_unitario
        preenchimento.valor_frete = valor_frete if valor_frete > 0 else None
        preenchimento.valor_total = novo_valor_total
        preenchimento.observacoes = observacoes
        
        # Registrar histórico se houver mudança significativa
        if abs(valor_unitario_anterior - valor_unitario) > 0.001 or \
           abs(valor_frete_anterior - valor_frete) > 0.001:
            
            historico = HistoricoDescontos(
                preenchimento_id=preenchimento.id,
                valor_unitario_anterior=valor_unitario_anterior,
                valor_unitario_novo=valor_unitario,
                valor_frete_anterior=valor_frete_anterior if valor_frete_anterior != 0 else None,
                valor_frete_novo=valor_frete if valor_frete > 0 else None,
                data_alteracao=get_local_time(),
                usuario=session['usuario']
            )
            db.session.add(historico)
            print(f"📝 Histórico registrado para preenchimento {preenchimento_id}")
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': 'Valores atualizados com sucesso',
            'novo_valor_total': novo_valor_total,
            'valor_unitario': valor_unitario,
            'valor_frete': valor_frete
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro ao atualizar valores: {str(e)}")
        return jsonify({
            'success': False, 
            'message': 'Erro interno. Tente novamente.'
        }), 500
    
def create_historico_descontos_table():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        create_table_query = """
        CREATE TABLE IF NOT EXISTS HistoricoDescontos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            preenchimento_id INTEGER NOT NULL,
            valor_unitario_anterior REAL NOT NULL,
            valor_unitario_novo REAL NOT NULL,
            valor_frete_anterior REAL,
            valor_frete_novo REAL,
            data_alteracao DATETIME NOT NULL,
            usuario TEXT NOT NULL,
            FOREIGN KEY (preenchimento_id) REFERENCES SolicitacoesPreenchidas(id)
        )
        """
        cursor.execute(create_table_query)
        conn.commit()
        print("Tabela 'HistoricoDescontos' criada com sucesso.")
        cursor.close()
        conn.close()
        return True
    except sqlite3.Error as e:
        print(f"Erro ao criar a tabela HistoricoDescontos: {str(e)}")
        return False
    

@routes_bp.route('/historico_descontos', methods=['GET'])
def listar_historico_descontos():
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        # Obter parâmetros de filtro
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        usuario = request.args.get('usuario')
        preenchimento_id = request.args.get('preenchimento_id')
        
        # Query base
        query = db.session.query(HistoricoDescontos).join(
            SolicitacoesPreenchidas
        ).join(
            SolicitacoesCompra
        ).join(
            Materiais
        )
        
        # Aplicar filtros
        if data_inicio:
            query = query.filter(HistoricoDescontos.data_alteracao >= data_inicio)
        if data_fim:
            data_fim_ajustada = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(HistoricoDescontos.data_alteracao <= data_fim_ajustada)
        if usuario:
            query = query.filter(HistoricoDescontos.usuario.ilike(f'%{usuario}%'))
        if preenchimento_id:
            query = query.filter(HistoricoDescontos.preenchimento_id == preenchimento_id)
        
        # Executar query
        historicos = query.order_by(HistoricoDescontos.data_alteracao.desc()).all()
        
        # Obter lista de usuários para filtro
        usuarios = db.session.query(
            HistoricoDescontos.usuario
        ).distinct().order_by(
            HistoricoDescontos.usuario
        ).all()
        
        return render_template(
            'historico_descontos.html',
            historicos=historicos,
            usuarios=[u[0] for u in usuarios if u[0]],
            filtros={
                'data_inicio': data_inicio,
                'data_fim': data_fim,
                'usuario': usuario,
                'preenchimento_id': preenchimento_id
            }
        )
    
    except Exception as e:
        _flash_erro_seguro('Erro ao carregar histórico de descontos.', e, app.logger)
        app.logger.error(f'Erro em listar_historico_descontos: {str(e)}', exc_info=True)
        return render_template(
            'historico_descontos.html',
            historicos=[],
            usuarios=[],
            filtros={}
        )
def migrate_solicitacoes_compra_status_aprovacao():
    """Adiciona a coluna status_aprovacao na tabela SolicitacoesCompra se não existir"""
    try:
        # Usar SQLAlchemy em vez de sqlite3 direto
        from sqlalchemy import inspect
        
        inspector = inspect(db.engine)
        columns = [col['name'] for col in inspector.get_columns('SolicitacoesCompra')]
        
        if 'status_aprovacao' not in columns:
            # Usar SQLAlchemy para executar o ALTER TABLE
            db.engine.execute("ALTER TABLE SolicitacoesCompra ADD COLUMN status_aprovacao TEXT")
            logging.info("Coluna status_aprovacao adicionada à tabela SolicitacoesCompra")
            print("✓ Coluna status_aprovacao adicionada com sucesso!")
        else:
            print("✓ Coluna status_aprovacao já existe na tabela")
        
        return True
    except Exception as e:
        logging.error(f"Erro na migração status_aprovacao: {str(e)}")
        print(f"✗ Erro na migração: {str(e)}")
        return False
    
def add_observacoes_column_to_solicitacoes_preenchidas():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # Verificar se a coluna já existe
        cursor.execute("PRAGMA table_info(SolicitacoesPreenchidas)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'observacoes' not in columns:
            cursor.execute("ALTER TABLE SolicitacoesPreenchidas ADD COLUMN observacoes TEXT")
            conn.commit()
            logging.info("Coluna observacoes adicionada à tabela SolicitacoesPreenchidas")
            print("✓ Coluna observacoes adicionada com sucesso!")
        else:
            print("✓ Coluna observacoes já existe na tabela")
        
        conn.close()
        return True
    except sqlite3.Error as e:
        logging.error(f"Erro ao adicionar coluna observacoes: {str(e)}")
        print(f"✗ Erro ao adicionar coluna: {str(e)}")
        return False
    
def migrate_solicitacoes_preenchidas_status():
    """Migra o status padrão para suportar 'Rascunho'"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # Verificar se existem registros com status vazio ou nulo
        cursor.execute("UPDATE SolicitacoesPreenchidas SET status = 'Rascunho' WHERE status IS NULL OR status = ''")
        conn.commit()
        
        # Verificar se a atualização foi bem-sucedida
        cursor.execute("SELECT COUNT(*) FROM SolicitacoesPreenchidas WHERE status = 'Rascunho'")
        count = cursor.fetchone()[0]
        
        logging.info(f"Migração de status concluída: {count} registros atualizados para 'Rascunho'")
        print(f"✓ Migração de status concluída: {count} registros atualizados")
        
        conn.close()
        return True
    except sqlite3.Error as e:
        logging.error(f"Erro na migração de status: {str(e)}")
        print(f"✗ Erro na migração: {str(e)}")
        return False
    

@routes_bp.route('/teste_data_hora')
def teste_data_hora():
    import os as _os
    if _os.getenv("FLASK_ENV", "production").lower() not in ("development", "dev"):
        abort(404)
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    from datetime import datetime
    return f"""
    <h3>Teste de Data/Hora</h3>
    <p>datetime.utcnow(): {datetime.utcnow()}</p>
    <p>datetime.now(): {datetime.now()}</p>
    <p>get_local_time(): {get_local_time()}</p>
    <p>Ano atual: {get_local_time().year}</p>
    """

@app.template_filter('format_brasil_time')
def format_brasil_time(dt):
    if dt is None:
        return ""
    
    # Converte para fuso horário de Brasília se não tiver informação de timezone
    if dt.tzinfo is None:
        from datetime import timezone, timedelta
        brasil_tz = timezone(timedelta(hours=-3))
        dt = dt.replace(tzinfo=timezone.utc).astimezone(brasil_tz)
    
    return dt.strftime('%d/%m/%Y %H:%M')

def add_aplicacao_geral_column():
    """Adiciona a coluna aplicacao_geral na tabela SolicitacoesCompra"""
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # Verificar se a coluna já existe
        cursor.execute("PRAGMA table_info(SolicitacoesCompra)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'aplicacao_geral' not in columns:
            cursor.execute("ALTER TABLE SolicitacoesCompra ADD COLUMN aplicacao_geral TEXT")
            conn.commit()
            logging.info("Coluna aplicacao_geral adicionada à tabela SolicitacoesCompra")
            print("✓ Coluna aplicacao_geral adicionada com sucesso!")
        else:
            print("✓ Coluna aplicacao_geral já existe na tabela")
        
        conn.close()
        return True
    except sqlite3.Error as e:
        logging.error(f"Erro ao adicionar coluna aplicacao_geral: {str(e)}")
        print(f"✗ Erro ao adicionar coluna aplicacao_geral: {str(e)}")
        return False
    
#Tela comprador
# No app.py, substitua a função get_compradores por esta versão
# Definir o caminho absoluto para senhas.txt
SENHAS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'senhas.txt')

# Verificar se o arquivo existe no início
print(f"Verificando caminho de senhas.txt: {SENHAS_FILE}")
print(f"Arquivo existe: {os.path.exists(SENHAS_FILE)}")
    
def get_next_comprador(aplicacao):
    """
    Retorna o comprador a ser atribuído com base na aplicação.
    Se já houver um comprador atribuído a essa aplicação, reutiliza o mesmo.
    Caso contrário, distribui de forma balanceada entre os compradores disponíveis.
    """
    try:
        compradores = get_compradores()
        if not compradores:
            return None

        # Normaliza nome da aplicação
        aplicacao = aplicacao.strip().lower()

        # Verifica se já há comprador atribuído para essa aplicação
        solicit_existente = SolicitacoesCompra.query.filter_by(aplicacao=aplicacao).filter(
            SolicitacoesCompra.comprador_atribuido.isnot(None)
        ).first()

        if solicit_existente and solicit_existente.comprador_atribuido:
            # Reutiliza o comprador já existente para essa aplicação
            return solicit_existente.comprador_atribuido

        # Caso seja uma nova aplicação (sem histórico), distribui automaticamente
        todas_solicitacoes = SolicitacoesCompra.query.filter(
            SolicitacoesCompra.comprador_atribuido.isnot(None)
        ).all()

        # Conta quantas solicitações cada comprador já tem (para balanceamento)
        contagem = {c: 0 for c in compradores}
        for sol in todas_solicitacoes:
            if sol.comprador_atribuido in contagem:
                contagem[sol.comprador_atribuido] += 1

        # Seleciona o comprador com menor carga de solicitações
        comprador_escolhido = min(contagem, key=contagem.get)
        return comprador_escolhido

    except Exception as e:
        logging.error(f"Erro em get_next_comprador: {str(e)}")
        return None


# Adicione esta migração no init_db ou em uma função separada (ex: no bloco de migrações no final do app.py)
def add_comprador_atribuido_column():
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(SolicitacoesCompra)")
        columns = [col[1] for col in cursor.fetchall()]
        if 'comprador_atribuido' not in columns:
            cursor.execute("ALTER TABLE SolicitacoesCompra ADD COLUMN comprador_atribuido TEXT")
            conn.commit()
            logging.info("Coluna comprador_atribuido adicionada à tabela SolicitacoesCompra")
            print("✓ Coluna comprador_atribuido adicionada com sucesso!")
        else:
            print("✓ Coluna comprador_atribuido já existe na tabela")
        conn.close()
        return True
    except sqlite3.Error as e:
        logging.error(f"Erro ao adicionar coluna comprador_atribuido: {str(e)}")
        print(f"✗ Erro ao adicionar coluna: {str(e)}")
        return False

# No bloco if __name__ == '__main__': adicione esta chamada à migração
add_comprador_atribuido_column()

# Adicione esta rota no app.py, no bloco de rotas (após outras rotas semelhantes)
@routes_bp.route('/api/solicitacoes/<int:solicitacao_id>/atribuir_comprador', methods=['POST'])
def atribuir_comprador(solicitacao_id):
    if 'usuario' not in session:
        return jsonify({'success': False, 'error': 'Usuário não autenticado'}), 401

    # RBAC: apenas master e supervisor podem reatribuir compradores
    _uc_atrib = session.get('usuario_completo', session['usuario'])
    _role_atrib = ler_senhas().get(_uc_atrib, {}).get('pagina', '')
    if not any(r in _role_atrib for r in ('menu_master', 'menu_supervisor', 'menu_aprovacao')):
        registrar_log(_uc_atrib, 'acesso_negado', f'Tentativa de atribuir comprador em solicitação ID {solicitacao_id}', request.remote_addr)
        return jsonify({'success': False, 'error': 'Sem permissão para atribuir compradores'}), 403
    try:
        data = request.get_json()
        comprador = data.get('comprador')
        if not comprador:
            return jsonify({'success': False, 'error': 'Nenhum comprador selecionado'}), 400
        solicitacao = SolicitacoesCompra.query.get_or_404(solicitacao_id)
        solicitacao.comprador_atribuido = comprador
        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'Solicitação atribuída a {comprador} com sucesso!'
        })
    except Exception as e:
        db.session.rollback()
        logging.error(f"Erro ao atribuir comprador: {str(e)}")
        return jsonify({'success': False, 'error': 'Erro interno. Tente novamente.'}), 500
    

@app.template_filter('basename')
def basename_filter(path):
    if not path:
        return ''
    from os.path import basename
    return basename(path)

app.jinja_env.filters['basename'] = basename_filter


@routes_bp.route('/listar_solicitacoes_finalizadas', methods=['GET'])
def listar_solicitacoes_finalizadas():
    if 'usuario' not in session:
        flash('Acesso não autorizado', 'error')
        return redirect(url_for('routes_bp.login'))
    
    try:
        # Filtros básicos para evitar carregar todo o banco em memória
        empresa_f   = request.args.get('empresa', '').strip()
        usuario_f   = request.args.get('usuario', '').strip()
        aplicacao_f = request.args.get('aplicacao', '').strip()
        comprador_f = request.args.get('comprador', '').strip()
        page        = request.args.get('page', 1, type=int)
        per_page    = 50

        query = SolicitacoesCompra.query.join(Materiais)
        if empresa_f:
            query = query.filter(SolicitacoesCompra.empresa == empresa_f)
        if usuario_f:
            query = query.filter(SolicitacoesCompra.usuario == usuario_f)
        if aplicacao_f:
            query = query.filter(SolicitacoesCompra.aplicacao == aplicacao_f)
        if comprador_f:
            query = query.filter(SolicitacoesCompra.comprador_atribuido == comprador_f)

        paginacao   = query.order_by(SolicitacoesCompra.data_solicitacao.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        solicitacoes = paginacao.items

        # Listas de filtros via helper cacheado
        _filtros = get_filtros_listagem(apenas_aprovadas=False)
        empresas   = _filtros['empresas']
        usuarios   = _filtros['usuarios']
        aplicacoes = _filtros['aplicacoes']
        compradores = get_compradores()

        if not compradores:
            flash('Nenhum comprador encontrado no arquivo senhas.txt. Verifique o arquivo ou a função get_compradores.', 'error')
        
        return render_template(
            'listar_solicitacoes.html',
            solicitacoes=solicitacoes,
            empresas=empresas,
            usuarios=usuarios,
            aplicacoes=aplicacoes,
            compradores=compradores,
            paginacao=paginacao,
        )
    
    except Exception as e:
        _flash_erro_seguro('Erro ao carregar solicitações.', e, app.logger)
        logging.error(f"Erro em listar_solicitacoes_finalizadas: {str(e)}")
        print(f"ERRO: Erro em listar_solicitacoes_finalizadas: {str(e)}")
        return render_template(
            'listar_solicitacoes.html',
            solicitacoes=[],
            empresas=[],
            usuarios=[],
            aplicacoes=[],
            compradores=[]  # Passa lista vazia em caso de erro
        )
    
def get_fornecedor_details(fornecedor_id):
    """Busca nome_fantasia e cnpj do fornecedor pelo ID no banco fornecedores.db"""
    try:
        conn = sqlite3.connect(DB_PATH_FORNECEDORES)
        cursor = conn.cursor()
        cursor.execute('SELECT nome_fantasia, cnpj FROM fornecedores WHERE id = ?', (fornecedor_id,))
        result = cursor.fetchone()
        conn.close()
        if result:
            return result[0], result[1]
        return None, None
    except sqlite3.Error as e:
        logging.error('get_fornecedor_details – erro: %s', type(e).__name__)
        return None, None


def get_fornecedores_bulk(fornecedor_ids: set) -> dict:
    """
    Busca fornecedores em lote com UMA única query e UMA conexão.
    Retorna {id: {'nome_fantasia': ..., 'cnpj': ...}}.
    Substitui o anti-pattern de abrir conn/cursor por linha no loop.
    Usa context manager para garantir fechamento mesmo em exceção.
    """
    if not fornecedor_ids:
        return {}
    result = {}
    try:
        with sqlite3.connect(DB_PATH_FORNECEDORES) as conn:
            placeholders = ','.join('?' * len(fornecedor_ids))
            rows = conn.execute(
                f'SELECT id, nome_fantasia, cnpj FROM fornecedores WHERE id IN ({placeholders})',
                list(fornecedor_ids)
            ).fetchall()
            for row_id, nome, cnpj in rows:
                result[row_id] = {
                    'nome_fantasia': nome or 'Fornecedor não encontrado',
                    'cnpj': format_cnpj(cnpj) if cnpj else 'N/A',
                }
    except sqlite3.Error as e:
        logging.error('get_fornecedores_bulk – erro de banco: %s', type(e).__name__, exc_info=True)
    except Exception as e:
        logging.error('get_fornecedores_bulk – erro inesperado: %s', type(e).__name__, exc_info=True)
    return result

@routes_bp.route('/api/fornecedores', methods=['GET'])
def api_fornecedores():
    """API para fornecedores com paginação no servidor"""
    if 'usuario' not in session:
        return jsonify({'error': 'Não autenticado'}), 401
    
    try:
        # Parâmetros de paginação
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 25, type=int)
        search_name = request.args.get('search_name', '').strip()
        search_material = request.args.get('search_material', '').strip()
        
        # Conexão com o banco de fornecedores
        conn = get_db_connection(DB_PATH_FORNECEDORES)
        if not conn:
            return jsonify({'error': 'Erro ao conectar ao banco'}), 500
        
        try:
            cursor = conn.cursor()
            
            # Query base com filtros
            query = '''
                SELECT id, nome_fantasia, cnpj, telefone, email, endereco, 
                       bairro, cidade, estado, contato, materiais
                FROM fornecedores 
                WHERE 1=1
            '''
            params = []
            
            if search_name:
                query += ' AND (nome_fantasia LIKE ? OR cnpj LIKE ?)'
                params.extend([f'%{search_name}%', f'%{search_name}%'])
            
            if search_material:
                query += ' AND materiais LIKE ?'
                params.append(f'%{search_material}%')
            
            # Ordenação
            query += ' ORDER BY nome_fantasia'
            
            # Contagem total (para paginação)
            count_query = 'SELECT COUNT(*) FROM fornecedores WHERE 1=1'
            count_params = []
            
            if search_name:
                count_query += ' AND (nome_fantasia LIKE ? OR cnpj LIKE ?)'
                count_params.extend([f'%{search_name}%', f'%{search_name}%'])
            
            if search_material:
                count_query += ' AND materiais LIKE ?'
                count_params.append(f'%{search_material}%')
            
            cursor.execute(count_query, count_params)
            total_count = cursor.fetchone()[0]
            
            # Query com paginação
            query += ' LIMIT ? OFFSET ?'
            offset = (page - 1) * per_page
            params.extend([per_page, offset])
            
            cursor.execute(query, params)
            fornecedores = []
            
            for row in cursor.fetchall():
                fornecedor = {
                    'id': row[0],
                    'nome_fantasia': row[1],
                    'cnpj': row[2],
                    'telefone': row[3],
                    'email': row[4],
                    'endereco': row[5],
                    'bairro': row[6],
                    'cidade': row[7],
                    'estado': row[8],
                    'contato': row[9],
                    'materiais': row[10],
                    'cnpj_formatado': format_cnpj(row[2]) if row[2] else ''
                }
                fornecedores.append(fornecedor)
            
            return jsonify({
                'success': True,
                'fornecedores': fornecedores,
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total_count': total_count,
                    'total_pages': (total_count + per_page - 1) // per_page
                }
            })
            
        finally:
            conn.close()
            
    except Exception as e:
        logging.error(f"Erro na API de fornecedores: {str(e)}")
        app.logger.error(f'API error: {type(e).__name__}', exc_info=False)
        return jsonify({'error': 'Erro interno. Tente novamente.'}), 500
    
@routes_bp.route('/api/excluir_grupo_aplicacao', methods=['POST'])
def excluir_grupo_aplicacao():
    """Exclui todas as solicitações de um grupo (aplicação) e seus relacionamentos"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'error': 'Usuário não autenticado'}), 401
    
    try:
        data = request.get_json()
        aplicacao = data.get('aplicacao')
        
        if not aplicacao:
            return jsonify({'success': False, 'error': 'Aplicação não especificada'}), 400
        
        usuario = session['usuario']
        print(f"🚨 Usuário {usuario} solicitando exclusão do grupo: {aplicacao}")
        
        # 1. Buscar todas as solicitações APROVADAS desta aplicação
        solicitacoes = SolicitacoesCompra.query.filter_by(
            aplicacao=aplicacao,
            status_aprovacao='Aprovado'
        ).all()
        
        if not solicitacoes:
            return jsonify({'success': False, 'error': f'Nenhuma solicitação APROVADA encontrada para "{aplicacao}"'}), 404
        
        ids_solicitacoes = [s.id for s in solicitacoes]
        count_solicitacoes = len(ids_solicitacoes)
        
        print(f"📋 Encontradas {count_solicitacoes} solicitações para exclusão")
        
        # 2. Buscar todos os preenchimentos relacionados
        preenchimentos = SolicitacoesPreenchidas.query.filter(
            SolicitacoesPreenchidas.solicitacao_id.in_(ids_solicitacoes)
        ).all()
        
        ids_preenchimentos = [p.id for p in preenchimentos]
        count_preenchimentos = len(ids_preenchimentos)
        
        print(f"📋 Encontrados {count_preenchimentos} preenchimentos para exclusão")
        
        # 3. Buscar e excluir histórico de descontos
        historicos = HistoricoDescontos.query.filter(
            HistoricoDescontos.preenchimento_id.in_(ids_preenchimentos)
        ).all()
        
        for historico in historicos:
            db.session.delete(historico)
        
        count_historicos = len(historicos)
        print(f"📋 Encontrados {count_historicos} históricos de desconto para exclusão")
        
        # 4. Verificar se há pedidos de compra associados
        pedidos_associados = []
        if ids_preenchimentos:
            # Buscar pedidos que têm esses preenchimentos
            pedidos = PedidosCompra.query.join(
                pedido_preenchimento_associacao,
                PedidosCompra.id == pedido_preenchimento_associacao.c.pedido_id
            ).filter(
                pedido_preenchimento_associacao.c.preenchimento_id.in_(ids_preenchimentos)
            ).distinct().all()
            
            pedidos_associados = [p.numero_pedido for p in pedidos]
        
        if pedidos_associados:
            return jsonify({
                'success': False,
                'error': f'Não é possível excluir: existem pedidos de compra associados ({len(pedidos_associados)} pedidos)',
                'pedidos': pedidos_associados
            }), 400
        
        # 5. Verificar se há registros em outras tabelas
        # Estoque
        estoques = Estoque.query.filter(
            Estoque.preenchimento_id.in_(ids_preenchimentos)
        ).all()
        
        if estoques:
            for estoque in estoques:
                db.session.delete(estoque)
            print(f"📋 Excluídos {len(estoques)} registros de estoque")
        
        # Requisições
        requisicoes = Requisicoes.query.filter(
            Requisicoes.preenchimento_id.in_(ids_preenchimentos)
        ).all()
        
        if requisicoes:
            for requisicao in requisicoes:
                db.session.delete(requisicao)
            print(f"📋 Excluídos {len(requisicoes)} registros de requisições")
        
        # Auditoria
        auditorias = Auditoria.query.filter(
            Auditoria.solicitacao_id.in_(ids_solicitacoes)
        ).all()
        
        if auditorias:
            for auditoria in auditorias:
                db.session.delete(auditoria)
            print(f"📋 Excluídos {len(auditorias)} registros de auditoria")
        
        # 6. Excluir preenchimentos
        for preenchimento in preenchimentos:
            db.session.delete(preenchimento)
        
        # 7. Excluir solicitações
        for solicitacao in solicitacoes:
            db.session.delete(solicitacao)
        
        # 8. Commit da transação
        db.session.commit()
        
        # 9. Log detalhado
        logging.info(f"""
        ✅ GRUPO EXCLUÍDO COM SUCESSO
        Aplicação: {aplicacao}
        Usuário: {usuario}
        Data/Hora: {get_local_time()}
        Estatísticas:
          - Solicitações: {count_solicitacoes}
          - Preenchimentos: {count_preenchimentos}
          - Históricos: {count_historicos}
          - Estoques: {len(estoques) if estoques else 0}
          - Requisições: {len(requisicoes) if requisicoes else 0}
          - Auditorias: {len(auditorias) if auditorias else 0}
        """)
        
        return jsonify({
            'success': True,
            'message': f'✅ Grupo "{aplicacao}" excluído com sucesso!',
            'details': {
                'solicitacoes_excluidas': count_solicitacoes,
                'preenchimentos_excluidas': count_preenchimentos,
                'historicos_excluidos': count_historicos,
                'estoques_excluidos': len(estoques) if estoques else 0,
                'requisicoes_excluidas': len(requisicoes) if requisicoes else 0,
                'auditorias_excluidas': len(auditorias) if auditorias else 0
            }
        })
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"❌ ERRO AO EXCLUIR GRUPO {aplicacao}: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Erro interno. Tente novamente.'
        }), 500

@routes_bp.route('/editar_fornecedor/<int:fornecedor_id>', methods=['GET', 'POST'])
def editar_fornecedor(fornecedor_id):
    """Rota para editar fornecedor (GET para carregar, POST para salvar)"""
    if 'usuario' not in session:
        flash('Você precisa estar logado para editar fornecedores.', 'warning')
        return redirect(url_for('routes_bp.login'))
    
    conn = get_db_connection(DB_PATH_FORNECEDORES)
    if not conn:
        flash('Erro ao conectar ao banco de dados.', 'error')
        return redirect(url_for('routes_bp.lista_fornecedores'))
    
    try:
        cursor = conn.cursor()
        
        if request.method == 'GET':
            # Buscar dados do fornecedor para edição
            cursor.execute('''
                SELECT id, nome_fantasia, cnpj, telefone, email, endereco, 
                       bairro, cidade, estado, contato, materiais
                FROM fornecedores WHERE id = ?
            ''', (fornecedor_id,))
            
            fornecedor_raw = cursor.fetchone()
            
            if not fornecedor_raw:
                flash('Fornecedor não encontrado.', 'error')
                return redirect(url_for('routes_bp.lista_fornecedores'))
            
            # Converter para dicionário
            fornecedor = {
                'id': fornecedor_raw[0],
                'nome_fantasia': fornecedor_raw[1],
                'cnpj': fornecedor_raw[2],
                'telefone': fornecedor_raw[3],
                'email': fornecedor_raw[4],
                'endereco': fornecedor_raw[5],
                'bairro': fornecedor_raw[6],
                'cidade': fornecedor_raw[7],
                'estado': fornecedor_raw[8],
                'contato': fornecedor_raw[9],
                'materiais': fornecedor_raw[10]
            }
            
            return render_template('editar_fornecedor.html', fornecedor=fornecedor)
        
        elif request.method == 'POST':
            # Obter dados do formulário de edição
            nome_fantasia = request.form.get('nome_fantasia', '').strip()
            telefone = request.form.get('telefone', '').strip()
            email = request.form.get('email', '').strip()
            endereco = request.form.get('endereco', '').strip()
            bairro = request.form.get('bairro', '').strip()
            cidade = request.form.get('cidade', '').strip()
            estado = request.form.get('estado', '').strip()
            contato = request.form.get('contato', '').strip()
            materiais = request.form.get('materiais', '').strip()
            
            # Validações (exceto CNPJ)
            if not all([nome_fantasia, telefone, email, endereco, bairro, cidade, estado, contato, materiais]):
                flash('Todos os campos são obrigatórios.', 'error')
                return redirect(url_for('routes_bp.editar_fornecedor', fornecedor_id=fornecedor_id))
            
            if not validate_email(email):
                flash('Email inválido. Insira um email válido.', 'error')
                return redirect(url_for('routes_bp.editar_fornecedor', fornecedor_id=fornecedor_id))
            
            if len(telefone) < 10:
                flash('Telefone inválido. Deve conter pelo menos 10 dígitos.', 'error')
                return redirect(url_for('routes_bp.editar_fornecedor', fornecedor_id=fornecedor_id))
            
            # Atualizar fornecedor (exceto CNPJ)
            cursor.execute('''
                UPDATE fornecedores SET 
                    nome_fantasia = ?,
                    telefone = ?,
                    email = ?,
                    endereco = ?,
                    bairro = ?,
                    cidade = ?,
                    estado = ?,
                    contato = ?,
                    materiais = ?
                WHERE id = ?
            ''', (nome_fantasia, telefone, email, endereco, bairro, cidade, estado, contato, materiais, fornecedor_id))
            
            conn.commit()
            flash('Fornecedor atualizado com sucesso!', 'success')
            return redirect(url_for('routes_bp.lista_fornecedores'))
            
    except Exception as e:
        if conn:
            conn.rollback()
        _flash_erro_seguro('Erro ao editar fornecedor.', e, app.logger)
        return redirect(url_for('routes_bp.lista_fornecedores'))
    finally:
        if conn:
            conn.close()

@routes_bp.route('/excluir_fornecedor/<int:fornecedor_id>', methods=['POST'])
def excluir_fornecedor(fornecedor_id):
    """Exclui um fornecedor - versão simplificada e testada"""
    if 'usuario' not in session:
        flash('Você precisa estar logado.', 'warning')
        return redirect(url_for('routes_bp.login'))
    
    try:
        # Conectar ao banco de fornecedores
        conn = sqlite3.connect(DB_PATH_FORNECEDORES)
        cursor = conn.cursor()
        
        # 1. Buscar nome do fornecedor para mensagem
        cursor.execute('SELECT nome_fantasia FROM fornecedores WHERE id = ?', (fornecedor_id,))
        resultado = cursor.fetchone()
        
        if not resultado:
            flash('Fornecedor não encontrado.', 'error')
            return redirect(url_for('routes_bp.lista_fornecedores'))
        
        nome_fornecedor = resultado[0]
        
        # 2. Verificar se está em uso (opcional - comente se quiser forçar exclusão)
        try:
            conn_main = sqlite3.connect(DATABASE)
            cursor_main = conn_main.cursor()
            cursor_main.execute('SELECT COUNT(*) FROM SolicitacoesPreenchidas WHERE fornecedor_id = ?', (fornecedor_id,))
            uso_count = cursor_main.fetchone()[0]
            conn_main.close()
            
            if uso_count > 0:
                flash(f'❌ Fornecedor "{nome_fornecedor}" está em {uso_count} cotações. Não pode ser excluído.', 'error')
                return redirect(url_for('routes_bp.lista_fornecedores'))
        except Exception as e:
            print(f"AVISO: Não foi possível verificar uso - {e}")
            # Continua mesmo sem verificar
        
        # 3. Excluir fornecedor
        cursor.execute('DELETE FROM fornecedores WHERE id = ?', (fornecedor_id,))
        conn.commit()
        conn.close()
        
        flash(f'✅ Fornecedor "{nome_fornecedor}" excluído com sucesso!', 'success')
        
    except Exception as e:
        print(f"ERRO CRÍTICO ao excluir: {str(e)}")
        _flash_erro_seguro('❌ Erro ao excluir fornecedor.', e, app.logger)
    
    return redirect(url_for('routes_bp.lista_fornecedores'))

    
@routes_bp.route('/api/fornecedor/<int:fornecedor_id>', methods=['GET'])
def api_get_fornecedor(fornecedor_id):
    """API para obter dados de um fornecedor específico"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'error': 'Não autenticado'}), 401
    
    try:
        conn = get_db_connection(DB_PATH_FORNECEDORES)
        if not conn:
            return jsonify({'success': False, 'error': 'Erro ao conectar ao banco'}), 500
        
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, nome_fantasia, cnpj, telefone, email, endereco, 
                   bairro, cidade, estado, contato, materiais
            FROM fornecedores WHERE id = ?
        ''', (fornecedor_id,))
        
        fornecedor_raw = cursor.fetchone()
        conn.close()
        
        if not fornecedor_raw:
            return jsonify({'success': False, 'error': 'Fornecedor não encontrado'}), 404
        
        fornecedor = {
            'id': fornecedor_raw[0],
            'nome_fantasia': fornecedor_raw[1],
            'cnpj': fornecedor_raw[2],
            'telefone': fornecedor_raw[3],
            'email': fornecedor_raw[4],
            'endereco': fornecedor_raw[5],
            'bairro': fornecedor_raw[6],
            'cidade': fornecedor_raw[7],
            'estado': fornecedor_raw[8],
            'contato': fornecedor_raw[9],
            'materiais': fornecedor_raw[10],
            'cnpj_formatado': format_cnpj(fornecedor_raw[2]) if fornecedor_raw[2] else ''
        }
        
        return jsonify({
            'success': True,
            'fornecedor': fornecedor
        })
        
    except Exception as e:
        logging.error(f"Erro na API de fornecedor específico: {str(e)}")
        app.logger.error(f'API error: {type(e).__name__}', exc_info=False)
        return jsonify({'success': False, 'error': 'Erro interno. Tente novamente.'}), 500

@routes_bp.route('/duplicar_grupo_aplicacao', methods=['POST'])
def duplicar_grupo_aplicacao():
    """Duplica todas as solicitações de um grupo (aplicação)"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'error': 'Usuário não autenticado'}), 401
    
    try:
        data = request.get_json()
        aplicacao_original = data.get('aplicacao')
        
        if not aplicacao_original:
            return jsonify({'success': False, 'error': 'Aplicação não especificada'}), 400
        
        # Buscar todas as solicitações da aplicação original
        solicitacoes_originais = SolicitacoesCompra.query.filter(
            SolicitacoesCompra.aplicacao == aplicacao_original,
            SolicitacoesCompra.status_aprovacao == "Aprovado"
        ).all()
        
        if not solicitacoes_originais:
            return jsonify({'success': False, 'error': 'Nenhuma solicitação encontrada para esta aplicação'}), 404
        
        # Criar nova aplicação com sufixo _copia
        aplicacao_nova = f"{aplicacao_original}_copia"
        usuario = session['usuario']
        
        # Contador de duplicações bem-sucedidas
        duplicacoes_criadas = 0
        
        # Duplicar cada solicitação
        for solicitacao_original in solicitacoes_originais:
            try:
                # Criar nova solicitação baseada na original
                nova_solicitacao = SolicitacoesCompra(
                    cod_material=solicitacao_original.cod_material,
                    especificacao=solicitacao_original.especificacao,
                    quantidade=solicitacao_original.quantidade,
                    unidade_medida=solicitacao_original.unidade_medida,
                    aplicacao=aplicacao_nova,
                    aplicacao_geral=aplicacao_nova,  # Aplicação geral também atualizada
                    empresa=solicitacao_original.empresa,
                    usuario=usuario,
                    foto_path=solicitacao_original.foto_path,
                    marca=solicitacao_original.marca,
                    ativo=solicitacao_original.ativo,
                    nome_ativo=solicitacao_original.nome_ativo,
                    prioridade=solicitacao_original.prioridade,
                    status_aprovacao=None,  # Nova solicitação começa sem aprovação
                    comprador_atribuido=None  # Resetar comprador atribuído
                )
                
                db.session.add(nova_solicitacao)
                duplicacoes_criadas += 1
                
            except Exception as e:
                db.session.rollback()
                logging.error(f"Erro ao duplicar solicitação {solicitacao_original.id}: {str(e)}")
                return jsonify({
                    'success': False, 
                    'error': 'Erro interno. Tente novamente.'
                }), 500
        
        # Commit de todas as novas solicitações
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Grupo duplicado com sucesso! {duplicacoes_criadas} solicitações criadas.',
            'aplicacao_nova': aplicacao_nova,
            'count': duplicacoes_criadas
        })
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Erro geral ao duplicar grupo: {str(e)}")
        return jsonify({
            'success': False, 
            'error': 'Erro interno. Tente novamente.'
        }), 500   
    
@routes_bp.route('/reprovar_solicitacao_individual/<int:id>', methods=['POST'])
def reprovar_solicitacao_individual(id):
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401

    # RBAC: apenas aprovação, supervisor e master podem reprovar
    _uc_repr = session.get('usuario_completo', session['usuario'])
    _role_repr = ler_senhas().get(_uc_repr, {}).get('pagina', '')
    if not any(r in _role_repr for r in ('menu_aprovacao', 'menu_master', 'menu_supervisor')):
        registrar_log(_uc_repr, 'acesso_negado', f'Tentativa de reprovar solicitação ID {id} sem permissão', request.remote_addr)
        return jsonify({'success': False, 'message': 'Sem permissão para reprovar solicitações'}), 403
    
    try:
        # Buscar a solicitação
        solicitacao = SolicitacoesCompra.query.get_or_404(id)
        usuario = session['usuario']
        
        # Verificar se já está reprovada
        if solicitacao.status_aprovacao == 'Reprovado':
            return jsonify({
                'success': False, 
                'message': 'Esta solicitação já está reprovada.'
            }), 400
        
        # REMOVIDA A VERIFICAÇÃO DE COTAÇÕES
        # A reprovação agora é permitida independentemente de ter cotações
        
        # Obter motivo do corpo da requisição
        data = request.get_json()
        motivo = data.get('motivo', '').strip() if data else ''
        
        # Atualizar status para reprovado
        solicitacao.status_aprovacao = 'Reprovado'
        
        # Adicionar observação se houver motivo
        if motivo:
            if solicitacao.observacoes_col:
                solicitacao.observacoes_col += f"\nREPROVADA INDIVIDUALMENTE - Motivo: {motivo} - Usuário: {usuario}"
            else:
                solicitacao.observacoes_col = f"REPROVADA INDIVIDUALMENTE - Motivo: {motivo} - Usuário: {usuario}"
        
        # Registrar log
        ip = request.remote_addr
        registrar_log(usuario, 'reprovar_solicitacao_individual', 
                     f'Solicitação ID {id} reprovada individualmente. Motivo: {motivo}', ip)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Solicitação reprovada com sucesso!'
        })
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Erro ao reprovar solicitação {id}: {str(e)}")
        return jsonify({
            'success': False, 
            'message': 'Erro interno. Tente novamente.'
        }), 500

# Segmentar Grupo

@routes_bp.route('/segmentar_grupo', methods=['POST'])
def segmentar_grupo():
    """Cria uma nova aplicação com as solicitações selecionadas"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'error': 'Usuário não autenticado'}), 401
    
    try:
        data = request.get_json()
        solicitacoes_ids = data.get('solicitacoes_ids', [])
        nova_aplicacao = data.get('nova_aplicacao', '').strip()
        senha = data.get('senha', '').strip()
        
        # Validações
        if not solicitacoes_ids:
            return jsonify({'success': False, 'error': 'Nenhuma solicitação selecionada'}), 400
        
        if not nova_aplicacao:
            return jsonify({'success': False, 'error': 'Nome da nova aplicação é obrigatório'}), 400
        
        if senha != '122004':
            return jsonify({'success': False, 'error': 'Senha administrativa inválida'}), 403
        
        # Buscar as solicitações originais
        solicitacoes_originais = SolicitacoesCompra.query.filter(
            SolicitacoesCompra.id.in_(solicitacoes_ids)
        ).all()
        
        if not solicitacoes_originais:
            return jsonify({'success': False, 'error': 'Solicitações não encontradas'}), 404
        
        usuario = session['usuario']
        solicitacoes_criadas = 0
        
        # Duplicar cada solicitação com a nova aplicação
        for sol_original in solicitacoes_originais:
            try:
                nova_solicitacao = SolicitacoesCompra(
                    cod_material=sol_original.cod_material,
                    especificacao=sol_original.especificacao,
                    quantidade=sol_original.quantidade,
                    unidade_medida=sol_original.unidade_medida,
                    aplicacao=nova_aplicacao,
                    aplicacao_geral=nova_aplicacao,
                    empresa=sol_original.empresa,
                    usuario=usuario,
                    foto_path=sol_original.foto_path,
                    marca=sol_original.marca,
                    ativo=sol_original.ativo,
                    nome_ativo=sol_original.nome_ativo,
                    prioridade=sol_original.prioridade,
                    status_aprovacao=None,  # Nova solicitação começa pendente
                    comprador_atribuido=None
                )
                db.session.add(nova_solicitacao)
                solicitacoes_criadas += 1
                
            except Exception as e:
                db.session.rollback()
                logging.error(f"Erro ao segmentar solicitação {sol_original.id}: {str(e)}")
                return jsonify({
                    'success': False, 
                    'error': 'Erro interno. Tente novamente.'
                }), 500
        
        # Commit de todas as novas solicitações
        db.session.commit()
        
        # Log da operação
        logging.info(f"""
        ✅ GRUPO SEGMENTADO COM SUCESSO
        Nova aplicação: {nova_aplicacao}
        Usuário: {usuario}
        Solicitações originais: {len(solicitacoes_ids)}
        Solicitações criadas: {solicitacoes_criadas}
        """)
        
        return jsonify({
            'success': True,
            'message': f'✅ Grupo segmentado com sucesso! {solicitacoes_criadas} solicitações criadas.',
            'nova_aplicacao': nova_aplicacao,
            'count': solicitacoes_criadas
        })
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Erro geral ao segmentar grupo: {str(e)}", exc_info=True)
        return jsonify({
            'success': False, 
            'error': 'Erro interno. Tente novamente.'
        }), 500

# ===================================================
# RELATÓRIO DE PEDIDOS DE COMPRA
# ===================================================

@routes_bp.route('/relatorio_pedidos', methods=['GET'])
def relatorio_pedidos():
    """Relatório agrupado por Pedido de Compra"""
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        # Parâmetros de filtro
        empresa = request.args.get('empresa')
        status = request.args.get('status')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        numero_pedido = request.args.get('numero_pedido')
        
        # Query base
        query = db.session.query(PedidosCompra)
        
        # Aplicar filtros
        if status:
            query = query.filter(PedidosCompra.status == status)
        
        if numero_pedido:
            query = query.filter(PedidosCompra.numero_pedido.ilike(f'%{numero_pedido}%'))
        
        if data_inicio:
            query = query.filter(PedidosCompra.data_criacao >= data_inicio)
        
        if data_fim:
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(PedidosCompra.data_criacao < dt_fim)
        
        # Executar query
        pedidos = query.order_by(PedidosCompra.data_criacao.desc()).all()
        
        # Estruturar dados para o relatório
        relatorio_data = []
        
        for pedido in pedidos:
            # Obter informações das solicitações associadas
            empresas_solicitacoes = set()
            compradores_set = set()
            materiais_list = []
            valor_total_pedido = 0
            
            for preenchimento in pedido.preenchimentos:
                solicitacao = preenchimento.solicitacao
                if solicitacao:
                    empresas_solicitacoes.add(solicitacao.empresa)
                    if solicitacao.comprador_atribuido:
                        compradores_set.add(solicitacao.comprador_atribuido)
                    
                    # Converter quantidade para float se for string
                    quantidade = solicitacao.quantidade
                    if isinstance(quantidade, str):
                        try:
                            quantidade = float(quantidade.replace(',', '.'))
                        except:
                            quantidade = 0
                    
                    material_info = {
                        'id_solicitacao': solicitacao.id,
                        'material': solicitacao.material.DescricaoMaterial if solicitacao.material else 'N/A',
                        'cod_material': solicitacao.cod_material,
                        'especificacao': solicitacao.especificacao,
                        'quantidade': quantidade,
                        'unidade': solicitacao.unidade_medida,
                        'marca': solicitacao.marca or 'Não informado',
                        'valor_unitario': preenchimento.valor_unitario or 0,
                        'valor_total_item': preenchimento.valor_total or 0,
                        'fornecedor': get_fornecedor_nome(preenchimento.fornecedor_id),
                        'prazo_entrega': preenchimento.prazo_entrega or 'N/A',
                        'condicao_pagamento': preenchimento.condicao_pagamento or 'N/A'
                    }
                    materiais_list.append(material_info)
                    valor_total_pedido += (preenchimento.valor_total or 0)
            
            # Obter fornecedores do pedido
            fornecedores_pedido = set()
            for p in pedido.preenchimentos:
                fornecedores_pedido.add(get_fornecedor_nome(p.fornecedor_id))
            
            # Aplicar filtro de empresa (se necessário)
            if empresa and empresa not in empresas_solicitacoes:
                continue
            
            # Calcular estatísticas
            total_itens = len(materiais_list)
            total_fornecedores = len(fornecedores_pedido)
            
            relatorio_data.append({
                'pedido': pedido,
                'numero_pedido': pedido.numero_pedido,
                'data_criacao': pedido.data_criacao,
                'status': pedido.status,
                'usuario': pedido.usuario,
                'empresas': ', '.join(sorted(empresas_solicitacoes)) if empresas_solicitacoes else 'N/A',
                'compradores': ', '.join(sorted(compradores_set)) if compradores_set else 'Não atribuído',
                'fornecedores': ', '.join(sorted(fornecedores_pedido)) if fornecedores_pedido else 'N/A',
                'valor_total': valor_total_pedido,
                'valor_frete': pedido.valor_frete or 0,
                'valor_liquido': pedido.valor_liquido,
                'forma_pagamento': pedido.forma_pagamento or 'N/A',
                'observacoes': pedido.observacoes or '',
                'total_itens': total_itens,
                'total_fornecedores': total_fornecedores,
                'materiais': materiais_list,
                'preenchimentos': pedido.preenchimentos
            })
        
        # Obter listas para filtros
        empresas_list = []
        try:
            empresas_query = db.session.query(SolicitacoesCompra.empresa).distinct().all()
            empresas_list = sorted([e[0] for e in empresas_query if e[0]])
        except:
            pass
        
        status_list = ['Gerado', 'Aprovado', 'AgPagamento', 'AgEntrega', 'Entregue', 'Reprovado']
        
        return render_template(
            'relatorio_pedidos.html',
            relatorio=relatorio_data,
            empresas=empresas_list,
            status_list=status_list,
            filtros={
                'empresa': empresa,
                'status': status,
                'data_inicio': data_inicio,
                'data_fim': data_fim,
                'numero_pedido': numero_pedido
            }
        )
        
    except Exception as e:
        _flash_erro_seguro('Erro ao gerar relatório.', e, app.logger)
        app.logger.error(f'Erro em relatorio_pedidos: {str(e)}', exc_info=True)
        return render_template('relatorio_pedidos.html', relatorio=[], empresas=[], status_list=[], filtros={})


@routes_bp.route('/exportar_relatorio_pedidos_excel', methods=['GET'])
def exportar_relatorio_pedidos_excel():
    """Exporta o relatório de pedidos para Excel"""
    if 'usuario' not in session:
        return redirect(url_for('routes_bp.login'))
    
    try:
        # Parâmetros de filtro
        empresa = request.args.get('empresa')
        status = request.args.get('status')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        numero_pedido = request.args.get('numero_pedido')
        
        # Query base
        query = db.session.query(PedidosCompra)
        
        if status:
            query = query.filter(PedidosCompra.status == status)
        if numero_pedido:
            query = query.filter(PedidosCompra.numero_pedido.ilike(f'%{numero_pedido}%'))
        if data_inicio:
            query = query.filter(PedidosCompra.data_criacao >= data_inicio)
        if data_fim:
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(PedidosCompra.data_criacao < dt_fim)
        
        EXPORT_ROW_LIMIT = 5000
        pedidos = query.order_by(PedidosCompra.data_criacao.desc()).limit(EXPORT_ROW_LIMIT).all()
        if len(pedidos) == EXPORT_ROW_LIMIT:
            app.logger.warning(f'Export de pedidos atingiu limite de {EXPORT_ROW_LIMIT} linhas.')
        
        # Criar workbook
        wb = Workbook()
        
        # ============================================
        # PLANILHA 1: Resumo dos Pedidos
        # ============================================
        ws_resumo = wb.active
        ws_resumo.title = "Resumo dos Pedidos"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="022855", end_color="022855", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Cabeçalhos do resumo
        headers_resumo = [
            'Nº Pedido', 'Data Criação', 'Status', 'Usuário', 'Empresas', 
            'Compradores', 'Fornecedores', 'Qtd Itens', 'Valor Total (R$)', 
            'Valor Frete (R$)', 'Valor Líquido (R$)', 'Forma Pagamento', 'Observações',
            'Data Envio p/ Aprovação', 'Dias Aguardando Aprovação'
        ]
        
        for col_num, header in enumerate(headers_resumo, 1):
            cell = ws_resumo.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Dados do resumo
        row_num = 2
        for pedido in pedidos:
            # Aplicar filtro de empresa em memória
            empresas_set = set()
            compradores_set = set()
            fornecedores_set = set()
            valor_total = 0
            total_itens = 0
            
            for preenchimento in pedido.preenchimentos:
                solicitacao = preenchimento.solicitacao
                if solicitacao:
                    empresas_set.add(solicitacao.empresa)
                    if solicitacao.comprador_atribuido:
                        compradores_set.add(solicitacao.comprador_atribuido)
                    valor_total += (preenchimento.valor_total or 0)
                    total_itens += 1
                fornecedores_set.add(get_fornecedor_nome(preenchimento.fornecedor_id))
            
            # Pular se filtro de empresa não corresponder
            if empresa and empresa not in empresas_set:
                continue
            
            ws_resumo.cell(row=row_num, column=1, value=pedido.numero_pedido)
            ws_resumo.cell(row=row_num, column=2, value=pedido.data_criacao.strftime('%d/%m/%Y %H:%M') if pedido.data_criacao else '')
            ws_resumo.cell(row=row_num, column=3, value=pedido.status)
            ws_resumo.cell(row=row_num, column=4, value=pedido.usuario)
            ws_resumo.cell(row=row_num, column=5, value=', '.join(sorted(empresas_set)))
            ws_resumo.cell(row=row_num, column=6, value=', '.join(sorted(compradores_set)) if compradores_set else 'Não atribuído')
            ws_resumo.cell(row=row_num, column=7, value=', '.join(sorted(fornecedores_set)))
            ws_resumo.cell(row=row_num, column=8, value=total_itens)
            ws_resumo.cell(row=row_num, column=9, value=round(valor_total, 2))
            ws_resumo.cell(row=row_num, column=10, value=round(pedido.valor_frete or 0, 2))
            ws_resumo.cell(row=row_num, column=11, value=round(pedido.valor_liquido or 0, 2))
            ws_resumo.cell(row=row_num, column=12, value=pedido.forma_pagamento or 'N/A')
            ws_resumo.cell(row=row_num, column=13, value=pedido.observacoes or '')
            
            # Buscar data_envio_aprovacao das solicitações deste pedido
            datas_envio = [
                p.solicitacao.data_envio_aprovacao
                for p in pedido.preenchimentos
                if p.solicitacao and p.solicitacao.data_envio_aprovacao
            ]
            data_envio = min(datas_envio) if datas_envio else None
            data_pedido = pedido.data_criacao

            ws_resumo.cell(row=row_num, column=14, value=data_envio.strftime('%d/%m/%Y %H:%M') if data_envio else '')
            ws_resumo.cell(row=row_num, column=15, value=(data_pedido - data_envio).days if (data_envio and data_pedido and data_pedido >= data_envio) else '')
            row_num += 1
        
        # ============================================
        # PLANILHA 2: Detalhamento dos Itens
        # ============================================
        ws_detalhes = wb.create_sheet("Detalhamento dos Itens")
        
        headers_detalhes = [
            'Nº Pedido', 'Data Pedido', 'Status Pedido', 'ID Solicitação', 
            'Material', 'Cód Material', 'Especificação', 'Quantidade', 'Unidade',
            'Marca', 'Valor Unitário (R$)', 'Valor Total Item (R$)', 
            'Fornecedor', 'Prazo Entrega', 'Condição Pagamento', 'Empresa', 'Comprador',
            'Data Envio p/ Aprovação', 'Dias Aguardando Aprovação'
        ]
        
        for col_num, header in enumerate(headers_detalhes, 1):
            cell = ws_detalhes.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        row_num = 2
        for pedido in pedidos:
            for preenchimento in pedido.preenchimentos:
                solicitacao = preenchimento.solicitacao
                if solicitacao:
                    # Aplicar filtro de empresa
                    if empresa and solicitacao.empresa != empresa:
                        continue
                    
                    # Converter quantidade
                    quantidade = solicitacao.quantidade
                    if isinstance(quantidade, str):
                        try:
                            quantidade = float(quantidade.replace(',', '.'))
                        except:
                            quantidade = 0
                    
                    ws_detalhes.cell(row=row_num, column=1, value=pedido.numero_pedido)
                    ws_detalhes.cell(row=row_num, column=2, value=pedido.data_criacao.strftime('%d/%m/%Y %H:%M') if pedido.data_criacao else '')
                    ws_detalhes.cell(row=row_num, column=3, value=pedido.status)
                    ws_detalhes.cell(row=row_num, column=4, value=solicitacao.id)
                    ws_detalhes.cell(row=row_num, column=5, value=solicitacao.material.DescricaoMaterial if solicitacao.material else 'N/A')
                    ws_detalhes.cell(row=row_num, column=6, value=solicitacao.cod_material)
                    ws_detalhes.cell(row=row_num, column=7, value=solicitacao.especificacao or '')
                    ws_detalhes.cell(row=row_num, column=8, value=float(quantidade) if isinstance(quantidade, (int, float)) else 0)
                    ws_detalhes.cell(row=row_num, column=9, value=solicitacao.unidade_medida or 'Un')
                    ws_detalhes.cell(row=row_num, column=10, value=solicitacao.marca or '')
                    ws_detalhes.cell(row=row_num, column=11, value=round(preenchimento.valor_unitario or 0, 2))
                    ws_detalhes.cell(row=row_num, column=12, value=round(preenchimento.valor_total or 0, 2))
                    ws_detalhes.cell(row=row_num, column=13, value=get_fornecedor_nome(preenchimento.fornecedor_id))
                    ws_detalhes.cell(row=row_num, column=14, value=preenchimento.prazo_entrega or '')
                    ws_detalhes.cell(row=row_num, column=15, value=preenchimento.condicao_pagamento or '')
                    ws_detalhes.cell(row=row_num, column=16, value=solicitacao.empresa)
                    ws_detalhes.cell(row=row_num, column=17, value=solicitacao.comprador_atribuido or '')
                    
                    data_envio = solicitacao.data_envio_aprovacao
                    data_pedido = pedido.data_criacao
                    ws_detalhes.cell(row=row_num, column=18, value=data_envio.strftime('%d/%m/%Y %H:%M') if data_envio else '')
                    ws_detalhes.cell(row=row_num, column=19, value=(data_pedido - data_envio).days if (data_envio and data_pedido and data_pedido >= data_envio) else '')
                    row_num += 1
        
        # ============================================
        # PLANILHA 3: Estatísticas por Empresa
        # ============================================
        ws_estatisticas = wb.create_sheet("Estatísticas por Empresa")
        
        headers_estatisticas = [
            'Empresa', 'Total Pedidos', 'Total Itens', 'Valor Total (R$)', 
            'Valor Médio por Pedido (R$)', 'Status Mais Frequente'
        ]
        
        for col_num, header in enumerate(headers_estatisticas, 1):
            cell = ws_estatisticas.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Agrupar dados por empresa
        empresas_stats = {}
        for pedido in pedidos:
            for preenchimento in pedido.preenchimentos:
                solicitacao = preenchimento.solicitacao
                if solicitacao:
                    empresa_nome = solicitacao.empresa
                    if empresa_nome not in empresas_stats:
                        empresas_stats[empresa_nome] = {
                            'total_pedidos': 0,
                            'total_itens': 0,
                            'valor_total': 0,
                            'status_count': {}
                        }
                    
                    empresas_stats[empresa_nome]['total_pedidos'] += 1
                    empresas_stats[empresa_nome]['total_itens'] += 1
                    empresas_stats[empresa_nome]['valor_total'] += (preenchimento.valor_total or 0)
                    
                    if pedido.status not in empresas_stats[empresa_nome]['status_count']:
                        empresas_stats[empresa_nome]['status_count'][pedido.status] = 0
                    empresas_stats[empresa_nome]['status_count'][pedido.status] += 1
        
        row_num = 2
        for empresa_nome, stats in empresas_stats.items():
            # Encontrar status mais frequente
            status_mais_frequente = max(stats['status_count'].items(), key=lambda x: x[1])[0] if stats['status_count'] else 'N/A'
            
            ws_estatisticas.cell(row=row_num, column=1, value=empresa_nome)
            ws_estatisticas.cell(row=row_num, column=2, value=stats['total_pedidos'])
            ws_estatisticas.cell(row=row_num, column=3, value=stats['total_itens'])
            ws_estatisticas.cell(row=row_num, column=4, value=round(stats['valor_total'], 2))
            ws_estatisticas.cell(row=row_num, column=5, value=round(stats['valor_total'] / stats['total_pedidos'] if stats['total_pedidos'] > 0 else 0, 2))
            ws_estatisticas.cell(row=row_num, column=6, value=status_mais_frequente)
            row_num += 1
        
        # Ajustar largura das colunas para todas as planilhas
        for ws in [ws_resumo, ws_detalhes, ws_estatisticas]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar arquivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nome do arquivo com timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"relatorio_pedidos_{timestamp}.xlsx"
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        from urllib.parse import quote as _url_quote
        response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{_url_quote(filename)}"
        
        return response
        
    except Exception as e:
        _flash_erro_seguro('Erro ao exportar relatório.', e, app.logger)
        app.logger.error(f'Erro em exportar_relatorio_pedidos_excel: {str(e)}', exc_info=True)
        return redirect(url_for('routes_bp.relatorio_pedidos'))
    
@routes_bp.route('/api/verificar_rascunhos', methods=['GET'])
def verificar_rascunhos():
    """API para verificar quais solicitações têm rascunhos salvos"""
    if 'usuario' not in session:
        return jsonify({'error': 'Não autenticado'}), 401
    
    try:
        # Buscar todas as solicitações que têm pelo menos um rascunho
        # Agrupar por aplicacao + data_hora (mesma lógica do grupo)
        rascunhos_por_grupo = {}
        
        # Buscar todos os rascunhos
        rascunhos = db.session.query(
            SolicitacoesPreenchidas.solicitacao_id,
            SolicitacoesCompra.aplicacao,
            SolicitacoesCompra.data_solicitacao
        ).join(
            SolicitacoesCompra,
            SolicitacoesPreenchidas.solicitacao_id == SolicitacoesCompra.id
        ).filter(
            SolicitacoesPreenchidas.status == 'Rascunho'
        ).all()
        
        for rascunho in rascunhos:
            # Criar chave do grupo (mesma lógica do template)
            aplic = rascunho.aplicacao or 'Sem Aplicação'
            data_hora_str = rascunho.data_solicitacao.strftime('%Y-%m-%d %H:%M')
            chave_grupo = f"{aplic}|{data_hora_str}"
            
            if chave_grupo not in rascunhos_por_grupo:
                rascunhos_por_grupo[chave_grupo] = {
                    'tem_rascunho': True,
                    'quantidade_rascunhos': 0
                }
            rascunhos_por_grupo[chave_grupo]['quantidade_rascunhos'] += 1
        
        return jsonify({
            'success': True,
            'rascunhos': rascunhos_por_grupo
        })
        
    except Exception as e:
        logging.error(f"Erro ao verificar rascunhos: {str(e)}")
        app.logger.error(f'API error: {type(e).__name__}', exc_info=False)
        return jsonify({'success': False, 'error': 'Erro interno. Tente novamente.'}), 500
    
  # Registro do Blueprint (apenas uma vez)
app.register_blueprint(routes_bp)  

# ============================================
# INICIALIZAÇÃO DO BANCO — roda sempre,
# tanto via `gunicorn app:app` quanto `python app.py`
# ============================================
executar_todas_migracoes()

create_database()
create_auditoria_table()
create_fornecedores_db()
create_historico_descontos_table()

with app.app_context():
    db.create_all()

# Garantir colunas APÓS create_all (adiciona colunas em tabelas já existentes)
garantir_colunas_necessarias()

# Execução local (desenvolvimento)
if __name__ == '__main__':
    print("\n🌐 Iniciando servidor Flask...")
    app.run(debug=False, host='0.0.0.0', port=5000, threaded=True, use_reloader=False)
    